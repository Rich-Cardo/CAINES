#Librerías generales
import os

import locale
locale.setlocale(locale.LC_TIME, 'es_ES.utf8')

from flask import Flask
from flask import render_template,request,redirect, flash, make_response, abort
from flaskext.mysql import MySQL

#Librerías para la generación de PDF

import jinja2
import pdfkit
import json

#Librería para enviar solicitudes al servidor
from flask import  jsonify

#Librería para servidor de producción
from waitress import serve

#Librería para el manejo de sesiones activas de los usuarios
from flask_login import UserMixin, login_user, LoginManager, login_required, logout_user, current_user

import pandas as pd

#Código general de Flask

app = Flask(__name__,static_folder='static')
app.secret_key="Caines"

mysql = MySQL()

app.config['MYSQL_DATABASE_HOST'] = 'localhost'
app.config['MYSQL_DATABASE_USER'] = 'root'
app.config['MYSQL_DATABASE_PASSWORD'] = ''
app.config['MYSQL_DATABASE_DB']='compras'


mysql.init_app(app)


#Codigo para gestionar el login con Flask_Login

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'


#Clase Usuario para el manejo de sesiones activas

class User(UserMixin):
    def __init__(self, id, cedula, nombre, rol):
        self.id = id
        self.cedula = cedula
        self.nombre = nombre
        self.rol = rol

    @staticmethod
    def get(id):
        conn = mysql.connect()
        cursor = conn.cursor()
        sql = "SELECT * FROM usuarios WHERE id_usuario = %s"
        cursor.execute(sql, (id,))
        user_data = cursor.fetchone()
        if user_data:
            return User(id=user_data[0], rol=user_data[1], cedula=user_data[2], nombre=user_data[3])
        return None

    def is_authenticated(self):
        return True

    def is_active(self):
        return True

    def is_anonymous(self):
        return False

class IDGenerator:
    def __init__(self, conn):
        self.conn = conn

    def _generar_id_base(self, tabla, campo_id, prefijo, filtro_ubicacion):
        """
        Método base para generación de IDs
        :param tabla: Nombre de la tabla a consultar
        :param campo_id: Nombre del campo ID a consultar
        :param prefijo: Prefijo para el ID
        :param filtro_ubicacion: Tupla con (campo, valor) para filtro de ubicación
        """

        # Validación de los campos para agregar seguridad en la base de datos
        campos_permitidos = {'id_salida', 'id_entrada', 'id_articulo'}
        if campo_id not in campos_permitidos:
            raise ValueError(f"Campo ID no válido: {campo_id}")
        
        # Inicio de la lógica de generación de IDs
        cursor = self.conn.cursor()
        query = f"SELECT {campo_id} FROM {tabla}"
        params = None
        
        if filtro_ubicacion:
            query += f" WHERE ubicacion = %s"
            params = (filtro_ubicacion[1],)
        
        query += f" ORDER BY {campo_id} DESC LIMIT 1"

        cursor.execute(query, params)
        last_id = cursor.fetchone()
        cursor.close()

        last_number = int(last_id[0].split('-')[1]) if last_id else 0
        return f"{prefijo}-{(last_number + 1):04d}"

    def generar_id_salida(self, ubicacion):
        """Genera ID de salida según ubicación"""
        prefijos = {
            'oficina': 'SOF',
            'galpon': 'SGA'
        }
        return self._generar_id_base(
            tabla='salidas_inventario',
            campo_id='id_salida',
            prefijo=prefijos[ubicacion],
            filtro_ubicacion=('ubicacion', ubicacion)
        )

    def generar_id_articulo(self, ubicacion):
        """Genera ID de artículo según ubicación"""
        prefijos = {
            'oficina': 'OFC',
            'galpon': 'ART'
        }
        return self._generar_id_base(
            tabla='articulos',
            campo_id='id_articulo',
            prefijo=prefijos[ubicacion],
            filtro_ubicacion=('ubicacion', ubicacion)
        )
    
    def generar_id_entrada(self, ubicacion):
        """Genera ID de entrada según ubicación"""
        prefijos = {
            'oficina': 'EOF',
            'galpon': 'EN'
        }
        return self._generar_id_base(
            tabla='entradas_inventario',
            campo_id='id_entrada',
            prefijo=prefijos[ubicacion],
            filtro_ubicacion=('ubicacion', ubicacion)
        )

class InventarioManager:
    def __init__(self, conn):
        self.conn = conn
        self.cursor = conn.cursor()

    def commit(self):
        """Confirma los cambios en la base de datos."""
        self.conn.commit()

    def rollback(self):
        """Revierte los cambios en caso de error."""
        self.conn.rollback()

    def cerrar_conexion(self):
        """Cierra la conexión con la base de datos."""
        self.cursor.close()
        self.conn.close()

    ### Módulo de Gestión de Artículos ###

    def _insertar_nuevo_articulo(self, id_articulo, nombre, unidad, area, ubicacion):
        """Inserta un nuevo artículo en la base de datos"""

        query = """
            INSERT INTO articulos (id_articulo, articulo, unidad, area, ubicacion)
            VALUES (%s, %s, %s, %s, %s)
        """
        self.cursor.execute(query, (id_articulo, nombre, unidad, area, ubicacion))

    ### Módulo de Adición de Salidas ###
    def agregar_salida(self, fecha, destino, ubicacion, items):
        """Registra una nueva salida y sus artículos asociados"""
        # self._validate_location(ubicacion)
        
        try:
            # Usar IDGenerator para generar el ID
            id_generator = IDGenerator(self.conn)
            id_salida = id_generator.generar_id_salida(ubicacion)
            
            self._insertar_cabecera_salida(id_salida, fecha, ubicacion, destino)
            self._procesar_articulos_salida(id_salida, items, ubicacion)
            
            return id_salida
            
        except Exception as e:
            self.rollback()
            raise (f"Error registrando salida: {str(e)}") from e

    def _insertar_cabecera_salida(self, id_salida, fecha, ubicacion, destino):
        """Registra la cabecera de la salida"""
        self.cursor.execute(
            "INSERT INTO salidas_inventario (id_salida, fecha, ubicacion, destino) "
            "VALUES (%s, %s, %s, %s)",
            (id_salida, fecha, ubicacion, destino)
        )

    def _procesar_articulos_salida(self, id_salida, items, ubicacion):
        """Registra todos los artículos de la salida"""
        for item in items:

            #Primero hay que obtener el ID de los artículos
            id_articulo = self._obtener_id_articulo_existente(
                item['articulo'],
                item['unidad'],
                item['area'], 
                ubicacion)
            
            # Luego se procede a registrar en aux_salidas_inventario
            self._registrar_detalle_salida(id_salida, id_articulo, item['cantidad'])
            
    # def _registrar_articulo(self, nombre, unidad, area, ubicacion):
    #     """Ojo, esta función ya no es necesaria; se dividió su lógica del UPSERT."""
    #     """Registra o recupera un artículo existente (optimizado)"""
    #     id_articulo = self._obtener_id_articulo_existente(nombre, unidad, area, ubicacion)
        
    #     if not id_articulo:
    #         id_articulo = IDGenerator(self.conn).generar_id_articulo(ubicacion)
    #         self._insertar_nuevo_articulo(id_articulo, nombre, unidad, area, ubicacion)
        
    #     return id_articulo
    
    def _obtener_id_articulo_existente(self, nombre, unidad, area, ubicacion):
        """Busca un artículo existente basado en sus características"""
        query = """
            SELECT id_articulo 
            FROM articulos 
            WHERE 
                articulo = %s AND
                unidad = %s AND
                area = %s AND
                ubicacion = %s
            LIMIT 1
        """
        self.cursor.execute(query, (nombre, unidad, area, ubicacion))
        result = self.cursor.fetchone()

        if result:
            return result[0]
        else:
            return None
    
    def _registrar_detalle_salida(self, id_salida, id_articulo, cantidad):
        """Inserta un artículo en aux_salidas_inventario"""

        self.cursor.execute(
            "INSERT INTO aux_salidas_inventario (id_salida, id_articulo, cantidad) "
            "VALUES (%s, %s, %s)",
            (id_salida, id_articulo, cantidad)
        )

    def registrar_articulo_extra(self, id_salida, articulo, cantidad, unidad, area):
        """Registra un artículo extra en la salida"""

        self.cursor.execute(
            "INSERT INTO articulos_extra_salida "
            "(id_salida, articulo, cantidad, unidad, area) "
            "VALUES (%s, %s, %s, %s, %s)",
            (id_salida, articulo, cantidad, unidad, area)
        )

    ### Módulo de Edición de Salidas ###

    def actualizar_salida(self, id_salida, fecha, destino):
        """Actualiza la cabecera de una salida existente"""
        self.cursor.execute(
            "UPDATE salidas_inventario SET fecha = %s, destino = %s "
            "WHERE id_salida = %s",
            (fecha, destino, id_salida)
        )

    def _actualizar_detalle_salida(self, id_salida, id_articulo, nueva_cantidad):
        """Actualiza la cantidad de un artículo existente"""
        self.cursor.execute(
            "UPDATE aux_salidas_inventario SET cantidad = %s WHERE id_salida = %s AND id_articulo = %s",
            (nueva_cantidad, id_salida, id_articulo)
        )
    
    # def registrar_articulos_extra(self, id_salida, articulos_extra):
    #     """Registra múltiples artículos extra en una salida"""
    #     for articulo in articulos_extra:
    #         self.registrar_articulo_extra(
    #             id_salida,
    #             articulo['articulo'],
    #             articulo['cantidad'],
    #             articulo['unidad'],
    #             articulo['area']
    #         )

    # def registrar_articulos_extra(self, id_salida, articulos):
    #     """Registra uno o múltiples artículos extra usando ejecución masiva (executemany)"""
    #     # Normalizar entrada: convertir artículo único a lista de un elemento
    #     if not isinstance(articulos, list):
    #         articulos = [articulos]
        
    #     # Construir datos para la query
    #     valores = [
    #         (
    #             id_salida,
    #             art["articulo"],
    #             art["cantidad"],
    #             art["unidad"],
    #             art["area"]
    #         ) 
    #         for art in articulos
    #     ]
        
    #     # Ejecutar inserción masiva (1 sola operación en BD)
    #     self.cursor.executemany(
    #         "INSERT INTO articulos_extra_salida "
    #         "(id_salida, articulo, cantidad, unidad, area) "
    #         "VALUES (%s, %s, %s, %s, %s)",
    #         valores
    #     )

    ### Módulo de Eliminación ###

    def _eliminar_articulo_salida(self, id_salida, id_articulo):
        """Elimina sólo un (1) artículo específico de la salida"""
        self.cursor.execute(
            "DELETE FROM aux_salidas_inventario WHERE id_salida = %s AND id_articulo = %s",
            (id_salida, id_articulo)
        )

    def eliminar_articulos_salida(self, id_salida):
        """Elimina TODOS los artículos (incluido los extras) asociados a una salida"""
        self.cursor.execute(
            "DELETE FROM aux_salidas_inventario WHERE id_salida = %s",
            (id_salida,)
        )
        self.cursor.execute(
            "DELETE FROM articulos_extra_salida WHERE id_salida = %s",
            (id_salida,)
        )

    def eliminar_articulos_viejos(self, id_entrada, existing_articles, new_articles):
        """
        Elimina los artículos que ya no están en la lista y actualiza el stock.
        """
        articles_to_remove = set(existing_articles.keys()) - new_articles
        for art_id in articles_to_remove:
            old_cantidad = existing_articles[art_id]

            # Reducir stock y eliminar artículo de la entrada
            # self.cursor.execute("UPDATE stock_inventario SET cantidad = cantidad - %s WHERE id_articulo = %s", (old_cantidad, art_id))
            self.cursor.execute("DELETE FROM aux_entradas_inventario WHERE id_entrada = %s AND id_articulo = %s", (id_entrada, art_id))

    ### Módulo de Entradas ###

    ### Agregar Entradas ###
    def agregar_entrada(self, fecha, ubicacion, items):
        """Registra una nueva entrada y sus artículos asociados"""
        try:
            id_generator = IDGenerator(self.conn)
            id_entrada = id_generator.generar_id_entrada(ubicacion)
            
            self._insertar_cabecera_entrada(id_entrada, fecha, ubicacion)
            self._procesar_articulos_entrada(id_entrada, items, ubicacion)
            
            self.actualizar_stock(ubicacion)
            return id_entrada
            
        except Exception as e:
            self.rollback()
            raise (f"Error registrando entrada: {str(e)}") from e

    def _insertar_cabecera_entrada(self, id_entrada, fecha, ubicacion):
        """Registra la cabecera de la entrada"""
        self.cursor.execute(
            "INSERT INTO entradas_inventario (id_entrada, fecha, ubicacion) "
            "VALUES (%s, %s, %s)",
            (id_entrada, fecha, ubicacion)
        )

    def _procesar_articulos_entrada(self, id_entrada, items, ubicacion):
        """Registra todos los artículos de la entrada"""
        for item in items:
            id_articulo = self._registrar_articulo(
                item['articulo'],
                item['unidad'],
                item['area'],
                ubicacion
            )
            self._registrar_detalle_entrada(id_entrada, id_articulo, item['cantidad'])

    def _registrar_detalle_entrada(self, id_entrada, id_articulo, cantidad):
        """Inserta un artículo en aux_entradas_inventario"""
        self.cursor.execute(
            "INSERT INTO aux_entradas_inventario (id_entrada, id_articulo, cantidad) "
            "VALUES (%s, %s, %s)",
            (id_entrada, id_articulo, cantidad)
        )

    ### Editar Entradas ###

    def editar_entrada(self, id_entrada, fecha, articulos, cantidades, unidades, areas, ubicacion):
        """
        Edita una entrada existente en la tabla `entradas_inventario` y actualiza el stock.
        """
        # Actualizar la fecha de la entrada
        self.cursor.execute("UPDATE entradas_inventario SET fecha = %s WHERE id_entrada = %s", (fecha, id_entrada))

        # Obtener artículos existentes en la entrada
        existing_entries = self.obtener_articulos_en_entrada(id_entrada)
        existing_articles = {row[0]: row[1] for row in existing_entries}

        # Artículos nuevos que serán procesados
        new_articles = set()

        # Actualizar o insertar artículos nuevos
        for articulo, cantidad, unidad, area in zip(articulos, cantidades, unidades, areas):
            id_articulo = self.agregar_articulo(articulo, unidad, area, ubicacion)
            new_articles.add(id_articulo)
            cantidad = int(cantidad)

            if id_articulo in existing_articles:
                # Artículo ya existe en la entrada, actualizar cantidad
                old_cantidad = existing_articles[id_articulo]
                diferencia = cantidad - old_cantidad

                # Actualizar cantidad en `aux_entradas_inventario` y stock
                self.actualizar_entrada(id_entrada, id_articulo, cantidad, diferencia)
            else:
                # Nuevo artículo, insertar en `aux_entradas_inventario` y actualizar stock
                self.insertar_nuevo_articulo(id_entrada, id_articulo, cantidad)

        # Eliminar artículos que ya no están en la lista
        self.eliminar_articulos_viejos(id_entrada, existing_articles, new_articles)

    def actualizar_entrada(self, id_entrada, id_articulo, cantidad, diferencia):
        """
        Actualiza la cantidad de un artículo en la entrada y el stock.
        """
        self.cursor.execute("UPDATE aux_entradas_inventario SET cantidad = %s WHERE id_entrada = %s AND id_articulo = %s", (cantidad, id_entrada, id_articulo))
        # self.cursor.execute("UPDATE stock_inventario SET cantidad = cantidad + %s WHERE id_articulo = %s", (diferencia, id_articulo))

    def insertar_nuevo_articulo(self, id_entrada, id_articulo, cantidad):
        """
        Inserta un nuevo artículo en la entrada y actualiza el stock.
        """
        self.cursor.execute("INSERT INTO aux_entradas_inventario (id_entrada, id_articulo, cantidad) VALUES (%s, %s, %s)", (id_entrada, id_articulo, cantidad))
        # self.cursor.execute("UPDATE stock_inventario SET cantidad = cantidad + %s WHERE id_articulo = %s", (cantidad, id_articulo))

    ### Módulo de Consultas ###

    def obtener_totales(self, tabla_aux):
        """
        Obtiene los totales de entradas o salidas agrupados por artículo.
        :param tabla_aux: Nombre de la tabla (aux_entradas_inventario o aux_salidas_inventario).
        Este método se utiliza en el FrontEnd de las Entradas, Inventario y Salidas.
        """
        query = f"""
            SELECT id_articulo, SUM(cantidad) AS total
            FROM {tabla_aux}
            GROUP BY id_articulo
        """
        self.cursor.execute(query)
        
        # Creamos un diccionario vacío para almacenar los resultados
        totales = {}
        
        # Iteramos sobre cada fila del resultado
        for row in self.cursor.fetchall():
            id_articulo = row[0]  # Primera columna: id_articulo
            total = row[1]        # Segunda columna: suma total
            totales[id_articulo] = total
        
        return totales
    
    #Método para poner el stock al día (mover de este módulo en un futuro)
    def actualizar_stock(self, ubicacion):
        """
        Actualiza el stock de todos los artículos para una ubicación específica,
        calculando el balance entre entradas y salidas registradas.
        
        Parámetros:
            ubicacion (str): Ubicación para la cual se actualizará el stock
        """
        with self.conn.cursor() as cursor:
            # 1. Obtener todos los artículos de la ubicación
            articulos = self._obtener_articulos_por_ubicacion(cursor, ubicacion)
            
            for id_articulo in articulos:
                # 2. Calcular movimientos totales
                total_entradas = self._calcular_total_movimientos(
                    cursor, 
                    'aux_entradas_inventario', 
                    id_articulo
                )
                total_salidas = self._calcular_total_movimientos(
                    cursor, 
                    'aux_salidas_inventario', 
                    id_articulo
                )
                
                # 3. Calcular y actualizar stock
                nuevo_stock = total_entradas - total_salidas
                self._actualizar_registro_stock(
                    cursor, 
                    id_articulo, 
                    nuevo_stock, 
                    ubicacion
                )
            
            self.conn.commit()
    
    #Métodos auxiliares de Actualizar_Stock    
    def _obtener_articulos_por_ubicacion(self, cursor, ubicacion):
        """Obtiene todos los IDs de artículos para una ubicación"""
        cursor.execute(
            "SELECT id_articulo FROM articulos WHERE ubicacion = %s",
            (ubicacion,)
        )
        return [row[0] for row in cursor.fetchall()]
    
    def _calcular_total_movimientos(self, cursor, tabla_movimientos, id_articulo):
        """Calcula el total acumulado de movimientos para un artículo"""
        query = f"""
            SELECT COALESCE(SUM(cantidad), 0)
            FROM {tabla_movimientos}
            WHERE id_articulo = %s
        """
        cursor.execute(query, (id_articulo,))
        return cursor.fetchone()[0]
    
    def _actualizar_registro_stock(self, cursor, id_articulo, cantidad, ubicacion):
        """Realiza el UPSERT del stock en la base de datos
        Se prefiere usar UPSERT ya que se puede reutilizar cuando se está agregando
        un nuevo artículo al inventario, sin crear otras funciones."""

        query = """
            INSERT INTO stock_inventario (id_articulo, cantidad, ubicacion)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE cantidad = VALUES(cantidad)
        """
        cursor.execute(query, (id_articulo, cantidad, ubicacion))
    
    ####

    def obtener_inventario(self, ubicacion):
        """Obtiene todos los artículos en stock con su cantidad y detalles para una ubicación específica."""
        self.cursor.execute('''
            SELECT
                stock.id_articulo,
                articulos.articulo,
                stock.cantidad,
                articulos.unidad,
                articulos.area
            FROM
                stock_inventario stock
            JOIN
                articulos ON stock.id_articulo = articulos.id_articulo
            WHERE
                stock.ubicacion = %s
            ORDER BY
                articulos.articulo ASC
        ''', (ubicacion,))
        return self.cursor.fetchall()
    
    def obtener_articulos_extra_por_salida(self, id_salida):
        """Obtiene los artículos extra para una salida dada"""
        try:
            cursor = self.conn.cursor()
            query = """
                SELECT articulo, cantidad, unidad, area
                FROM articulos_extra_salida
                WHERE id_salida = %s
            """
            cursor.execute(query, (id_salida,))
            articulos_extra = cursor.fetchall()
            return articulos_extra
        except Exception as e:
            self.conn.rollback()
            raise e
        
    def obtener_articulos_en_entrada(self, id_entrada):
        """
        Obtiene los artículos y cantidades de la entrada.
        """
        self.cursor.execute("SELECT id_articulo, cantidad FROM aux_entradas_inventario WHERE id_entrada = %s", (id_entrada,))
        return self.cursor.fetchall()

    def obtener_total_cantidad(self, tipo, ubicacion):
        """
        Obtiene la sumatoria total de las cantidades de artículos según el tipo de movimiento y la ubicación.
        :param tipo: Tipo de movimiento ('entrada', 'salida', 'inventario').
        :param ubicacion: Ubicación ('galpon' o 'palacio').
        """
        if tipo == 'inventario':
            query = '''
                    SELECT
                        SUM(stock.cantidad) AS total_cantidad
                    FROM
                        stock_inventario stock
                    WHERE
                        stock.cantidad > 0 AND stock.ubicacion = %s
            '''

        elif tipo == 'entradas':
            # Para entradas, se consulta la tabla `aux_entradas_inventario`.
            query = '''
                SELECT SUM(ae.cantidad) AS total_cantidad
                FROM aux_entradas_inventario ae
                JOIN entradas_inventario ei ON ae.id_entrada = ei.id_entrada
                JOIN articulos a ON ae.id_articulo = a.id_articulo
                WHERE ei.ubicacion = %s
            '''
        elif tipo == 'salidas':
            # Para salidas, se consulta la tabla `aux_salidas_inventario`.
            query = '''
                SELECT SUM(asl.cantidad) AS total_cantidad
                FROM aux_salidas_inventario asl
                JOIN salidas_inventario si ON asl.id_salida = si.id_salida
                JOIN articulos a ON asl.id_articulo = a.id_articulo
                WHERE si.ubicacion = %s
            '''
        else:
            raise ValueError("Tipo de movimiento no válido. Debe ser 'entrada', 'salida' o 'inventario'.")

        self.cursor.execute(query, (ubicacion,))
        result = self.cursor.fetchone()
        return result[0] if result else 0

    def obtener_articulos_con_stock(self, ubicacion):
        """
        Obtiene los artículos presentes y su cantidad total en stock según la ubicación.
        :param ubicacion: Ubicación del stock ('galpon' o 'oficina').
        """
        query = '''
            SELECT
                articulos.id_articulo,
                articulos.articulo,
                articulos.unidad,
                articulos.area,
                IFNULL(SUM(stock_inventario.cantidad), 0) AS total_cantidad
            FROM
                articulos articulos
            LEFT JOIN
                stock_inventario stock_inventario ON articulos.id_articulo = stock_inventario.id_articulo
            WHERE
                stock_inventario.ubicacion = %s
            GROUP BY
                articulos.id_articulo, articulos.articulo, articulos.unidad, articulos.area
            ORDER BY
                articulos.articulo
        '''

        # Ejecutar la consulta parametrizada
        self.cursor.execute(query, (ubicacion,))

        # Obtener los resultados
        resultados = self.cursor.fetchall()

        # Devolver los resultados
        return resultados

    def obtener_entradas_o_salidas(self, tipo, ubicacion):
        """
        Obtiene las entradas o salidas de artículos según el tipo y la ubicación.
        :param tipo: 'entradas' o 'salidas' para saber qué consultar.
        :param ubicacion: 'galpon' o 'palacio' para filtrar por ubicación.
        :return: Lista de entradas o salidas formateadas.
        """
        # Seleccionar la tabla y el campo dinámico según el tipo
        if tipo == "entradas":
            tabla = "entradas_inventario"
            tabla_aux = "aux_entradas_inventario"
            campo_id = "id_entrada"  # Campo dinámico para entradas
            extra_campos = ""
        elif tipo == "salidas":
            tabla = "salidas_inventario"
            tabla_aux = "aux_salidas_inventario"
            campo_id = "id_salida"  # Campo dinámico para salidas
            extra_campos = f", {tabla}.destino"
        else:
            raise ValueError("El tipo debe ser 'entradas' o 'salidas'")

        # Construcción de la consulta principal
        query = f'''
            SELECT
                {tabla_aux}.id,
                {tabla}.fecha,
                articulos.articulo,
                {tabla_aux}.cantidad,
                articulos.unidad,
                articulos.area,
                {tabla}.{campo_id} {extra_campos}
            FROM
                {tabla}
            LEFT JOIN
                {tabla_aux} ON {tabla}.{campo_id} = {tabla_aux}.{campo_id}
            LEFT JOIN
                articulos ON {tabla_aux}.id_articulo = articulos.id_articulo
            WHERE
                {tabla}.ubicacion = %s
            ORDER BY
                {tabla}.fecha DESC, {tabla}.{campo_id} DESC
        '''

        # Ejecución de la consulta principal
        self.cursor.execute(query, (ubicacion,))
        resultados = self.cursor.fetchall()

        if tipo == "salidas":
            # Construcción de la consulta para artículos extra en salidas
            query_extras = f'''
                SELECT
                    id_salida,
                    articulo,
                    cantidad,
                    unidad,
                    area
                FROM
                    articulos_extra_salida
                WHERE
                    id_salida IN (SELECT DISTINCT {campo_id} FROM {tabla} WHERE ubicacion = %s)
            '''

            # Ejecución de la consulta de artículos extra
            self.cursor.execute(query_extras, (ubicacion,))
            resultados_extras = self.cursor.fetchall()

            # Formatear los resultados extras
            articulos_extra = {}
            for extra in resultados_extras:
                id_salida, articulo, cantidad, unidad, area = extra
                if id_salida not in articulos_extra:
                    articulos_extra[id_salida] = []
                articulos_extra[id_salida].append((articulo, cantidad, unidad, area))

        # Formatear los resultados de la consulta principal
        resultados_formateados = []
        for resultado in resultados:
            id, fecha, articulo, cantidad, unidad, area, campo_id_valor, *destino = resultado

            # Formatear la fecha si es datetime
            if hasattr(fecha, "strftime"):
                fecha = fecha.strftime('%d/%m/%Y')

            # Agregar destino al resultado si aplica
            if tipo == "salidas":
                destino = destino[0] if destino else None
                salida_formateada = [id, fecha, articulo, cantidad, unidad, area, campo_id_valor, destino]
                if campo_id_valor in articulos_extra:
                    salida_formateada.append(articulos_extra[campo_id_valor])
                resultados_formateados.append(salida_formateada)
            else:
                resultados_formateados.append((id, fecha, articulo, cantidad, unidad, area, campo_id_valor))

        # print("\n\nRESULTADOS FORMATEADOS ES: \n\n",  resultados_formateados, "\n\n")

        return resultados_formateados
    
    #Metodos Antiguos
    def agregar_articulo(self, articulo, unidad, area, ubicacion):
        """
        Verifica si un artículo existe, si no, lo agrega.
        Devuelve el ID del artículo.
        """
        self.cursor.execute("SELECT id_articulo FROM articulos WHERE articulo = %s AND unidad = %s AND ubicacion = %s", (articulo, unidad, ubicacion))
        result = self.cursor.fetchone()

        if result:  # El artículo ya existe en la base de datos
            return result[0] #retorna el ID del artículo.
        
        else:  # El artículo no existe, por lo tanto agregarlo.
            id_articulo = IDGenerator.generar_id_articulo(self, ubicacion)
            self.cursor.execute(
                "INSERT INTO articulos (id_articulo, articulo, unidad, area, ubicacion) VALUES (%s, %s, %s, %s, %s)",
                (id_articulo, articulo, unidad, area, ubicacion)
            )

            return id_articulo #Retorna el ID del nuevo articulo.


@login_manager.user_loader
def load_user(user_id):
    return User.get(user_id)

@app.errorhandler(404)
def page_not_found(e):
    return render_template('general/notfound.html'), 404

@app.route('/entrar', methods=['POST'])
def entrar():
    cedula = request.form['txtCedula']
    contrasena = request.form['txtContrasena']

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("SELECT id_usuario FROM usuarios WHERE cedula=%s AND clave=%s", (cedula, contrasena))
    user_id = cursor.fetchone()

    if user_id:
        user = User.get(user_id[0])
        login_user(user)  # Funcion de Flask_Login para Iniciar sesión del usuario actual
        return render_template('general/principal.html')
    else:
        flash('Cédula o contraseña incorrecta.')
        return render_template('general/login.html')

#Funciones para acceder a las rutas de la página

@app.route('/cerrar_sesion', methods=['GET'])
@login_required
def cerrar_sesion():

    logout_user()  # Función de Flask-Login para cerrar sesión
    flash('¡Has cerrado sesión exitosamente!')  # Mostramos un mensaje personalizado

    return redirect('/login')  # Redirigir al inicio de sesión

@app.route('/menu')
@login_required
def menu():
    return render_template('/general/menu.html')

@app.route('/medicinas/inventario')
@login_required
def medicinas_inventario():

    conn = mysql.connect()
    cursor = conn.cursor()

    # Consulta para obtener las medicinas junto con la cantidad disponible
    cursor.execute('''
        SELECT
            med.id_medicina,
            med.descripcion,
            med.marca,
            med.dosis,
            med.unidad,
            IFNULL(stock.cantidad, 0) AS cantidad
        FROM
            medicinas med
        LEFT JOIN
            stock_medicinas stock ON med.id_medicina = stock.id_medicina
        ORDER BY
            med.descripcion ASC
    ''')

    inventario = cursor.fetchall()

    # Obtener la sumatoria total de todas las cantidades en stock
    cursor.execute('''
        SELECT
            SUM(cantidad) AS total_cantidad
        FROM
            stock_medicinas
        WHERE
            cantidad > 0
    ''')

    # Para obtener el resultado
    total_cantidad = cursor.fetchone()[0]
    # print(f"La sumatoria de las cantidades es: {total_cantidad}")

    return render_template('/medicinas/inventario.html', inventario=inventario, total = total_cantidad)

@app.route('/medicinas/salidas')
@login_required
def medicinas_salidas():
    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT
            aux_salidas_medicinas.id,
            DATE_FORMAT(salidas_medicinas.fecha, '%d/%m/%Y') as fecha,
            medicinas.descripcion,
            aux_salidas_medicinas.cantidad,
            medicinas.marca,
            medicinas.dosis,
            medicinas.unidad,
            salidas_medicinas.id_salida,
            salidas_medicinas.destino  -- Aquí agregamos la columna destino
        FROM
            salidas_medicinas
        JOIN
            aux_salidas_medicinas ON salidas_medicinas.id_salida = aux_salidas_medicinas.id_salida
        JOIN
            medicinas ON aux_salidas_medicinas.id_medicina = medicinas.id_medicina
        ORDER BY
            salidas_medicinas.id_salida ASC, salidas_medicinas.fecha ASC
    ''')

    salidas = cursor.fetchall()

    # Agrupar las salidas por ID Salida
    grouped_salidas = []
    current_id_salida = None
    group = []

    for salida in salidas:
        if salida[7] != current_id_salida:  # Index 7 corresponde a id_salida
            if group:
                grouped_salidas.append(group)
            group = [salida]
            current_id_salida = salida[7]
        else:
            group.append(salida)

    if group:
        grouped_salidas.append(group)

    # Calcular la cantidad total de artículos salidos
    cursor.execute('''
        SELECT
            SUM(aux_sal.cantidad) as cantidad_total
        FROM
            aux_salidas_medicinas aux_sal
    ''')
    cantidad_total = cursor.fetchone()[0]

    cursor.close()
    conn.close()

    return render_template('/medicinas/salidas.html', grouped_salidas=grouped_salidas, cantidad_total=cantidad_total)

@app.route('/recuperar_contrasena')
def recuperar_contrasena():
    return render_template('/general/contrasena.html')

@app.route('/crear_parroquia', methods=['POST'])
def crear_parroquia():

    _parroquia = request.form['txtParroquia']
    _cod_municipio = request.form['txtMunicipio']

    conn = mysql.connect()
    cursor = conn.cursor()

    # Consultar si el parroquia ya existe en la base de datos

    sql = "SELECT * FROM parroquias WHERE nombre = %s"
    cursor.execute(sql, (_parroquia,))
    result = cursor.fetchone()
    if result is not None:

        # La cita ya existe, mostrar mensaje de flash y redirigir a la página anterior
        flash('Ya posee un parroquia con el mismo nombre', 'error')
        return redirect(request.referrer)

    # Insertar el parroquia si no existe
    sql = "INSERT INTO parroquias (cod_par, cod_mun, nombre) VALUES (NULL, %s, %s);"

    datos = (_cod_municipio, _parroquia, )

    cursor.execute(sql,datos)
    conn.commit()

    flash('¡Parroquia agregada exitosamente!', 'error')

    return redirect('/parroquias')

@app.route('/crear_sector', methods=['POST'])
def crear_sector():

    _sector = request.form['txtSector']
    _cod_municipio = request.form['txtMunicipio']

    conn = mysql.connect()
    cursor = conn.cursor()

    # Consultar si el sector ya existe en la base de datos

    sql = "SELECT * FROM sectors WHERE nombre = %s"
    cursor.execute(sql, (_sector,))
    result = cursor.fetchone()
    if result is not None:

        # La cita ya existe, mostrar mensaje de flash y redirigir a la página anterior
        flash('Ya posee un sector con el mismo nombre', 'error')
        return redirect(request.referrer)

    # Insertar el sector si no existe
    sql = "INSERT INTO sectores (cod_par, cod_mun, nombre) VALUES (NULL, %s, %s);"

    datos = (_cod_municipio, _sector, )

    cursor.execute(sql,datos)
    conn.commit()

    flash('¡sector agregada exitosamente!', 'error')

    return redirect('/sectores')

@app.route('/crear_municipio', methods=['POST'])
def crear_municipio():

    _municipio = request.form['txtMunicipio']

    conn = mysql.connect()
    cursor = conn.cursor()

    # Consultar si el municipio ya existe en la base de datos

    sql = "SELECT * FROM municipios WHERE nombre = %s"
    cursor.execute(sql, (_municipio,))
    result = cursor.fetchone()
    if result is not None:

        # La cita ya existe, mostrar mensaje de flash y redirigir a la página anterior
        flash('Ya posee un municipio con el mismo nombre', 'error')
        return redirect(request.referrer)

    # Insertar el Municipio si no existe
    sql = "INSERT INTO municipios (cod_mun, nombre) VALUES (NULL, %s);"

    datos = (_municipio)

    cursor.execute(sql,datos)
    conn.commit()

    flash('¡Municipio agregado exitosamente!', 'error')

    return redirect('/municipios')

@app.route('/login')
def login():
    return render_template('general/login.html')

@app.route('/')
def index():
    return render_template('general/principal.html')

@app.route('/registro')
def registro():
    return render_template('general/registro.html')

@app.route('/municipios')
def municipios():
    conn = mysql.connect()
    cursor = conn.cursor()
    cursor.execute("SELECT cod_mun AS codigo_municipio, nombre AS nombre_municipio FROM municipios")

    municipios = cursor.fetchall()
    conn.commit()

    return render_template('general/municipios.html', municipios=municipios)

@app.route('/parroquias')
def parroquias():
    conn = mysql.connect()
    cursor = conn.cursor()
    cursor.execute("SELECT p.*, m.nombre as municipio FROM parroquias p LEFT JOIN municipios m ON p.cod_mun = m.cod_mun")

    parroquias = cursor.fetchall()

    conn.commit()

    return render_template('general/parroquias.html', parroquias=parroquias)

# Codigo para Editar Usuario

@app.route('/edit_user/<int:id>')
def edit_user(id):

    if current_user.id != id:
        abort(404)

    conn = mysql.connect()
    cursor = conn.cursor()

    # Consulta para obtener los datos del usuario y sus teléfonos
    cursor.execute("SELECT * FROM usuarios WHERE id_usuario =%s",(id))
    usuarios = cursor.fetchall()

    conn.commit()

    return render_template("general/edit_user.html", usuarios=usuarios)

@app.route('/edit_password/<int:id>')
def edit_password(id):

    if current_user.id != id:
        abort(404)

    conn = mysql.connect()
    cursor = conn.cursor()

    # Consulta para obtener los datos del usuario y sus teléfonos
    cursor.execute("SELECT * FROM usuarios WHERE id_usuario=%s",(id))
    usuarios = cursor.fetchall()

    conn.commit()

    return render_template("general/edit_password.html", usuarios=usuarios)

#Funcion para guardar los cambios en el usuario editado

@app.route('/update_user', methods=['POST'])
def update_user():

    _nombre = request.form['txtNombre']
    _apellido = request.form['txtApellido']
    _telefono = request.form['txtTelefono']
    _id = request.form['txtID']
    _correo = request.form['txtCorreo']
    # _password = request.form['txtPassword']

    conn = mysql.connect()
    cursor = conn.cursor()

    # Obtener información de la bomba antes de la actualización (para el registro)
    cursor.execute("SELECT nombre FROM usuarios WHERE id_usuario = %s", (_id,))
    usuario_anterior = cursor.fetchone()

    if usuario_anterior[0] != _nombre:

        # Registrar la actualización en el historial de acciones
        _id_usuario = current_user.id
        _area = "USUARIOS"
        _accion = "ACTUALIZACION"

        ahora = datetime.now()
        _hora = ahora.strftime('%H:%M:%S')  # Formato de 24 horas para almacenamiento
        _fecha_actual = ahora.strftime('%Y-%m-%d')

        detalles = f'EL USUARIO: "{usuario_anterior[0]}"CAMBIÓ SU NOMBRE A: "{_nombre}"'
        sql_historial = '''
            INSERT INTO historial
            (id_historial, id_usuario, accion, area, detalles, hora, fecha)
            VALUES (NULL, %s, %s, %s, %s, %s, %s)
        '''
        datos_historial = (_id_usuario, _accion, _area, detalles, _hora, _fecha_actual)
        cursor.execute(sql_historial, datos_historial)
        conn.commit()


    # Update los datos del usuario en la tabla 'usuarios'
    sql_usuario = "UPDATE usuarios SET nombre=%s, apellido=%s, telefono=%s, correo=%s WHERE id_usuario=%s"
    datos_usuario = (_nombre, _apellido, _telefono, _correo, _id)



    try:
        cursor.execute(sql_usuario, datos_usuario)
        conn.commit()
        flash('¡Usuario modificado exitosamente!', 'error')

        return redirect('/menu')

    except Exception as e:
        conn.rollback()

        flash('¡Ha ocurrido un error!', 'error')

        return str(e)


    finally:
        cursor.close()
        conn.close()

@app.route('/update_password', methods=['POST'])
def update_password():

    usuario_id = request.form['txtID']
    new_password = request.form['txtPassword']
    old_password = request.form['txtOldPassword']

    conn = mysql.connect()
    cursor = conn.cursor()

    # Recuperar la contraseña antigua de la base de datos
    sql_select_old_password = "SELECT clave FROM usuarios WHERE id_usuario = %s"
    cursor.execute(sql_select_old_password, (usuario_id,))
    result = cursor.fetchone()

    if result is None:
        flash('¡Usuario no encontrado!', 'error')
        return redirect('/menu')

    stored_old_password = result[0]

    # Verificar si la contraseña antigua coincide
    if stored_old_password != old_password:
        flash('¡La contraseña antigua NO coincide!', 'error')
        return redirect('/menu')

    # Actualizar la contraseña si la antigua coincide
    sql_update_password = "UPDATE usuarios SET clave=%s WHERE id_usuario=%s"
    cursor.execute(sql_update_password, (new_password, usuario_id))
    conn.commit()

    flash('¡Contraseña modificada exitosamente!', 'success')
    return redirect('/menu')


#Fin del código para manipular los usuarios

### Aplicación de Buenas Prácticas ###

class MedicinaManager:
    def __init__(self, conn):
        self.conn = conn
        self.cursor = conn.cursor()

    def generar_id_medicina(self, prefijo):
        self.cursor.execute("SELECT id_medicina FROM medicinas WHERE id_medicina LIKE %s ORDER BY id_medicina DESC LIMIT 1", (f"{prefijo}-%",))
        last_id = self.cursor.fetchone()
        new_id_num = int(last_id[0].split('-')[1]) + 1 if last_id else 1
        return f"{prefijo}-{new_id_num:03d}"

    def agregar_medicina(self, descripcion, marca, dosis, unidad):
        id_medicina = self.generar_id_medicina("MED")
        self.cursor.execute("INSERT INTO medicinas (id_medicina, descripcion, marca, dosis, unidad) VALUES (%s, %s, %s, %s, %s)",
                            (id_medicina, descripcion, marca, dosis, unidad))
        return id_medicina

    def actualizar_stock(self, id_medicina, cantidad):
        self.cursor.execute("SELECT cantidad FROM stock_medicinas WHERE id_medicina = %s", (id_medicina,))
        stock_result = self.cursor.fetchone()
        nueva_cantidad = stock_result[0] + cantidad if stock_result else cantidad
        self.cursor.execute("INSERT INTO stock_medicinas (id_medicina, cantidad) VALUES (%s, %s) ON DUPLICATE KEY UPDATE cantidad=%s",
                            (id_medicina, nueva_cantidad, nueva_cantidad))

    def obtener_medicina_por_detalle(self, descripcion, unidad, marca, dosis):
        self.cursor.execute(
            "SELECT id_medicina FROM medicinas WHERE descripcion = %s AND unidad = %s AND marca = %s AND dosis = %s",
            (descripcion, unidad, marca, dosis)
        )
        return self.cursor.fetchone()





















# Ejemplo de uso
@app.route('/agregar_medicina_BD', methods=['POST'])
def agregar_medicina_BD():

    conn = mysql.connect()
    manager = MedicinaManager(conn)   # Instancia de Medicina Manager
    id_generator = IDGenerator(conn)  # Instancia de IDGenerator

    try:
        # Datos del formulario
        fecha = request.form['txtFecha']
        medicinas = [medicina.upper() for medicina in request.form.getlist('medicina[]')]
        marcas = [marca.upper() for marca in request.form.getlist('marca[]')]
        dosis = [dosis.upper() for dosis in request.form.getlist('dosis[]')]
        cantidades = request.form.getlist('cantidad[]')
        unidades = [unidad.upper() for unidad in request.form.getlist('unidad[]')]

        # Generar y agregar el ID de entrada
        id_entrada = id_generator.generar_id_entrada_medicina()
        conn.cursor().execute("INSERT INTO entradas_medicinas (id_entrada, fecha) VALUES (%s, %s)", (id_entrada, fecha))

        # Variables para historial y mensajes
        medicinas_existentes = []
        detalles_historial = []

        # Procesar cada medicina
        for medicina, marca, dosis, cantidad, unidad in zip(medicinas, marcas, dosis, cantidades, unidades):
            # Verificar si la medicina existe o agregarla
            result = manager.obtener_medicina_por_detalle(medicina, marca, dosis, unidad)

            if result:
                id_medicina = result[0]
                medicinas_existentes.append(medicina)
            else:
                id_medicina = manager.agregar_medicina(medicina, marca, dosis, unidad)

            # Insertar en aux_entradas_medicinas y actualizar stock
            manager.cursor.execute(
                "INSERT INTO aux_entradas_medicinas (id_entrada, id_medicina, cantidad) VALUES (%s, %s, %s)",
                (id_entrada, id_medicina, cantidad))

            # Agregar detalles al historial
            detalles_historial.append(f"MEDICINA: {medicina}, CANTIDAD: {cantidad}, UNIDAD: {unidad}")

        manager.actualizar_stock(id_medicina, int(cantidad))

        # Confirmar cambios en la base de datos
        conn.commit()

        # Mensajes de éxito o advertencia
        if medicinas_existentes:
            flash(f'Algunos artículos ya existían y fueron sumados al inventario: {", ".join(medicinas_existentes)}', 'warning')
        else:
            flash('¡Artículos agregados exitosamente!', 'success')

        # Registrar en el historial
        if detalles_historial:
            registrar_historial(conn, current_user.id, "REGISTRO", "MEDICINAS", "\n".join(detalles_historial))

    except Exception as e:
        conn.rollback()
        flash(f'Error al agregar artículos: {str(e)}', 'error')
    finally:
        conn.close()

    return redirect('/medicinas/inventario')










### Gestion de Medicinas ###

@app.route('/medicinas/agregar_medicina')
def agregar_medicina():

    conn = mysql.connect()
    cursor = conn.cursor()

    #Vista para el usuario director (o administrador)

    conn.commit()

    return render_template('/medicinas/agregar_medicina.html')

@app.route('/medicinas/editar_medicina/<string:id_medicina>')
def editar_medicina(id_medicina):

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM medicinas WHERE id_medicina = %s", (id_medicina))

    medicina = cursor.fetchone()

    return render_template('medicinas/editar_medicina.html', medicina = medicina)

@app.route('/medicinas/editar_medicina_BD', methods=['POST'])
def medicinas_editar_medicina_BD():

    try:

        id_medicina = request.form['txtID']
        _medicina = request.form["medicina[]"].upper()
        _marca = request.form["marca[]"].upper()
        _dosis = request.form["dosis[]"].upper()
        _unidad = request.form["unidad[]"].upper()

        conn = mysql.connect()
        cursor = conn.cursor()

        # Actualizar la fecha en la tabla de entradas_medicinas
        cursor.execute("UPDATE medicinas SET descripcion = %s, marca = %s, dosis = %s, unidad = %s WHERE id_medicina = %s", (_medicina, _marca, _dosis,_unidad, id_medicina))

        conn.commit()

        flash('¡Medicina actualizada exitosamente!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Error al actualizar el medicina: {str(e)}', 'error')

    finally:
        cursor.close()
        conn.close()

    return redirect('/medicinas/inventario')

@app.route('/eliminar_multiples_medicinas', methods=['POST'])
@login_required
def eliminar_multiples_medicinas():
    ids = request.json.get('ids', [])
    if isinstance(ids, str):  # Si ids es un solo ID en lugar de una lista
        ids = [ids]
    elif not ids:  # Si no se proporciona ningún ID
        return jsonify(success=False, message='No IDs provided'), 400

    try:
        conn = mysql.connect()
        cursor = conn.cursor()

        # Obtener información de los artículos antes de eliminarlos (para el registro)
        cursor.execute("SELECT id_medicina, descripcion, marca, dosis, unidad FROM medicinas WHERE id_medicina IN %s", (tuple(ids),))
        medicinas = cursor.fetchall()

        # Eliminar los artículos
        format_strings = ','.join(['%s'] * len(ids))
        query = f"DELETE FROM medicinas WHERE id_medicina IN ({format_strings})"
        cursor.execute(query, ids)

        conn.commit()

        # Registrar cada eliminación en el historial de acciones
        _id_usuario = current_user.id
        _area = "MEDICINAS"
        _accion = "ELIMINACION"

        for medicina in medicinas:
            codigo_medicina = medicina[0]
            nombre_medicina = medicina[1]
            _detalles = f"SE ELIMINÓ: {nombre_medicina}, CODIGO: {codigo_medicina}"
            registrar_historial(conn, _id_usuario, _accion, _area, _detalles)

        return jsonify(success=True)

    except Exception as e:
        conn.rollback()
        return jsonify(success=False, message=str(e)), 500

    finally:
        cursor.close()
        conn.close()

@app.route('/medicinas/inventario_pdf', methods=['POST'])
def medicinas_inventario_pdf():
    conn = mysql.connect()
    cursor = conn.cursor()

    # Extraer solo medicinas con cantidad mayor a 0
    cursor.execute('''
        SELECT
            med.id_medicina,
            med.descripcion,
            med.marca,
            med.dosis,
            med.unidad,
            IFNULL(stock.cantidad, 0) AS cantidad
        FROM
            medicinas med
        LEFT JOIN
            stock_medicinas stock ON med.id_medicina = stock.id_medicina
        WHERE
            stock.cantidad > 0
        ORDER BY
            med.descripcion ASC
    ''')

    rows = cursor.fetchall()

    # Obtener la sumatoria total de todas las cantidades en stock mayores a 0
    cursor.execute('''
        SELECT
            SUM(cantidad) AS total_cantidad
        FROM
            stock_medicinas
        WHERE
            cantidad > 0
    ''')

    # Convertir el resultado de la consulta a un número entero
    total_result = cursor.fetchone()
    total = int(total_result[0]) if total_result[0] is not None else 0
    print("total es: ", total)

    conn.close()

    # Renderizar el template con los datos agrupados y el total
    rendered = render_template('medicinas/reportes/reporte_inventario.html', datos=rows, total=total)

    # Generar el PDF con WeasyPrint
    html = HTML(string=rendered)
    pdf = html.write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=reporte_inventario.pdf'

    return response


### Gestion de Medicinas ENTRADAS ###

@app.route('/medicinas/entradas')
@login_required
def medicinas_entradas():

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT
            aux_entradas_medicinas.id,
            DATE_FORMAT(entradas_medicinas.fecha, '%d/%m/%Y') as fecha,
            medicinas.descripcion,
            aux_entradas_medicinas.cantidad,
            medicinas.marca,
            medicinas.dosis,
            medicinas.unidad,
            entradas_medicinas.id_entrada
        FROM
            entradas_medicinas
        JOIN
            aux_entradas_medicinas ON entradas_medicinas.id_entrada = aux_entradas_medicinas.id_entrada
        JOIN
            medicinas ON aux_entradas_medicinas.id_medicina = medicinas.id_medicina
        ORDER BY
            entradas_medicinas.id_entrada ASC, entradas_medicinas.fecha ASC
    ''')

    entradas = cursor.fetchall()

    # Agrupar las entradas por ID Entrada
    grouped_entradas = []
    current_id_entrada = None
    group = []

    for entrada in entradas:
        if entrada[7] != current_id_entrada:  # Index 7 corresponde a id_entrada
            if group:
                grouped_entradas.append(group)
            group = [entrada]
            current_id_entrada = entrada[7]
        else:
            group.append(entrada)

    if group:
        grouped_entradas.append(group)

    # Calcular la cantidad total de artículos ingresados
    cursor.execute('''
        SELECT
            SUM(ae.cantidad) as cantidad_total
        FROM
            aux_entradas_medicinas ae
    ''')
    cantidad_total = cursor.fetchone()[0]

    cursor.close()
    conn.close()

    return render_template('/medicinas/entradas.html', grouped_entradas=grouped_entradas, cantidad_total=cantidad_total)

@app.route('/medicinas/agregar_entrada')
def agregar_entrada_medicinas():

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT med.id_medicina,
                med.descripcion,
                med.marca,
                med.dosis,
                med.unidad,
                IFNULL(SUM(stock.cantidad), 0) AS total_cantidad
        FROM medicinas med
        LEFT JOIN stock_medicinas stock ON med.id_medicina = stock.id_medicina
        GROUP BY med.id_medicina, med.descripcion, med.marca, med.dosis, med.unidad
        ORDER BY med.descripcion
    """)

    articulos = cursor.fetchall()

    conn.commit()

    return render_template('medicinas/agregar_entrada.html', medicinas = articulos)

@app.route('/medicinas/eliminar_multiples_entradas', methods=['POST'])
@login_required
def eliminar_multiples_entradas_medicinas():
    ids = request.json.get('ids')

    if not ids:
        return jsonify({'success': False, 'message': 'No se proporcionaron IDs'})

    conn = mysql.connect()
    cursor = conn.cursor()

    try:
        entradas_a_verificar = set()
        medicinas_cache = {}

        for id_entrada in ids:
            # Obtener todos los registros asociados al id_entrada en aux_entradas_medicinas
            cursor.execute('''
                            SELECT id_medicina, cantidad
                            FROM aux_entradas_medicinas
                            WHERE id_entrada=%s
                           ''', (id_entrada,))
            resultados = cursor.fetchall()

            if not resultados:
                continue

            for id_medicina, cantidad in resultados:
                # Obtener nombre del artículo desde la caché o la base de datos si no está en la caché
                if id_medicina in medicinas_cache:
                    nombre_medicina = medicinas_cache[id_medicina]
                else:
                    cursor.execute('''
                                    SELECT descripcion
                                    FROM medicinas
                                    WHERE id_medicina=%s
                                   ''', (id_medicina,))
                    nombre_medicina = cursor.fetchone()[0]
                    medicinas_cache[id_medicina] = nombre_medicina

                # Actualizar la tabla stock_medicinas disminuyendo la cantidad eliminada
                cursor.execute('''
                                UPDATE stock_medicinas
                                SET cantidad = GREATEST(cantidad - %s, 0)
                                WHERE id_medicina=%s
                               ''', (cantidad, id_medicina))

                # Registrar la eliminación de cada artículo en el historial
                detalles = f'SE ELIMINÓ {cantidad} DEL ARTÍCULO: "{nombre_medicina}" EN LA SALIDA: {id_entrada}.'
                registrar_historial(conn, current_user.id, "ELIMINACIÓN", "MEDICINAS", detalles)

            # Eliminar todos los registros de aux_entradas_medicinas asociados a esta salida
            cursor.execute('''
                            DELETE FROM aux_entradas_medicinas
                            WHERE id_entrada=%s
                           ''', (id_entrada,))

            # Registrar la eliminación de la salida completa en el historial
            detalles = f"SE ELIMINÓ COMPLETAMENTE LA SALIDA: ID {id_entrada}."
            registrar_historial(conn, current_user.id, "ELIMINACIÓN", "FERRETERÍA", detalles)

            # Eliminar la salida completa de entradas_medicinas
            cursor.execute('''
                            DELETE FROM entradas_medicinas
                            WHERE id_entrada=%s
                           ''', (id_entrada,))

        conn.commit()
        return jsonify({'success': True})

    except Exception as e:
        conn.rollback()
        print(f"Error: {e}")
        return jsonify({'success': False, 'message': str(e)})

    finally:
        cursor.close()
        conn.close()

@app.route('/medicinas/reportes/reporte_entradas_pdf', methods=['POST'])
def medicinas_entradas_pdf():
    data = request.get_json()
    fecha_inicio = data.get('fechaInicio')
    fecha_fin = data.get('fechaFin')
    search_input = data.get('searchInput', '').upper()  # Convertir a mayúsculas

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT
            entradas_medicinas.id_entrada,
            DATE_FORMAT(entradas_medicinas.fecha, '%d/%m/%Y') as fecha,
            med.descripcion,
            med.marca,
            med.dosis,
            aux_ent.cantidad,
            med.unidad
        FROM
            entradas_medicinas
        JOIN
            aux_entradas_medicinas aux_ent ON entradas_medicinas.id_entrada = aux_ent.id_entrada
        JOIN
            medicinas med ON aux_ent.id_medicina = med.id_medicina
        ORDER BY
            entradas_medicinas.id_entrada ASC, entradas_medicinas.fecha ASC
    ''')

    rows = cursor.fetchall()
    conn.close()

    # Procesar los datos para agrupar por id_entrada
    grouped_data = {}
    for row in rows:
        id_entrada = row[0]
        if id_entrada not in grouped_data:
            grouped_data[id_entrada] = {
                'id_entrada': id_entrada,
                'fecha': row[1],
                'entradas': []
            }
        grouped_data[id_entrada]['entradas'].append({
            'medicina': row[2],
            'marca': row[3],
            'dosis': row[4],
            'unidad': row[6],
            'cantidad': row[5]
        })

    # Convertir el diccionario a una lista para facilitar el manejo en la plantilla
    final_data = []
    for idx, (id_entrada, details) in enumerate(grouped_data.items(), start=1):
        final_data.append({
            'index': idx,
            'id_entrada': details['id_entrada'],
            'fecha': details['fecha'],
            'rowspan': len(details['entradas']),
            'entradas': details['entradas']
        })

    # Renderizar el template con los datos agrupados
    rendered = render_template('medicinas/reportes/reporte_entradas.html', datos=final_data)

    # Generar el PDF con WeasyPrint
    html = HTML(string=rendered)
    pdf = html.write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=reporte_entradas.pdf'

    return response


@app.route('/bombas')
@login_required
def bombas():
    conn = mysql.connect()
    cursor = conn.cursor()

    #Vista para el usuario director (o administrador)

    cursor.execute("SELECT * FROM bombas ")

    bombas = cursor.fetchall()

    conn.commit()

    return render_template('general/bombas.html', bombas = bombas)

@app.route('/historial')
@login_required
def historial():
    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT h.id_historial,
               DATE_FORMAT(h.fecha, '%d/%m/%Y') as fecha,
               TIME_FORMAT(h.hora, '%l:%i %p') AS hora,
               u.nombre AS nombre_usuario,
               h.area,
               h.accion,
               h.detalles
        FROM historial h
        JOIN usuarios u ON h.id_usuario = u.id_usuario
        ORDER BY h.id_historial DESC;

    ''')

    historial = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('general/historial.html', historial=historial)

@app.route('/entregas')
@login_required
def entregas():
    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute('''SELECT ID,
                   DATE_FORMAT(fecha, '%d/%m/%Y') as fecha,
                   solicitado,
                   cantidad,
                   fuerza,
                   tipo,
                   voltios,
                   municipio,
                   parroquia,
                   sector
                   FROM bombas_entregadas
                   ORDER BY fecha ASC''')  # Puedes ordenar los resultados por fecha u otra columna si lo deseas

    entregas = cursor.fetchall()

    # Calcular la cantidad total de bombas entregadas
    cursor.execute('''SELECT SUM(cantidad) as cantidad_total
                   FROM bombas_entregadas''')
    cantidad_total = cursor.fetchone()[0]  # Obtenemos el valor de cantidad_total desde la primera columna del primer resultado

    conn.commit()
    cursor.close()
    conn.close()

    return render_template('bombas/entregas.html', entregas=entregas, cantidad_total=cantidad_total)


#Ir a la plantilla HTML de agregar bombas
@app.route('/agregar_entrega')
def agregar_entrega():

    conn = mysql.connect()
    cursor = conn.cursor()

    #Vista para el usuario director (o administrador)

    cursor.execute("SELECT * FROM municipios ORDER BY nombre")

    municipios = cursor.fetchall()

    conn.commit()

    return render_template('bombas/agregar_entrega.html', municipios = municipios)

@app.route('/medicinas/agregar_articulo_BD', methods=['POST'])
def medicinas_agregar_articulo_BD():
    try:
        fecha = request.form['txtFecha']
        articulos = [articulo.upper() for articulo in request.form.getlist('articulo[]')]
        cantidades = request.form.getlist('cantidad[]')
        unidades = [unidad.upper() for unidad in request.form.getlist('unidad[]')]
        areas = [area.upper() for area in request.form.getlist('area[]')]

        conn = mysql.connect()
        cursor = conn.cursor()

        # Obtener el último ID de entrada
        cursor.execute("SELECT id_entrada FROM entradas_medicinas ORDER BY id_entrada DESC LIMIT 1")
        last_id = cursor.fetchone()

        if last_id:
            last_id_num = int(last_id[0].split('-')[1])
            new_id_num = last_id_num + 1
        else:
            new_id_num = 1

        new_id_entrada = f"EN-{new_id_num:04d}"

        # Insertar en la tabla de entradas_medicinas
        cursor.execute("INSERT INTO entradas_medicinas (id_entrada, fecha) VALUES (%s, %s)", (new_id_entrada, fecha))

        # Lista para almacenar los artículos que ya existen
        articulos_existentes = []

        detalles_historial = []

        for articulo, cantidad, unidad, area in zip(articulos, cantidades, unidades, areas):

            prefijo = area[:3]

            # Verificar si el artículo ya existe en la tabla articulos
            cursor.execute("SELECT id_articulo FROM articulos WHERE articulo = %s AND unidad = %s", (articulo, unidad))
            result = cursor.fetchone()

            if result:
                id_articulo = result[0]
                articulos_existentes.append(articulo)
            else:
                # Encontrar el último código insertado para el área
                cursor.execute("SELECT id_articulo FROM articulos WHERE id_articulo LIKE %s ORDER BY id_articulo DESC LIMIT 1", (f"{prefijo}-%",))
                last_id = cursor.fetchone()

                if last_id:
                    last_number = int(last_id[0].split('-')[1])
                    new_number = last_number + 1
                else:
                    new_number = 1

                # Generar el nuevo código
                id_articulo = f"{prefijo}-{new_number:03}"

                # Insertar el artículo en la tabla articulos
                cursor.execute("INSERT INTO articulos (id_articulo, articulo, unidad, area) VALUES (%s, %s, %s, %s)", (id_articulo, articulo, unidad, area))

            # Insertar en la tabla aux_entradas_medicinas usando el correcto id_entrada
            cursor.execute("INSERT INTO aux_entradas_medicinas (id_entrada, id_articulo, cantidad) VALUES (%s, %s, %s)", (new_id_entrada, id_articulo, cantidad))

            # Actualizar la tabla de stock
            cursor.execute("SELECT cantidad FROM stock_medicinas WHERE id_articulo = %s", (id_articulo,))
            stock_result = cursor.fetchone()

            if stock_result:
                nueva_cantidad = stock_result[0] + int(cantidad)
                cursor.execute("UPDATE stock_medicinas SET cantidad = %s WHERE id_articulo = %s", (nueva_cantidad, id_articulo))
            else:
                cursor.execute("INSERT INTO stock_medicinas (id_articulo, cantidad) VALUES (%s, %s)", (id_articulo, cantidad))

            # Agregar detalles al historial
            detalle = f"ARTÍCULO: {articulo}, CANTIDAD: {cantidad}, UNIDAD: {unidad}, ÁREA: {area}"
            detalles_historial.append(detalle)

        conn.commit()

        # Generar el mensaje de éxito y error
        if articulos_existentes:
            mensaje = 'Algunos artículos ya existían y fueron sumados al inventario: ' + ', '.join(articulos_existentes)
            flash(mensaje, 'warning')
        else:
            flash('¡Artículos agregados exitosamente!', 'success')

        # Registrar en el historial
        print("Intentando registrar en el historial...")
        if detalles_historial:
            _detalles = "\n".join(detalles_historial)
            print(f"Detalles del historial: {_detalles}")
            registrar_historial(conn, current_user.id, "REGISTRO", "FERRETERÍA", _detalles)
        else:
            print("No hay detalles del historial para registrar.")


    except Exception as e:
        conn.rollback()
        flash('Error al agregar artículos: {}'.format(str(e)), 'error')
    finally:
        cursor.close()
        conn.close()

    return redirect('/medicinas/inventario')

@app.route('/medicinas/agregar_articulo')
def agregar_articulo_medicinas():

    conn = mysql.connect()
    cursor = conn.cursor()

    #Vista para el usuario director (o administrador)

    cursor.execute("SELECT * FROM municipios ORDER BY nombre")

    municipios = cursor.fetchall()

    conn.commit()

    return render_template('medicinas/agregar_articulo.html', municipios = municipios)

### Gestion de Medicinas (SALIDAS) ###

@app.route('/medicinas/agregar_salida')
def agregar_salida_medicinas():

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT med.id_medicina,
                med.descripcion,
                med.marca,
                med.dosis,
                med.unidad,
                IFNULL(SUM(stock.cantidad), 0) AS total_cantidad
        FROM medicinas med
        LEFT JOIN stock_medicinas stock ON med.id_medicina = stock.id_medicina
        GROUP BY med.id_medicina, med.descripcion, med.marca, med.dosis, med.unidad
        ORDER BY med.descripcion
    """)

    articulos = cursor.fetchall()

    conn.commit()

    return render_template('medicinas/agregar_salida.html', articulos = articulos)

@app.route('/medicinas/agregar_salida_BD', methods=['POST'])
@login_required
def medicinas_agregar_salida_BD():

    try:
        fecha = request.form['txtFecha']
        destino = request.form['txtDestino'].upper()
        medicinas = [medicina.upper() for medicina in request.form.getlist('medicina[]')]
        marcas = [marca.upper() for marca in request.form.getlist('marca[]')]
        dosis = [dosis.upper() for dosis in request.form.getlist('dosis[]')]
        cantidades = request.form.getlist('cantidad[]')
        unidades = [unidad.upper() for unidad in request.form.getlist('unidad[]')]

        conn = mysql.connect()
        cursor = conn.cursor()

        # Obtener el último ID de salida
        cursor.execute("SELECT id_salida FROM salidas_medicinas ORDER BY id_salida DESC LIMIT 1")
        last_id = cursor.fetchone()

        if last_id:
            last_id_num = int(last_id[0].split('-')[1])
            new_id_num = last_id_num + 1
        else:
            new_id_num = 1

        new_id_salida = f"SA-{new_id_num:04d}"

        # Insertar en la tabla de salidas_medicinas con el nuevo ID
        cursor.execute("INSERT INTO salidas_medicinas (id_salida, fecha, destino) VALUES (%s, %s, %s)", (new_id_salida, fecha, destino))

        _id_usuario = current_user.id
        _accion = "SALIDAS"
        detalles_historial = []

        for medicina, marca, dosis, cantidad, unidad in zip(medicinas, marcas, dosis, cantidades, unidades):

            # Obtener el ID del artículo directamente, asumiendo que ya existe en la tabla `articulos`
            cursor.execute("SELECT id_medicina FROM medicinas WHERE descripcion = %s AND unidad = %s AND marca = %s AND dosis =%s", (medicina, unidad, marca, dosis))
            result = cursor.fetchone()

            if not result:
                # Si no existe el artículo, crearlo
                prefijo = "MED"
                cursor.execute("SELECT id_medicina FROM medicinas WHERE id_medicina LIKE %s ORDER BY id_medicina DESC LIMIT 1", (f"{prefijo}-%",))
                last_id_medicina = cursor.fetchone()

                if last_id_medicina:
                    last_number = int(last_id_medicina[0].split('-')[1])
                    new_number = last_number + 1
                else:
                    new_number = 1

                # Generar el nuevo código para el artículo
                id_medicina = f"{prefijo}-{new_number:03}"
                cursor.execute("INSERT INTO medicinas (id_medicina, descripcion, marca, dosis, unidad) VALUES (%s, %s, %s, %s, %s)", (id_medicina, medicina, marca, dosis, unidad))

                # Obtener el último ID de entrada
                cursor.execute("SELECT id_entrada FROM entradas_medicinas ORDER BY id_entrada DESC LIMIT 1")
                last_id = cursor.fetchone()

                if last_id:
                    last_id_num = int(last_id[0].split('-')[1])
                    new_id_num = last_id_num + 1
                else:
                    new_id_num = 1

                new_id_entrada = f"EN-{new_id_num:04d}"

                # Insertar en la tabla de entradas_medicinas
                cursor.execute("INSERT INTO entradas_medicinas (id_entrada, fecha) VALUES (%s, %s)", (new_id_entrada, fecha))

                # Insertar en la tabla aux_entradas_medicinas
                cursor.execute("INSERT INTO aux_entradas_medicinas (id_entrada, id_medicina, cantidad) VALUES (%s, %s, %s)", (new_id_entrada, id_medicina, cantidad))

                # Insertar el articulo en la tabla Stock medicinas
                cursor.execute("INSERT INTO stock_medicinas (id_medicina, cantidad) VALUES (%s, %s)", (id_medicina, cantidad))

            else: #El articulo sí existe.
                id_medicina = result[0]

            # Insertar en la tabla aux_salidas_medicinas
            cursor.execute("INSERT INTO aux_salidas_medicinas (id_salida, id_medicina, cantidad) VALUES (%s, %s, %s)", (new_id_salida, id_medicina, cantidad))

            # Actualizar la tabla de stock_medicinas
            cursor.execute("SELECT cantidad FROM stock_medicinas WHERE id_medicina = %s", (id_medicina,))
            stock_result = cursor.fetchone()

            nueva_cantidad = stock_result[0] - int(cantidad)

            if nueva_cantidad >= 0:
                cursor.execute("UPDATE stock_medicinas SET cantidad = %s WHERE id_medicina = %s", (nueva_cantidad, id_medicina))
            else:
                raise ValueError("La cantidad de stock no puede ser negativa.")

            # Agregar detalles al historial
            detalle = f'ENTREGADO {cantidad} {unidad} DE "{medicina}"'
            detalles_historial.append(detalle)

        conn.commit()

        if detalles_historial:
            _detalles = "\n".join(detalles_historial)
            registrar_historial(conn, _id_usuario, _accion, "MEDICINAS", _detalles)

        conn.rollback()

        flash('¡Salidas agregadas y stock actualizado exitosamente!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Error al agregar salidas: {str(e)}', 'error')
    finally:
        cursor.close()
        conn.close()

    return redirect('/medicinas/salidas')

@app.route('/medicinas/editar_salida/<string:id_salida>')
def editar_salida_medicinas(id_salida):

    conn = mysql.connect()
    cursor = conn.cursor()

    # Vista para el usuario director (o administrador)
    cursor.execute("""
        SELECT m.id_medicina,
               m.descripcion,
               m.marca,
               m.dosis,
               m.unidad,
               COALESCE(s.cantidad, 0) AS cantidad
        FROM medicinas m
        LEFT JOIN stock_medicinas s ON m.id_medicina = s.id_medicina
        ORDER BY m.descripcion
    """)

    articulos = cursor.fetchall()

    # Modificación para incluir stock disponible
    cursor.execute("""
      SELECT
            ent.id_salida,
            med.descripcion,
            med.marca,
            med.dosis,
            med.unidad,
            aux_ent.cantidad,
            ent.fecha,
            med.id_medicina,
            ent.destino,
            COALESCE(stock.cantidad, 0) AS stock_disponible  -- Añadido el stock disponible
        FROM medicinas med
        LEFT JOIN aux_salidas_medicinas aux_ent
            ON med.id_medicina = aux_ent.id_medicina
        LEFT JOIN salidas_medicinas ent
            ON aux_ent.id_salida = ent.id_salida
        LEFT JOIN stock_medicinas stock  -- Unir la tabla de stock
            ON med.id_medicina = stock.id_medicina
        WHERE aux_ent.id_salida = %s
        ORDER BY med.descripcion
    """, (id_salida,))

    salidas = cursor.fetchall()

    # Convertir los resultados de 'salidas' en una lista de diccionarios
    salidas_dict = [
        {
            'id_salida': salida[0],
            'descripcion': salida[1],
            'marca': salida[2],
            'dosis': salida[3],
            'unidad': salida[4],
            'cantidad': salida[5],
            'fecha': salida[6],
            'id_medicina': salida[7],
            'destino': salida[8],
            'stock_disponible': salida[9]  # Añadir stock disponible al diccionario
        }
        for salida in salidas
    ]

    return render_template('medicinas/editar_salida.html', medicinas=articulos, salidas=salidas_dict)


@app.route('/medicinas/editar_salida_BD', methods=['POST'])
def medicinas_editar_salida_BD():
    # try:
        id_salida = request.form['txtID']
        fecha = request.form['txtFecha']
        destino = request.form['txtDestino'].upper()
        medicinas = [medicina.upper() for medicina in request.form.getlist('medicina[]')]
        marcas = [marca.upper() for marca in request.form.getlist('marca[]')]
        dosis = [dosis.upper() for dosis in request.form.getlist('dosis[]')]
        cantidades = request.form.getlist('cantidad[]')
        unidades = [unidad.upper() for unidad in request.form.getlist('unidad[]')]

        conn = mysql.connect()
        cursor = conn.cursor()

        # Actualizar la fecha en la tabla de salidas_medicinas
        cursor.execute("UPDATE salidas_medicinas SET fecha = %s, destino = %s WHERE id_salida = %s", (fecha, destino, id_salida))

        # Obtener los artículos existentes en la salida
        cursor.execute("SELECT id_medicina, cantidad FROM aux_salidas_medicinas WHERE id_salida = %s", (id_salida,))
        existing_entries = cursor.fetchall()
        existing_articles = {row[0]: row[1] for row in existing_entries}

        new_articles = set()
        for medicina, marca, dosis, cantidad, unidad in zip(medicinas, marcas, dosis, cantidades, unidades):

            # Obtener el ID del artículo directamente, asumiendo que ya existe en la tabla `articulos`
            cursor.execute("SELECT id_medicina FROM medicinas WHERE descripcion = %s AND unidad = %s AND marca =%s AND dosis = %s", (medicina, unidad, marca, dosis))
            result = cursor.fetchone()

            if result:
                id_medicina = result[0]
                new_articles.add(id_medicina)
            else:
                # Si no existe el artículo, crearlo
                prefijo = "MED"
                cursor.execute("SELECT id_medicina FROM medicinas WHERE id_medicina LIKE %s ORDER BY id_medicina DESC LIMIT 1", (f"{prefijo}-%",))
                last_id_medicina = cursor.fetchone()

                if last_id_medicina:
                    last_number = int(last_id_medicina[0].split('-')[1])
                    new_number = last_number + 1
                else:
                    new_number = 1

                # Generar el nuevo código para el artículo
                id_medicina = f"{prefijo}-{new_number:03}"
                cursor.execute("INSERT INTO medicinas (id_medicina, descripcion, marca, dosis, unidad) VALUES (%s, %s, %s, %s, %s)", (id_medicina, medicina, marca, dosis, unidad))

            cantidad = int(cantidad)

            # Verificar si el artículo ya existe en la salida

            if id_medicina in existing_articles:
                old_cantidad = existing_articles[id_medicina]
                print("Old cantidad es: ", old_cantidad)
                diferencia = cantidad - old_cantidad
                print("Diferencia es: ", diferencia)

                # Actualizar la cantidad en aux_salidas_medicinas
                cursor.execute("UPDATE aux_salidas_medicinas SET cantidad = %s WHERE id_salida = %s AND id_medicina = %s", (cantidad, id_salida, id_medicina))

                # Actualizar la tabla stock_medicinas
                # cursor.execute("UPDATE stock_medicinas SET cantidad = cantidad + %s WHERE id_medicina = %s", (diferencia, id_medicina))
            else:
                # Si no existe, insertarlo en aux_salidas_medicinas
                print(id_salida, id_medicina, cantidad )
                cursor.execute("INSERT INTO aux_salidas_medicinas (id_salida, id_medicina, cantidad) VALUES (%s, %s, %s)", (id_salida, id_medicina, cantidad))

                # Actualizar el stock aumentando la cantidad
                # cursor.execute("UPDATE stock_medicinas SET cantidad = cantidad + %s WHERE id_medicina= %s", (cantidad, id_medicina))

        # Eliminar artículos que ya no están en la lista
        articles_to_remove = set(existing_articles.keys()) - new_articles
        for art_id in articles_to_remove:
            old_cantidad = existing_articles[art_id]

            # Reducir el stock por la cantidad eliminada
            # cursor.execute("UPDATE stock_medicinas SET cantidad = cantidad - %s WHERE id_medicina = %s", (old_cantidad, art_id))

            # Eliminar el artículo de aux_salidas_medicinas
            cursor.execute("DELETE FROM aux_salidas_medicinas WHERE id_salida = %s AND id_medicina = %s", (id_salida, art_id))

        conn.commit()

        # Registrar la acción en el historial
        for medicina, marca, dosis, cantidad, unidad in zip(medicinas, marcas, dosis, cantidades, unidades):
            detalles = f"SE ACTUALIZÓ {cantidad} {unidad} DE {medicina} EN LA salida {id_salida} EL {fecha}"
            registrar_historial(conn, current_user.id, "EDICIÓN", "MEDICINAS", detalles)

        flash('¡Salida actualizada exitosamente!', 'success')


        Actualizar_Stock_Medicinas()

    # except Exception as e:
    #     conn.rollback()
    #     flash(f'Error al actualizar la salida: {str(e)}', 'error')
    # finally:
        # conn.rollback()
        cursor.close()
        conn.close()

        return redirect('/medicinas/salidas')

@app.route('/medicinas/eliminar_multiples_salidas', methods=['POST'])
@login_required
def eliminar_multiples_salidas_medicinas():
    ids = request.json.get('ids')

    if not ids:
        return jsonify({'success': False, 'message': 'No se proporcionaron IDs'})

    conn = mysql.connect()
    cursor = conn.cursor()

    try:
        salidas_a_verificar = set()
        medicinas_cache = {}

        for id_salida in ids:
            # Obtener todos los registros asociados al id_salida en aux_salidas_medicinas
            cursor.execute('''
                            SELECT id_medicina, cantidad
                            FROM aux_salidas_medicinas
                            WHERE id_salida=%s
                           ''', (id_salida,))
            resultados = cursor.fetchall()

            if not resultados:
                continue

            for id_medicina, cantidad in resultados:
                # Obtener nombre del artículo desde la caché o la base de datos si no está en la caché
                if id_medicina in medicinas_cache:
                    nombre_medicina = medicinas_cache[id_medicina]
                else:
                    cursor.execute('''
                                    SELECT descripcion
                                    FROM medicinas
                                    WHERE id_medicina=%s
                                   ''', (id_medicina,))
                    nombre_medicina = cursor.fetchone()[0]
                    medicinas_cache[id_medicina] = nombre_medicina

                # Actualizar la tabla stock_medicinas aumentando la cantidad eliminada
                cursor.execute('''
                                UPDATE stock_medicinas
                                SET cantidad = GREATEST(cantidad + %s, 0)
                                WHERE id_medicina=%s
                               ''', (cantidad, id_medicina))

                # Registrar la eliminación de cada artículo en el historial
                detalles = f'SE ELIMINÓ {cantidad} DEL ARTÍCULO: "{nombre_medicina}" EN LA SALIDA: {id_salida}.'
                registrar_historial(conn, current_user.id, "ELIMINACIÓN", "MEDICINAS", detalles)

            # Eliminar todos los registros de aux_salidas_medicinas asociados a esta salida
            cursor.execute('''
                            DELETE FROM aux_salidas_medicinas
                            WHERE id_salida=%s
                           ''', (id_salida,))

            # Registrar la eliminación de la salida completa en el historial
            detalles = f"SE ELIMINÓ COMPLETAMENTE LA SALIDA: ID {id_salida}."
            registrar_historial(conn, current_user.id, "ELIMINACIÓN", "MEDICINAS", detalles)

            # Eliminar la salida completa de salidas_medicinas
            cursor.execute('''
                            DELETE FROM salidas_medicinas
                            WHERE id_salida=%s
                           ''', (id_salida,))

        conn.commit()
        return jsonify({'success': True})

    except Exception as e:
        conn.rollback()
        print(f"Error: {e}")
        return jsonify({'success': False, 'message': str(e)})

    finally:
        cursor.close()
        conn.close()

@app.route('/medicinas/reportes/reporte_salidas_pdf', methods=['POST'])
def medicinas_salidas_pdf():
    data = request.get_json()
    fecha_inicio = data.get('fechaInicio')
    fecha_fin = data.get('fechaFin')
    search_input = data.get('searchInput', '').upper()  # Convertir a mayúsculas

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT
            salidas_medicinas.id_salida,
            DATE_FORMAT(salidas_medicinas.fecha, '%d/%m/%Y') as fecha,
            salidas_medicinas.destino,  -- Agregamos el destino
            med.descripcion,
            med.marca,
            med.dosis,
            aux_ent.cantidad,
            med.unidad
        FROM
            salidas_medicinas
        JOIN
            aux_salidas_medicinas aux_ent ON salidas_medicinas.id_salida = aux_ent.id_salida
        JOIN
            medicinas med ON aux_ent.id_medicina = med.id_medicina
        ORDER BY
            salidas_medicinas.id_salida ASC, salidas_medicinas.fecha ASC
    ''')

    rows = cursor.fetchall()
    conn.close()

    # Procesar los datos para agrupar por id_salida, fecha y destino
    grouped_data = {}
    for row in rows:
        id_salida = row[0]
        fecha = row[1]
        destino = row[2]  # Nuevo campo destino
        if id_salida not in grouped_data:
            grouped_data[id_salida] = {
                'id_salida': id_salida,
                'fecha': fecha,
                'destino': destino,  # Agregamos el destino al grupo
                'salidas': []
            }
        grouped_data[id_salida]['salidas'].append({
            'medicina': row[3],
            'marca': row[4],
            'dosis': row[5],
            'unidad': row[7],
            'cantidad': row[6]
        })

    # Convertir el diccionario a una lista para facilitar el manejo en la plantilla
    final_data = []
    for idx, (id_salida, details) in enumerate(grouped_data.items(), start=1):
        final_data.append({
            'index': idx,
            'id_salida': details['id_salida'],
            'fecha': details['fecha'],
            'destino': details['destino'],  # Pasamos el destino a la plantilla
            'rowspan': len(details['salidas']),
            'salidas': details['salidas']
        })

    # Renderizar el template con los datos agrupados
    rendered = render_template('medicinas/reportes/reporte_salidas.html', datos=final_data)

    # Generar el PDF con WeasyPrint
    html = HTML(string=rendered)
    pdf = html.write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=reporte_salidas.pdf'

    return response












@app.route('/medicinas/agregar_entrada_BD', methods=['POST'])
@login_required
def medicinas_agregar_entrada_BD():
    # try:
    fecha = request.form['txtFecha']
    medicinas = [medicina.upper() for medicina in request.form.getlist('medicina[]')]
    marcas = [marca.upper() for marca in request.form.getlist('marca[]')]
    dosis = [dosis.upper() for dosis in request.form.getlist('dosis[]')]
    cantidades = request.form.getlist('cantidad[]')
    unidades = [unidad.upper() for unidad in request.form.getlist('unidad[]')]

    conn = mysql.connect()
    cursor = conn.cursor()

    # Obtener el último ID de entrada
    cursor.execute("SELECT id_entrada FROM entradas_medicinas ORDER BY id_entrada DESC LIMIT 1")
    last_id = cursor.fetchone()

    if last_id:
        last_id_num = int(last_id[0].split('-')[1])
        new_id_num = last_id_num + 1
    else:
        new_id_num = 1

    new_id_entrada = f"EN-{new_id_num:04d}"

    # Insertar en la tabla de entradas_medicinas con el nuevo ID
    cursor.execute("INSERT INTO entradas_medicinas (id_entrada, fecha) VALUES (%s, %s)", (new_id_entrada, fecha))

    _id_usuario = current_user.id
    _accion = "REGISTRO"

    for medicina, marca, dosis, cantidad, unidad in zip(medicinas, marcas, dosis, cantidades, unidades):

        # Obtener el ID del artículo directamente, asumiendo que ya existe en la tabla `articulos`
        cursor.execute("SELECT id_medicina FROM medicinas WHERE descripcion = %s AND unidad = %s AND marca = %s AND dosis =%s", (medicina, unidad, marca, dosis))
        result = cursor.fetchone()

        if result:
            id_medicina = result[0]
        else:
            # Si no existe el artículo, crearlo
            prefijo = "MED"
            cursor.execute("SELECT id_medicina FROM medicinas WHERE id_medicina LIKE %s ORDER BY id_medicina DESC LIMIT 1", (f"{prefijo}-%",))
            last_id_medicina = cursor.fetchone()

            if last_id_medicina:
                last_number = int(last_id_medicina[0].split('-')[1])
                new_number = last_number + 1
            else:
                new_number = 1

            # Generar el nuevo código para el artículo
            id_medicina = f"{prefijo}-{new_number:03}"
            cursor.execute("INSERT INTO medicinas (id_medicina, descripcion, marca, dosis, unidad) VALUES (%s, %s, %s, %s, %s)", (id_medicina, medicina, marca, dosis, unidad))

        # Insertar en la tabla aux_entradas_medicinas
        cursor.execute("INSERT INTO aux_entradas_medicinas (id_entrada, id_medicina, cantidad) VALUES (%s, %s, %s)", (new_id_entrada, id_medicina, cantidad))

        # Actualizar la tabla de stock_medicinas
        cursor.execute("SELECT cantidad FROM stock_medicinas WHERE id_medicina = %s", (id_medicina,))
        stock_result = cursor.fetchone()

        if stock_result:
            # Si ya existe, actualizar la cantidad sumando la nueva entrada
            nueva_cantidad = stock_result[0] + int(cantidad)
            cursor.execute("UPDATE stock_medicinas SET cantidad = %s WHERE id_medicina = %s", (nueva_cantidad, id_medicina))
        else:
            # Si no existe, insertarlo con la nueva cantidad
            cursor.execute("INSERT INTO stock_medicinas (id_medicina, cantidad) VALUES (%s, %s)", (id_medicina, cantidad))

        # Registrar la acción en el historial
        _area = "medicinas"
        _detalles = f'AGREGADA {cantidad} {unidad} DE "{medicina}", {marca}, {dosis}'
        registrar_historial(conn, _id_usuario, _accion, _area, _detalles)

    conn.commit()
    flash('¡Entradas agregadas y stock actualizado exitosamente!', 'success')
    # except Exception as e:
    #     conn.rollback()
        # flash(f'Error al agregar entradas: {str(e)}', 'error')
    # finally:
    #     cursor.close()
    #     conn.close()

    return redirect('/medicinas/entradas')


@app.route('/medicinas/editar_articulo/<string:id_articulo>')
def editar_articulo_medicinas(id_articulo):

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM articulos WHERE id_articulo = %s", (id_articulo))

    articulo = cursor.fetchone()

    return render_template('medicinas/editar_articulo.html', articulo = articulo)

@app.route('/medicinas/editar_articulo_BD', methods=['POST'])
def medicinas_editar_articulo_BD():

    try:

        id_articulo = request.form['txtID']
        _articulo = request.form["articulo[]"].upper()
        _unidad = request.form["unidad[]"].upper()
        _area = request.form["area[]"].upper()

        conn = mysql.connect()
        cursor = conn.cursor()

        # Actualizar la fecha en la tabla de entradas_medicinas
        cursor.execute("UPDATE articulos SET articulo = %s, unidad =%s, area =%s WHERE id_articulo = %s", (_articulo, _unidad, _area, id_articulo))

        conn.commit()

        flash('¡Articulo editado exitosamente!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Error al actualizar el articulo: {str(e)}', 'error')

    finally:
        cursor.close()
        conn.close()

    return redirect('/medicinas/inventario')

@app.route('/medicinas/editar_entrada/<string:id_entrada>')
def editar_entrada_medicinas(id_entrada):

    conn = mysql.connect()
    cursor = conn.cursor()

    # Vista para el usuario director (o administrador)
    cursor.execute("SELECT * FROM medicinas ORDER BY descripcion")
    articulos = cursor.fetchall()

    cursor.execute("""
        SELECT
            ent.id_entrada,
            med.descripcion,
            med.marca,
            med.dosis,
            med.unidad,
            aux_ent.cantidad,
            ent.fecha,
            med.id_medicina
        FROM medicinas med
        LEFT JOIN aux_entradas_medicinas aux_ent
            ON med.id_medicina = aux_ent.id_medicina
        LEFT JOIN entradas_medicinas ent
            ON aux_ent.id_entrada = ent.id_entrada
        WHERE aux_ent.id_entrada = %s
        ORDER BY med.descripcion
    """, (id_entrada,))

    entradas = cursor.fetchall()

    # Convertir los resultados de 'entradas' en una lista de diccionarios
    entradas_dict = [
        {
            'id_entrada': entrada[0],
            'descripcion': entrada[1],
            'marca': entrada[2],
            'dosis': entrada[3],
            'unidad': entrada[4],
            'cantidad': entrada[5],
            'fecha': entrada[6],
            'id_medicina': entrada[7]
        }
        for entrada in entradas
    ]

    return render_template('medicinas/editar_entrada.html', medicinas=articulos, entradas=entradas_dict)

def Actualizar_Stock_Medicinas():
    try:
        conn = mysql.connect()
        cursor = conn.cursor()

        # Obtener todas las medicinas en stock_medicinas
        cursor.execute("SELECT id_medicina FROM stock_medicinas")
        todas_las_medicinas = {row[0] for row in cursor.fetchall()}

        # Obtener las entradas totales
        cursor.execute("""
            SELECT id_medicina, SUM(cantidad) AS total_entradas
            FROM aux_entradas_medicinas
            GROUP BY id_medicina
        """)
        entradas = {row[0]: row[1] for row in cursor.fetchall()}
        print("Entradas es: ", entradas)

        # Obtener las salidas totales
        cursor.execute("""
            SELECT id_medicina, SUM(cantidad) AS total_salidas
            FROM aux_salidas_medicinas
            GROUP BY id_medicina
        """)
        salidas = {row[0]: row[1] for row in cursor.fetchall()}
        print("Salidas es: ", salidas)

        # Actualizar el stock_medicinas
        for id_medicina in todas_las_medicinas:
            total_entrada = entradas.get(id_medicina, 0)
            total_salida = salidas.get(id_medicina, 0)

            # Calcular nuevo stock
            nuevo_stock = total_entrada - total_salida

            # Si no hay entradas ni salidas, establecer stock a 0
            if total_entrada == 0 and total_salida == 0:
                nuevo_stock = 0

            # Actualizar la tabla stock_medicinas
            cursor.execute("""
                UPDATE stock_medicinas
                SET cantidad = %s
                WHERE id_medicina = %s
            """, (nuevo_stock, id_medicina))

        conn.commit()
        flash('¡Stock de medicamentos actualizado exitosamente!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error al actualizar el stock: {str(e)}', 'error')
    finally:
        cursor.close()
        conn.close()



@app.route('/medicinas/editar_entrada_BD', methods=['POST'])
def medicinas_editar_entrada_BD():
    # try:
        id_entrada = request.form['txtID']
        fecha = request.form['txtFecha']
        medicinas = [medicina.upper() for medicina in request.form.getlist('medicina[]')]
        marcas = [marca.upper() for marca in request.form.getlist('marca[]')]
        dosis = [dosis.upper() for dosis in request.form.getlist('dosis[]')]
        cantidades = request.form.getlist('cantidad[]')
        unidades = [unidad.upper() for unidad in request.form.getlist('unidad[]')]

        conn = mysql.connect()
        cursor = conn.cursor()

        # Actualizar la fecha en la tabla de entradas_medicinas
        cursor.execute("UPDATE entradas_medicinas SET fecha = %s WHERE id_entrada = %s", (fecha, id_entrada))

        # Obtener los artículos existentes en la entrada
        cursor.execute("SELECT id_medicina, cantidad FROM aux_entradas_medicinas WHERE id_entrada = %s", (id_entrada,))
        existing_entries = cursor.fetchall()
        existing_articles = {row[0]: row[1] for row in existing_entries}

        new_articles = set()
        for medicina, marca, dosis, cantidad, unidad in zip(medicinas, marcas, dosis, cantidades, unidades):

            # Obtener el ID del artículo directamente, asumiendo que ya existe en la tabla `articulos`
            cursor.execute("SELECT id_medicina FROM medicinas WHERE descripcion = %s AND unidad = %s AND marca =%s AND dosis = %s", (medicina, unidad, marca, dosis))
            result = cursor.fetchone()

            if result:
                id_medicina = result[0]
                new_articles.add(id_medicina)
            else:
                # Si no existe el artículo, crearlo
                prefijo = "MED"
                cursor.execute("SELECT id_medicina FROM medicinas WHERE id_medicina LIKE %s ORDER BY id_medicina DESC LIMIT 1", (f"{prefijo}-%",))
                last_id_medicina = cursor.fetchone()

                if last_id_medicina:
                    last_number = int(last_id_medicina[0].split('-')[1])
                    new_number = last_number + 1
                else:
                    new_number = 1

                # Generar el nuevo código para el artículo
                id_medicina = f"{prefijo}-{new_number:03}"
                cursor.execute("INSERT INTO medicinas (id_medicina, descripcion, marca, dosis, unidad) VALUES (%s, %s, %s, %s, %s)", (id_medicina, medicina, marca, dosis, unidad))

            cantidad = int(cantidad)

            # Verificar si el artículo ya existe en la entrada

            if id_medicina in existing_articles:
                old_cantidad = existing_articles[id_medicina]
                print("Old cantidad es: ", old_cantidad)
                diferencia = cantidad - old_cantidad
                print("Diferencia es: ", diferencia)

                # Actualizar la cantidad en aux_entradas_medicinas
                cursor.execute("UPDATE aux_entradas_medicinas SET cantidad = %s WHERE id_entrada = %s AND id_medicina = %s", (cantidad, id_entrada, id_medicina))

                # Actualizar la tabla stock_medicinas
                # cursor.execute("UPDATE stock_medicinas SET cantidad = cantidad + %s WHERE id_medicina = %s", (diferencia, id_medicina))
            else:
                # Si no existe, insertarlo en aux_entradas_medicinas
                print(id_entrada, id_medicina, cantidad )
                cursor.execute("INSERT INTO aux_entradas_medicinas (id_entrada, id_medicina, cantidad) VALUES (%s, %s, %s)", (id_entrada, id_medicina, cantidad))

                # Actualizar el stock aumentando la cantidad
                # cursor.execute("UPDATE stock_medicinas SET cantidad = cantidad + %s WHERE id_medicina= %s", (cantidad, id_medicina))

        # Eliminar artículos que ya no están en la lista
        articles_to_remove = set(existing_articles.keys()) - new_articles
        for art_id in articles_to_remove:
            old_cantidad = existing_articles[art_id]

            # Reducir el stock por la cantidad eliminada
            # cursor.execute("UPDATE stock_medicinas SET cantidad = cantidad - %s WHERE id_medicina = %s", (old_cantidad, art_id))

            # Eliminar el artículo de aux_entradas_medicinas
            cursor.execute("DELETE FROM aux_entradas_medicinas WHERE id_entrada = %s AND id_medicina = %s", (id_entrada, art_id))



        conn.commit()

        # Registrar la acción en el historial
        for medicina, marca, dosis, cantidad, unidad in zip(medicinas, marcas, dosis, cantidades, unidades):
            detalles = f"SE ACTUALIZÓ {cantidad} {unidad} DE {medicina} EN LA ENTRADA {id_entrada} EL {fecha}"
            registrar_historial(conn, current_user.id, "EDICIÓN", "MEDICINAS", detalles)

        flash('¡Entrada actualizada exitosamente!', 'success')


        Actualizar_Stock_Medicinas()

    # except Exception as e:
    #     conn.rollback()
    #     flash(f'Error al actualizar la entrada: {str(e)}', 'error')
    # finally:
        # conn.rollback()
        cursor.close()
        conn.close()

        return redirect('/medicinas/entradas')







from datetime import datetime

def registrar_historial(conn, id_usuario, accion, area, detalles):
    cursor = conn.cursor()
    ahora = datetime.now()
    hora = ahora.strftime('%H:%M:%S')  # Formato de 24 horas para almacenamiento
    fecha_actual = ahora.strftime('%Y-%m-%d')

    sql_historial = '''
        INSERT INTO historial
        (id_historial, id_usuario, accion, area, detalles, hora, fecha)
        VALUES (NULL, %s, %s, %s, %s, %s, %s)
    '''
    datos_historial = (id_usuario, accion, area, detalles, hora, fecha_actual)
    cursor.execute(sql_historial, datos_historial)
    conn.commit()


# Registrar la bomba en la base de datos
@app.route('/agregar_bomba_BD', methods=['POST'])
def agregar_bomba_BD():

    _fecha = request.form['txtFecha']
    _tipo = request.form['txtTipo'].upper()
    _solicitado = request.form['txtSolicitud'].upper()
    _municipio = request.form['txtMunicipio']
    _parroquia = request.form['txtParroquia']
    _sector = request.form['txtSector'].upper()
    _cantidad = int(request.form['txtCantidad'])
    _fuerza = f"{request.form['txtFuerza']} HP"
    _voltios = request.form['txtVoltios']

    try:
        conn = mysql.connect()
        cursor = conn.cursor()

        sql = '''
            INSERT INTO bombas_entregadas (ID, sector, fecha, solicitado, cantidad, fuerza, tipo, voltios, municipio, parroquia)
            VALUES (NULL, %s, %s, %s, %s, %s, %s, %s, %s, %s);
        '''
        datos = (_sector, _fecha, _solicitado, _cantidad, _fuerza, _tipo, _voltios, _municipio, _parroquia)
        cursor.execute(sql, datos)
        conn.commit()

        _id_usuario = current_user.id
        _area = "BOMBAS"
        _accion = "REGISTRO"
        _detalles = f"SE REGISTRÓ {_cantidad} BOMBA(S) DE {_fuerza} {_tipo} DE {_voltios} PARA {_municipio}, SECTOR: {_sector}, \n FECHA DE ENTREGA: {_fecha}"

        registrar_historial(conn, _id_usuario, _accion, _area, _detalles)

        flash('¡Entrega agregada exitosamente!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Ocurrió un error: {str(e)}', 'error')

    finally:
        cursor.close()
        conn.close()

    return redirect('/entregas')


#Formulario de Edicion de bombas
@app.route('/editar_bomba/<int:id>')
def editar_bomba(id):

    conn = mysql.connect()
    cursor = conn.cursor()

    # Consulta para obtener los datos de avances de un niño especifico
    cursor.execute("SELECT * FROM bombas_entregadas WHERE ID=%s",(id))
    bomba = cursor.fetchall()

    # Consulta para obtener los municipios
    cursor.execute("SELECT * FROM municipios ORDER BY nombre")
    municipios = cursor.fetchall()

    conn.commit()

    return render_template("bombas/editar_bomba.html", bomba = bomba, municipios = municipios)

@app.route('/editar_bomba_BD', methods=['POST'])
def editar_bomba_BD():

    _id = request.form['txtID']
    _fecha = request.form['txtFecha']
    _tipo = request.form['txtTipo'].upper()
    _solicitado = request.form['txtSolicitud'].upper()
    _municipio = request.form['txtMunicipio']
    _parroquia = request.form['txtParroquia']
    _sector = request.form['txtSector'].upper()
    _cantidad = int(request.form['txtCantidad'])
    _fuerza = f"{request.form['txtFuerza']} HP"
    _voltios = request.form['txtVoltios']

    try:
        conn = mysql.connect()
        cursor = conn.cursor()

        cursor.execute("SELECT cantidad, tipo, fuerza, voltios, municipio, sector FROM bombas_entregadas WHERE ID = %s", (_id,))
        bomba_anterior = cursor.fetchone()

        sql = '''
            UPDATE bombas_entregadas
            SET sector = %s, fecha = %s, solicitado = %s, cantidad = %s, fuerza = %s, tipo = %s, voltios = %s, municipio = %s, parroquia = %s
            WHERE ID = %s;
        '''
        datos = (_sector, _fecha, _solicitado, _cantidad, _fuerza, _tipo, _voltios, _municipio, _parroquia, _id)
        cursor.execute(sql, datos)
        conn.commit()

        _id_usuario = current_user.id
        _area = "BOMBAS"
        _accion = "ACTUALIZACION"

        if bomba_anterior:
            detalles_anterior = f"{bomba_anterior[0]} BOMBA(S) {bomba_anterior[1]} DE {bomba_anterior[2]} DE {bomba_anterior[3]} PARA {bomba_anterior[4]}, SECTOR: {bomba_anterior[5]}"
        else:
            detalles_anterior = "No se encontró información anterior."

        detalles_nuevo = f"{_cantidad} BOMBA(S) {_tipo} DE {_fuerza} DE {_voltios} PARA {_municipio}, SECTOR: {_sector}"
        detalles = f"ANTES: {detalles_anterior} \n DESPUÉS: {detalles_nuevo}"

        registrar_historial(conn, _id_usuario, _accion, _area, detalles)

        flash('¡Bomba actualizada exitosamente!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Ocurrió un error: {str(e)}', 'error')

    finally:
        cursor.close()
        conn.close()

    return redirect('/entregas')



@app.route('/eliminar_bomba/<int:id>')
@login_required
def eliminar_bomba(id):
    conn = mysql.connect()
    cursor = conn.cursor()

    try:
        # Obtener información de la bomba antes de eliminarla (para el registro)
        cursor.execute("SELECT sector, fecha, solicitado, cantidad, tipo, fuerza, voltios, municipio, parroquia FROM bombas_entregadas WHERE ID = %s", (id,))
        bomba = cursor.fetchone()

        # Eliminar la bomba de la base de datos
        cursor.execute("DELETE FROM bombas_entregadas WHERE ID = %s", (id,))
        conn.commit()

        # Registrar la acción en el historial de acciones
        _id_usuario = current_user.id
        _area = "BOMBAS"
        _accion = "ELIMINACION"
        _detalles = f"SE ELIMINÓ {bomba[3]} BOMBA(S) DE {bomba[5]} {bomba[4]} DE {bomba[6]} PARA {bomba[7]}, SECTOR: {bomba[0]}, FECHA DE ENTREGA: {bomba[1]}"

        # Usar la función registrar_historial
        registrar_historial(conn, _id_usuario, _accion, _area, _detalles)

        flash('¡La bomba fue eliminada exitosamente!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Ocurrió un error al eliminar la bomba: {str(e)}', 'error')

    finally:
        cursor.close()
        conn.close()

    return redirect('/entregas')


@app.route('/eliminar_multiples_bombas', methods=['POST'])
@login_required
def eliminar_multiples_bombas():
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify(success=False, message='No IDs provided'), 400

    try:
        conn = mysql.connect()
        cursor = conn.cursor()

        # Obtener información de las bombas antes de eliminarlas (para el registro)
        cursor.execute("SELECT ID, cantidad, tipo, fuerza, voltios, municipio, sector FROM bombas_entregadas WHERE ID IN %s", (tuple(ids),))
        bombas = cursor.fetchall()

        # Eliminar las bombas
        format_strings = ','.join(['%s'] * len(ids))
        query = f"DELETE FROM bombas_entregadas WHERE ID IN ({format_strings})"
        cursor.execute(query, ids)

        conn.commit()

        # Registrar cada eliminación en el historial de acciones
        _id_usuario = current_user.id
        _area = "BOMBAS"
        _accion = "ELIMINACION"

        for bomba in bombas:
            cantidad = bomba[1]
            tipo = bomba[2]
            fuerza = bomba[3]
            voltios = bomba[4]
            municipio = bomba[5]
            sector = bomba[6]

            _detalles = f"SE ELIMINÓ {cantidad} BOMBA(S) {tipo} DE {fuerza} DE {voltios} PARA {municipio}, SECTOR: {sector}"

            # Usar la función registrar_historial
            registrar_historial(conn, _id_usuario, _accion, _area, _detalles)

        return jsonify(success=True)

    except Exception as e:
        conn.rollback()
        return jsonify(success=False, message=str(e)), 500

    finally:
        cursor.close()
        conn.close()

@app.route('/agregar_municipio')
def agregar_municipio():
    return render_template('general/agregar_municipio.html')

@app.route('/agregar_parroquia')
def agregar_parroquia():

    conn = mysql.connect()
    cursor = conn.cursor()

    #Vista para el usuario director (o administrador)

    cursor.execute("SELECT * FROM municipios ")

    municipios = cursor.fetchall()

    conn.commit()

    return render_template('general/agregar_parroquia.html',municipios=municipios)

@app.route('/agregar_sector')
def agregar_sector():

    conn = mysql.connect()
    cursor = conn.cursor()

    #Vista para el usuario director (o administrador)

    cursor.execute("SELECT * FROM municipios ")

    municipios = cursor.fetchall()

    cursor.execute("SELECT * FROM parroquias ")

    parroquias = cursor.fetchall()

    conn.commit()

    return render_template('general/agregar_sector.html',municipios=municipios, parroquias=parroquias)


#Funcion para eliminar municipio
@app.route('/eliminar_municipio/<string:id>')
def eliminar_municipio(id):

    conn = mysql.connect()
    cursor = conn.cursor()

    # Eliminar el municipio de la base de datos
    cursor.execute("DELETE FROM municipios WHERE cod_mun = %s", (id,))

    conn.commit()

    flash('¡El municipio fue eliminado exitosamente!')

    return redirect('/municipios')

#Funcion para actualizar niño después de su edición
@app.route('/actualizar_municipio', methods=['POST'])
def actualizar_municipio():

    _nombre = request.form['txtNombre']
    _cod_mun = request.form['txtID']

    sql = "UPDATE municipios SET nombre = %s WHERE cod_mun = %s;"

    datos=(_nombre, _cod_mun)


    conn = mysql.connect()
    cursor = conn.cursor()
    cursor.execute(sql,datos)
    conn.commit()

    return redirect('/municipios')

@app.route('/bombas_excel', methods=['POST'])
def bombas_excel():
    data = request.get_json()
    fecha_inicio = data.get('fechaInicio')
    fecha_fin = data.get('fechaFin')
    search_input = data.get('searchInput', '').upper()

    conn = mysql.connect()
    cur = conn.cursor()

    # Consulta SQL para las bombas entregadas
    query = '''
        SELECT
            fecha,
            solicitado,
            cantidad,
            fuerza,
            tipo,
            voltios,
            municipio,
            parroquia,
            sector
        FROM bombas_entregadas
        WHERE fecha BETWEEN %s AND %s
        AND (UPPER(solicitado) LIKE %s
        OR UPPER(cantidad) LIKE %s
        OR UPPER(fuerza) LIKE %s
        OR UPPER(tipo) LIKE %s
        OR UPPER(voltios) LIKE %s
        OR UPPER(municipio) LIKE %s
        OR UPPER(parroquia) LIKE %s
        OR UPPER(sector) LIKE %s)
    '''

    like_term = f"%{search_input}%"
    cur.execute(query, (fecha_inicio, fecha_fin, like_term, like_term, like_term, like_term, like_term, like_term, like_term, like_term))
    rows = cur.fetchall()

    # Crear un archivo Excel en memoria
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Bombas Entregadas"

    # Estilo para los encabezados
    header_font = Font(bold=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Agregar el título "Bombas Entregadas" en la primera fila
    sheet.merge_cells('A1:I1')
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "Bombas Entregadas"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = alignment_center

    # Crear encabezados
    headers = ['Fecha', 'Solicitado por', 'Cantidad', 'Fuerza', 'Tipo', 'Voltios', 'Municipio', 'Parroquia', 'Sector']
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = border_style
        cell.fill = header_fill

    # Rellenar datos a partir de la tercera fila
    row_num = 3
    total_cantidad = 0
    for row in rows:
        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            if col_num == 1:  # Formatear la fecha en la columna 1 (Fecha)
                if value:
                    cell.value = value.strftime('%d/%m/%Y')  # Formato dd/mm/yyyy
            else:
                cell.value = value
            cell.alignment = alignment_center
            cell.border = border_style
        total_cantidad += row[2] if row[2] else 0  # Sumar la cantidad total
        row_num += 1

    # Agregar total de cantidades
    sheet.cell(row=row_num, column=1, value='TOTAL:')
    sheet.cell(row=row_num, column=3, value=total_cantidad)
    sheet.cell(row=row_num, column=1).font = Font(bold=True)
    sheet.cell(row=row_num, column=3).font = Font(bold=True)
    sheet.cell(row=row_num, column=1).alignment = alignment_center
    sheet.cell(row=row_num, column=3).alignment = alignment_center



    # Ajustar el ancho de las columnas automáticamente
    for col in sheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Usar get_column_letter para obtener la letra de la columna
        for cell in col:
            if not isinstance(cell, MergedCell):  # Verifica que no sea una celda fusionada
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    workbook.save(output)
    output.seek(0)

    # Asignar un nombre específico al archivo Excel
    filename = "Bombas_Entregadas.xlsx"

    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

#PDF de las bombas entregadas
@app.route('/bombas_pdf', methods=['POST'])
def bombas_pdf():
    data = request.get_json()
    fecha_inicio = data.get('fechaInicio')
    fecha_fin = data.get('fechaFin')
    search_input = data.get('searchInput', '').upper()  # Convertir a mayúsculas para coincidencias insensibles a mayúsculas/minúsculas

    conn = mysql.connect()
    cur = conn.cursor()

    # Incluir el término de búsqueda en la consulta SQL
    query = ''' SELECT
                fecha,
                solicitado,
                cantidad,
                fuerza,
                tipo,
                voltios,
                municipio,
                parroquia,
                sector
                FROM bombas_entregadas
                WHERE fecha BETWEEN %s AND %s
                AND (UPPER(solicitado) LIKE %s
                OR UPPER(cantidad) LIKE %s
                OR UPPER(fuerza) LIKE %s
                OR UPPER(tipo) LIKE %s
                OR UPPER(voltios) LIKE %s
                OR UPPER(municipio) LIKE %s
                OR UPPER(parroquia) LIKE %s
                OR UPPER(sector) LIKE %s) '''

    like_term = f"%{search_input}%"
    cur.execute(query, (fecha_inicio, fecha_fin, like_term, like_term, like_term, like_term, like_term, like_term, like_term, like_term))
    rows = cur.fetchall()

    # Crear un índice consecutivo para los datos
    indexed_rows = [(i+1,) + row for i, row in enumerate(rows)]

    # Rutas relativas a los archivos CSS, encabezado y pie de página
    css_path = os.path.join(os.path.dirname(__file__), 'static', 'css', 'style2.css')
    header_path = os.path.join(os.path.dirname(__file__), 'static', 'bombas','header_bombas.html')
    footer_path = os.path.join(os.path.dirname(__file__), 'static', 'bombas','footer_bombas.html')

    # Renderizar el template con los datos indexados
    rendered = render_template('bombas/reporte_bombas.html', datos=indexed_rows)

    pdf = pdfkit.from_string(rendered, False, css=css_path, options={
        "enable-local-file-access": "",
        "margin-top": "20mm",
        "margin-right": "12mm",
        "margin-bottom": "28mm",
        "margin-left": "10mm",
        "header-html": header_path,
        "footer-html": footer_path,
        "header-spacing": "4",
        "footer-spacing": "4",
        "zoom": "0.9"
    })

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=output.pdf'

    return response

@app.route('/inventario_pdf', methods=['POST'])
def inventario_pdf():
    data = request.get_json()
    search_input = data.get('searchInput', '').upper()  # Convertir a mayúsculas para coincidencias insensibles a mayúsculas/minúsculas

    conn = mysql.connect()
    cur = conn.cursor()

    # Incluir el término de búsqueda en la consulta SQL
    query = '''
        SELECT
            stock.id_articulo,
            articulos.articulo,
            stock.cantidad,
            articulos.unidad,
            articulos.area
        FROM
            stock_inventario stock
        JOIN
            articulos ON stock.id_articulo = articulos.id_articulo
        WHERE
            stock.cantidad > 0
            AND (UPPER(stock.id_articulo) LIKE %s
            OR UPPER(articulos.articulo) LIKE %s
            OR UPPER(stock.cantidad) LIKE %s
            OR UPPER(articulos.unidad) LIKE %s
            OR UPPER(articulos.area) LIKE %s)
        ORDER BY
            articulos.articulo ASC '''

    like_term = f"%{search_input}%"
    cur.execute(query, (like_term, like_term, like_term, like_term, like_term))
    rows = cur.fetchall()

    # Crear un índice consecutivo para los datos
    indexed_rows = [(i+1,) + row for i, row in enumerate(rows)]

    # Rutas relativas a los archivos CSS, encabezado y pie de página
    css_path = os.path.join(os.path.dirname(__file__), 'static', 'css', 'style2.css')
    header_path = os.path.join(os.path.dirname(__file__), 'templates', 'ferreteria', 'header_inventario.html')
    footer_path = os.path.join(os.path.dirname(__file__), 'static', 'footer_bombas.html')

    # Renderizar el template con los datos indexados
    rendered = render_template('ferreteria/reporte_inventario.html', datos=indexed_rows)

    pdf = pdfkit.from_string(rendered, False, css=css_path, options={
        "enable-local-file-access": "",
        "margin-top": "20mm",
        "margin-right": "12mm",
        "margin-bottom": "28mm",
        "margin-left": "10mm",
        "header-html": header_path,
        "footer-html": footer_path,
        "header-spacing": "4",
        "footer-spacing": "4",
        "zoom": "0.9"
    })

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=inventario.pdf'

    return response

@app.route('/entradas_pdf', methods=['POST'])
def entradas_pdf():
    data = request.get_json()
    fecha_inicio = data.get('fechaInicio')
    fecha_fin = data.get('fechaFin')
    search_input = data.get('searchInput', '').upper()  # Convertir a mayúsculas para coincidencias insensibles a mayúsculas/minúsculas

    conn = mysql.connect()
    cur = conn.cursor()

    # Incluir el término de búsqueda en la consulta SQL
    query = '''
                SELECT
                    ent.id_entrada,
                    ent.fecha,
                    art.articulo,
                    aux_ent.cantidad,
                    art.unidad,
                    art.area,
                    aux_ent.id_articulo
                FROM entradas_inventario ent
                JOIN aux_entradas_inventario aux_ent ON ent.id_entrada = aux_ent.id_entrada
                JOIN articulos art ON aux_ent.id_articulo = art.id_articulo
                WHERE ent.fecha BETWEEN %s AND %s
                AND (UPPER(art.articulo) LIKE %s
                OR UPPER(aux_ent.id_articulo) LIKE %s
                OR UPPER(art.unidad) LIKE %s)
            '''

    like_term = f"%{search_input}%"
    cur.execute(query, (fecha_inicio, fecha_fin, like_term, like_term, like_term))
    rows = cur.fetchall()

    # Crear un índice consecutivo para los datos
    indexed_rows = [(i+1,) + row for i, row in enumerate(rows)]

    # Rutas relativas a los archivos CSS, encabezado y pie de página
    css_path = os.path.join(os.path.dirname(__file__), 'static', 'css', 'style2.css')
    header_path = os.path.join(os.path.dirname(__file__), 'static', 'header_bombas.html')
    footer_path = os.path.join(os.path.dirname(__file__), 'static', 'footer_bombas.html')

    # Renderizar el template con los datos indexados
    rendered = render_template('ferreteria/reporte_entradas.html', datos=indexed_rows)

    pdf = pdfkit.from_string(rendered, False, css=css_path, options={
        "enable-local-file-access": "",
        "margin-top": "20mm",
        "margin-right": "12mm",
        "margin-bottom": "28mm",
        "margin-left": "10mm",
        "header-html": header_path,
        "footer-html": footer_path,
        "header-spacing": "4",
        "footer-spacing": "4",
        "zoom": "0.9"
    })

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=output.pdf'

    return response
@app.route('/salidas_pdf', methods=['POST'])
def salidas_pdf():
    data = request.get_json()
    fecha_inicio = data.get('fechaInicio')
    fecha_fin = data.get('fechaFin')
    search_input = data.get('searchInput', '').upper()

    conn = mysql.connect()
    cursor = conn.cursor()

    query_normal = '''
        SELECT
            aux_sal.id,
            salidas_inventario.fecha,
            art.articulo,
            aux_sal.cantidad,
            art.unidad,
            art.area,
            salidas_inventario.id_salida,
            salidas_inventario.destino
        FROM
            salidas_inventario
        JOIN
            aux_salidas_inventario aux_sal ON salidas_inventario.id_salida = aux_sal.id_salida
        JOIN
            articulos art ON aux_sal.id_articulo = art.id_articulo
        WHERE
            (salidas_inventario.fecha BETWEEN %s AND %s)
            AND (UPPER(art.articulo) LIKE %s
            OR UPPER(aux_sal.cantidad) LIKE %s
            OR UPPER(art.unidad) LIKE %s
            OR UPPER(art.area) LIKE %s
            OR UPPER(salidas_inventario.destino) LIKE %s)
        ORDER BY
            salidas_inventario.id_salida ASC, salidas_inventario.fecha ASC
    '''

    query_extra = '''
        SELECT
            artic_extra.id,
            salidas_inventario.fecha,
            artic_extra.articulo,
            artic_extra.cantidad,
            artic_extra.unidad,
            artic_extra.area,
            salidas_inventario.id_salida,
            salidas_inventario.destino
        FROM
            salidas_inventario
        JOIN
            articulos_extra_salida artic_extra ON salidas_inventario.id_salida = artic_extra.id_salida
        WHERE
            (salidas_inventario.fecha BETWEEN %s AND %s)
            AND (UPPER(artic_extra.articulo) LIKE %s
            OR UPPER(artic_extra.cantidad) LIKE %s
            OR UPPER(artic_extra.unidad) LIKE %s
            OR UPPER(artic_extra.area) LIKE %s
            OR UPPER(salidas_inventario.destino) LIKE %s)
        ORDER BY
            salidas_inventario.id_salida ASC, salidas_inventario.fecha ASC
    '''

    like_term = f"%{search_input}%"
    cursor.execute(query_normal, (fecha_inicio, fecha_fin, like_term, like_term, like_term, like_term, like_term))
    salidas_normales = cursor.fetchall()

    cursor.execute(query_extra, (fecha_inicio, fecha_fin, like_term, like_term, like_term, like_term, like_term))
    salidas_extras = cursor.fetchall()

    salidas = salidas_normales + salidas_extras

    grouped_salidas = []
    current_id_salida = None
    group = []

    for salida in salidas:
        if salida[6] != current_id_salida:
            if group:
                grouped_salidas.append(group)
            group = [salida]
            current_id_salida = salida[6]
        else:
            group.append(salida)

    if group:
        grouped_salidas.append(group)

    cursor.close()
    conn.close()

    # Calcular la cantidad total de artículos ingresados y artículos extras
    total_cantidad = sum(int(salida[3]) for salida in salidas)

    css_path = os.path.join(os.path.dirname(__file__), 'static', 'css', 'style2.css')
    header_path = os.path.join(os.path.dirname(__file__), 'templates', 'ferreteria', 'header_salidas.html')
    footer_path = os.path.join(os.path.dirname(__file__), 'templates', 'ferreteria', 'footer_general.html')

    rendered = render_template('ferreteria/reporte_salidas.html', grouped_salidas=grouped_salidas, cantidad_total=total_cantidad)

    pdf = pdfkit.from_string(rendered, False, css=css_path, options={
        "enable-local-file-access": "",
        "margin-top": "23mm",
        "margin-right": "12mm",
        "margin-bottom": "40mm",
        "margin-left": "10mm",
        "header-html": header_path,
        "footer-html": footer_path,
        "header-spacing": "4",
        "footer-spacing": "4",
        "zoom": "0.9"
    })

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=output.pdf'

    return response



from weasyprint import HTML

# @app.route('/actas_pdf/<string:id_salida>')
# def actas_pdf(id_salida):
#     conn = mysql.connect()
#     cur = conn.cursor()

#     # Consulta SQL filtrando por ID de salida
#     query = '''
#         SELECT
#             salidas_inventario.id_salida,
#             salidas_inventario.fecha,
#             salidas_inventario.destino,
#             articulos.id_articulo,
#             articulos.articulo AS nombre_articulo,
#             articulos.unidad,
#             aux_salidas_inventario.cantidad
#         FROM
#             salidas_inventario
#         JOIN
#             aux_salidas_inventario ON salidas_inventario.id_salida = aux_salidas_inventario.id_salida
#         JOIN
#             articulos ON aux_salidas_inventario.id_articulo = articulos.id_articulo
#         WHERE
#             salidas_inventario.id_salida = %s;
#     '''

#     cur.execute(query, (id_salida,))
#     rows = cur.fetchall()

#     # Renderizar el template con los datos
#     rendered = render_template('reportes/acta_prueba.html', datos=rows)

#     # Generar el PDF con WeasyPrint
#     html = HTML(string=rendered)
#     pdf = html.write_pdf()

#     response = make_response(pdf)
#     response.headers['Content-Type'] = 'application/pdf'
#     response.headers['Content-Disposition'] = 'inline; filename=output.pdf'

#     return response

@app.route('/actas_pdf/<string:id_salida>')
def actas_pdf(id_salida):
    conn = mysql.connect()
    cur = conn.cursor()

    # Consulta SQL para los artículos normales
    query_normal = '''
        SELECT
            salidas_inventario.id_salida,
            salidas_inventario.fecha,
            salidas_inventario.destino,
            articulos.id_articulo,
            articulos.articulo AS nombre_articulo,
            articulos.unidad,
            aux_salidas_inventario.cantidad
        FROM
            salidas_inventario
        JOIN
            aux_salidas_inventario ON salidas_inventario.id_salida = aux_salidas_inventario.id_salida
        JOIN
            articulos ON aux_salidas_inventario.id_articulo = articulos.id_articulo
        WHERE
            salidas_inventario.id_salida = %s;
    '''

    cur.execute(query_normal, (id_salida,))
    rows_normales = cur.fetchall()

    # Consulta SQL para los artículos extras
    query_extra = '''
        SELECT
            salidas_inventario.id_salida,
            salidas_inventario.fecha,
            salidas_inventario.destino,
            articulos_extra_salida.id,
            articulos_extra_salida.articulo AS nombre_articulo,
            articulos_extra_salida.unidad,
            articulos_extra_salida.cantidad
        FROM
            salidas_inventario
        JOIN
            articulos_extra_salida ON salidas_inventario.id_salida = articulos_extra_salida.id_salida
        WHERE
            salidas_inventario.id_salida = %s;
    '''

    cur.execute(query_extra, (id_salida,))
    rows_extras = cur.fetchall()

    # Combinar normales y extras
    all_rows = rows_normales + rows_extras

    # Renderizar el template con los datos
    rendered = render_template('reportes/acta_prueba.html', datos=all_rows)

    # Generar el PDF con WeasyPrint
    html = HTML(string=rendered)
    pdf = html.write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=output.pdf'

    return response



@app.route('/medicinas/actas_pdf/<string:id_salida>')
def medicinas_actas_pdf(id_salida):
    conn = mysql.connect()
    cur = conn.cursor()

    # Consulta SQL filtrando por ID de salida en medicinas
    query = '''
        SELECT
            salidas_medicinas.id_salida,
            salidas_medicinas.fecha,
            salidas_medicinas.destino,
            medicinas.id_medicina,
            medicinas.descripcion AS nombre_medicina,
            medicinas.marca,
            medicinas.dosis,
            medicinas.unidad,
            aux_salidas_medicinas.cantidad
        FROM
            salidas_medicinas
        JOIN
            aux_salidas_medicinas ON salidas_medicinas.id_salida = aux_salidas_medicinas.id_salida
        JOIN
            medicinas ON aux_salidas_medicinas.id_medicina = medicinas.id_medicina
        WHERE
            salidas_medicinas.id_salida = %s;
    '''

    cur.execute(query, (id_salida,))
    rows = cur.fetchall()

    # Renderizar el template con los datos
    rendered = render_template('medicinas/reportes/acta.html', datos=rows)

    # Generar el PDF con WeasyPrint
    html = HTML(string=rendered)
    pdf = html.write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=acta_salida_{id_salida}.pdf'

    return response


import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file

import io
import xlsxwriter
from flask import send_file
from datetime import datetime

@app.route('/actas_excel/<string:id_salida>')
def actas_excel(id_salida):
    conn = mysql.connect()
    cur = conn.cursor()

    # Consulta SQL para obtener los datos
    query = '''
        SELECT
            salidas_inventario.id_salida,
            salidas_inventario.fecha,
            salidas_inventario.destino,
            articulos.id_articulo,
            articulos.articulo AS nombre_articulo,
            articulos.unidad,
            aux_salidas_inventario.cantidad
        FROM
            salidas_inventario
        JOIN
            aux_salidas_inventario ON salidas_inventario.id_salida = aux_salidas_inventario.id_salida
        JOIN
            articulos ON aux_salidas_inventario.id_articulo = articulos.id_articulo
        WHERE
            salidas_inventario.id_salida = %s;
    '''
    cur.execute(query, (id_salida,))
    rows = cur.fetchall()

    # Crear el archivo Excel en memoria
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Acta de Entrega')

    # 1) Configurar encabezado con imagen en el área de impresión
    worksheet.set_header('&L&G', {'image_left': 'encabezado_mejorado.png'})

    # 2) Configurar pie de página con imagen en el área de impresión
    worksheet.set_footer('&L&G', {'image_left': 'pie_de_pagina_mejorado.png'})

    # Estilos
    bold_centered = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'font_size': 18})  # Fuente más grande
    normal_centered = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'font_size': 14})  # Ajuste de tamaño de letra
    fecha = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'font_size': 11})  # Ajuste de tamaño de letra
    table_header_format = workbook.add_format({'bold': True, 'font_size': 14, 'border': 1, 'align': 'center'})  # Estilo para encabezado de tabla
    table_data_format = workbook.add_format({'font_size': 12, 'border': 1, 'text_wrap': True})  # Estilo para las celdas de datos con ajuste de texto
    destination_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'font_size': 14})  # Formato para el destino

    # Agregar tres filas vacías
    worksheet.write_blank('A1', None)
    worksheet.write_blank('A2', None)
    worksheet.write_blank('A3', None)

    # Agregar fecha en la siguiente fila
    worksheet.write('C3', f'Maturín, {rows[0][1].strftime("%d/%m/%Y")}', fecha)

    # Encabezado del acta centrado en la primera fila
    worksheet.merge_range('A5:D5', 'ACTA DE ENTREGA DE MATERIALES', bold_centered)

    # Texto luego del encabezado con ajuste de texto
    worksheet.merge_range('A6:D8', 'Mediante la presente, hago constar que he recibido a través de la COORDINACIÓN DE COMPRAS, adscrita a la DIRECCIÓN DE ADMINISTRACIÓN DE LA GOBERNACIÓN DEL ESTADO MONAGAS, el siguiente material:', workbook.add_format({'font_size': 10, 'text_wrap': True}))

    # Fila vacía en la fila 9
    worksheet.write_blank('A9', None)

    # Encabezado de la tabla con bordes y texto en mayúsculas
    worksheet.write_row('A10', ['N°', 'DESCRIPCIÓN', 'CANT', 'UNIDAD'], table_header_format)

    # Ajustar el ancho de las columnas "N°", "Descripción", "Cantidad", y "Unidad"
    worksheet.set_column('A:A', 5)   # Columna N°
    worksheet.set_column('B:B', 57)  # Columna Descripcion
    worksheet.set_column('C:C', 7)   # Columna Cantidad
    worksheet.set_column('D:D', 13)  # Columna Unidades

    # Agregar datos de las filas con bordes y ajuste de texto
    for i, row in enumerate(rows, start=1):
        worksheet.write_row(f'A{i + 10}', [i, row[4].upper(), row[6], row[5].upper()], table_data_format)

    # Insertar una fila vacía después de la tabla
    row_count = len(rows) + 11
    worksheet.write_blank(f'A{row_count}', None)

    # Insertar "DESTINO DE LO SOLICITADO" y combinar de A hasta D
    worksheet.merge_range(f'A{row_count+1}:D{row_count+1}', 'DESTINO DE LO SOLICITADO', bold_centered)

    # Fila vacía
    worksheet.write_blank(f'A{row_count+2}', None)

    # Insertar el destino
    worksheet.merge_range(f'A{row_count+3}:D{row_count+3}', rows[0][2].upper(), destination_format)

    # Insertar tres filas vacías
    worksheet.write_blank(f'A{row_count+4}', None)
    worksheet.write_blank(f'A{row_count+5}', None)
    worksheet.write_blank(f'A{row_count+6}', None)

    # Añadir pie de firma
    worksheet.write(f'A{row_count+7}', 'Entregado por:')
    worksheet.write(f'C{row_count+7}', 'Recibido por:')
    worksheet.write(f'A{row_count+8}', 'Nombre y Apellido: ')
    worksheet.write(f'C{row_count+8}', 'Nombre y Apellido: ')
    worksheet.write(f'A{row_count+9}', 'Firma y Sello: ')
    worksheet.write(f'C{row_count+9}', 'Firma y Sello: ')

    # Guardar y enviar el archivo Excel
    workbook.close()
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=f'acta_{id_salida}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/salidas_excel', methods=['POST'])
def salidas_excel():
    data = request.get_json()
    fecha_inicio = data.get('fechaInicio')
    fecha_fin = data.get('fechaFin')
    search_input = data.get('searchInput', '').upper()

    conn = mysql.connect()
    cursor = conn.cursor()

    query_normal = '''
        SELECT
            aux_sal.id,
            salidas_inventario.fecha,
            art.articulo,
            aux_sal.cantidad,
            art.unidad,
            art.area,
            salidas_inventario.id_salida,
            salidas_inventario.destino
        FROM
            salidas_inventario
        JOIN
            aux_salidas_inventario aux_sal ON salidas_inventario.id_salida = aux_sal.id_salida
        JOIN
            articulos art ON aux_sal.id_articulo = art.id_articulo
        WHERE
            (salidas_inventario.fecha BETWEEN %s AND %s)
            AND (UPPER(art.articulo) LIKE %s
            OR UPPER(aux_sal.cantidad) LIKE %s
            OR UPPER(art.unidad) LIKE %s
            OR UPPER(art.area) LIKE %s
            OR UPPER(salidas_inventario.destino) LIKE %s)
        ORDER BY
            salidas_inventario.id_salida ASC, salidas_inventario.fecha ASC
    '''

    query_extra = '''
        SELECT
            artic_extra.id,
            salidas_inventario.fecha,
            artic_extra.articulo,
            artic_extra.cantidad,
            artic_extra.unidad,
            artic_extra.area,
            salidas_inventario.id_salida,
            salidas_inventario.destino
        FROM
            salidas_inventario
        JOIN
            articulos_extra_salida artic_extra ON salidas_inventario.id_salida = artic_extra.id_salida
        WHERE
            (salidas_inventario.fecha BETWEEN %s AND %s)
            AND (UPPER(artic_extra.articulo) LIKE %s
            OR UPPER(artic_extra.cantidad) LIKE %s
            OR UPPER(artic_extra.unidad) LIKE %s
            OR UPPER(artic_extra.area) LIKE %s
            OR UPPER(salidas_inventario.destino) LIKE %s)
        ORDER BY
            salidas_inventario.id_salida ASC, salidas_inventario.fecha ASC
    '''

    like_term = f"%{search_input}%"
    cursor.execute(query_normal, (fecha_inicio, fecha_fin, like_term, like_term, like_term, like_term, like_term))
    salidas_normales = cursor.fetchall()

    cursor.execute(query_extra, (fecha_inicio, fecha_fin, like_term, like_term, like_term, like_term, like_term))
    salidas_extras = cursor.fetchall()

    salidas = salidas_normales + salidas_extras

    grouped_salidas = []
    current_id_salida = None
    group = []

    for salida in salidas:
        if salida[6] != current_id_salida:
            if group:
                grouped_salidas.append(group)
            group = [salida]
            current_id_salida = salida[6]
        else:
            group.append(salida)

    if group:
        grouped_salidas.append(group)

    cursor.close()
    conn.close()

    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Salidas"

    header_font = Font(bold=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Agregar el título "Salidas" en la primera fila
    sheet.merge_cells('A1:G1')
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "Salidas"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = alignment_center

    # Crear encabezados (en la segunda fila)
    headers = ['ID', 'Fecha', 'Artículo', 'Cantidad', 'Unidad', 'Área', 'Destino']
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = border_style
        cell.fill = header_fill

    row_num = 3
    total_cantidad = 0
    for group in grouped_salidas:
        rowspan = len(group)
        for i, salida in enumerate(group):
            sheet.cell(row=row_num, column=3).value = salida[2]
            sheet.cell(row=row_num, column=4).value = salida[3]
            sheet.cell(row=row_num, column=5).value = salida[4]
            sheet.cell(row=row_num, column=6).value = salida[5]
            total_cantidad += salida[3]
            for col in range(3, 7):
                sheet.cell(row=row_num, column=col).border = border_style
                sheet.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True)

            if i == 0:
                sheet.cell(row=row_num, column=1).value = salida[6]
                sheet.cell(row=row_num, column=2).value = salida[1]
                sheet.cell(row=row_num, column=7).value = salida[7]
                for col in [1, 2, 7]:
                    sheet.merge_cells(start_row=row_num, start_column=col, end_row=row_num+rowspan-1, end_column=col)
                    for merged_row in range(row_num, row_num + rowspan):
                        merged_cell = sheet.cell(row=merged_row, column=col)
                        merged_cell.alignment = alignment_center
                        merged_cell.border = border_style
                        merged_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            row_num += 1

    # Agregar la fila con el total de cantidades al final (sin borde inferior en la celda TOTAL)
    sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
    total_label_cell = sheet.cell(row=row_num, column=1)
    total_label_cell.value = "TOTAL:"
    total_label_cell.font = Font(bold=True)
    total_label_cell.alignment = alignment_center
    no_bottom_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style=None)
    )
    total_label_cell.border = no_bottom_border

    total_value_cell = sheet.cell(row=row_num, column=4)
    total_value_cell.value = total_cantidad
    total_value_cell.font = Font(bold=True)
    total_value_cell.alignment = alignment_center
    total_value_cell.border = border_style

    for col in sheet.columns:
        max_length = 0
        column = col[0].coordinate[0]
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    workbook.save(output)
    output.seek(0)

    filename = "Compras (Salidas).xlsx"

    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# @app.route('/salidas_excel', methods=['POST'])
# def salidas_excel():
#     data = request.get_json()
#     fecha_inicio = data.get('fechaInicio')
#     fecha_fin = data.get('fechaFin')
#     search_input = data.get('searchInput', '').upper()

#     conn = mysql.connect()
#     cursor = conn.cursor()

#     query = '''
#         SELECT
#             aux_sal.id,
#             salidas_inventario.fecha,
#             art.articulo,
#             aux_sal.cantidad,
#             art.unidad,
#             art.area,
#             salidas_inventario.id_salida,
#             salidas_inventario.destino
#         FROM
#             salidas_inventario
#         JOIN
#             aux_salidas_inventario aux_sal ON salidas_inventario.id_salida = aux_sal.id_salida
#         JOIN
#             articulos art ON aux_sal.id_articulo = art.id_articulo
#         WHERE
#             (salidas_inventario.fecha BETWEEN %s AND %s)
#             AND (UPPER(art.articulo) LIKE %s
#             OR UPPER(aux_sal.cantidad) LIKE %s
#             OR UPPER(art.unidad) LIKE %s
#             OR UPPER(art.area) LIKE %s
#             OR UPPER(salidas_inventario.destino) LIKE %s)
#         ORDER BY
#             salidas_inventario.id_salida ASC, salidas_inventario.fecha ASC
#     '''

#     like_term = f"%{search_input}%"
#     cursor.execute(query, (fecha_inicio, fecha_fin, like_term, like_term, like_term, like_term, like_term))
#     salidas = cursor.fetchall()

#     grouped_salidas = []
#     current_id_salida = None
#     group = []

#     for salida in salidas:
#         if salida[6] != current_id_salida:
#             if group:
#                 grouped_salidas.append(group)
#             group = [salida]
#             current_id_salida = salida[6]
#         else:
#             group.append(salida)

#     if group:
#         grouped_salidas.append(group)

#     cursor.close()
#     conn.close()

#     # Crear un archivo Excel en memoria
#     output = BytesIO()
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = "Salidas"

#     # Estilo para los encabezados
#     header_font = Font(bold=True)
#     alignment_center = Alignment(horizontal='center', vertical='center')
#     border_style = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )
#     header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

#     # Agregar el título "Salidas" en la primera fila
#     sheet.merge_cells('A1:G1')  # Combina las celdas de la A1 a la G1
#     title_cell = sheet.cell(row=1, column=1)
#     title_cell.value = "Salidas"
#     title_cell.font = Font(bold=True, size=14)
#     title_cell.alignment = alignment_center

#     # Crear encabezados (en la segunda fila)
#     headers = ['ID', 'Fecha', 'Artículo', 'Cantidad', 'Unidad', 'Área', 'Destino']
#     for col_num, header in enumerate(headers, 1):
#         cell = sheet.cell(row=2, column=col_num)
#         cell.value = header
#         cell.font = header_font
#         cell.alignment = alignment_center
#         cell.border = border_style
#         cell.fill = header_fill

#     # Rellenar datos a partir de la tercera fila
#     row_num = 3
#     total_cantidad = 0  # Variable para almacenar la suma de cantidades
#     for group in grouped_salidas:
#         rowspan = len(group)
#         for i, salida in enumerate(group):
#             sheet.cell(row=row_num, column=3).value = salida[2]
#             sheet.cell(row=row_num, column=4).value = salida[3]
#             sheet.cell(row=row_num, column=5).value = salida[4]
#             sheet.cell(row=row_num, column=6).value = salida[5]
#             total_cantidad += salida[3]  # Sumar la cantidad
#             for col in range(3, 7):
#                 sheet.cell(row=row_num, column=col).border = border_style
#                 sheet.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True)

#             if i == 0:
#                 sheet.cell(row=row_num, column=1).value = salida[6]
#                 sheet.cell(row=row_num, column=2).value = salida[1]
#                 sheet.cell(row=row_num, column=7).value = salida[7]
#                 for col in [1, 2, 7]:
#                     sheet.merge_cells(start_row=row_num, start_column=col, end_row=row_num+rowspan-1, end_column=col)
#                     for merged_row in range(row_num, row_num + rowspan):
#                         merged_cell = sheet.cell(row=merged_row, column=col)
#                         merged_cell.alignment = alignment_center
#                         merged_cell.border = border_style
#                         merged_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
#             row_num += 1

#     # Agregar la fila con el total de cantidades al final (sin borde inferior en la celda TOTAL)
#     sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
#     total_label_cell = sheet.cell(row=row_num, column=1)
#     total_label_cell.value = "TOTAL:"
#     total_label_cell.font = Font(bold=True)
#     total_label_cell.alignment = alignment_center
#     # Remover el borde inferior
#     no_bottom_border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style=None)  # Sin borde inferior
#     )
#     total_label_cell.border = no_bottom_border

#     total_value_cell = sheet.cell(row=row_num, column=4)
#     total_value_cell.value = total_cantidad
#     total_value_cell.font = Font(bold=True)
#     total_value_cell.alignment = alignment_center
#     total_value_cell.border = border_style  # Mantener el borde normal aquí si es necesario


#     # Ajustar el ancho de las columnas automáticamente
#     for col in sheet.columns:
#         max_length = 0
#         column = col[0].coordinate[0]  # Obtén la letra de la columna a partir del atributo 'coordinate'
#         for cell in col:
#             if isinstance(cell, MergedCell):  # Saltar celdas fusionadas
#                 continue
#             try:
#                 if cell.value and len(str(cell.value)) > max_length:
#                     max_length = len(str(cell.value))
#             except:
#                 pass
#         adjusted_width = (max_length + 2)
#         sheet.column_dimensions[column].width = adjusted_width


#     workbook.save(output)
#     output.seek(0)

#     # Asignar un nombre específico al archivo Excel
#     filename = "Compras (Salidas).xlsx"

#     return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')





@app.route('/medicinas/salidas_excel', methods=['POST'])
def medicinas_salidas_excel():
    data = request.get_json()
    fecha_inicio = data.get('fechaInicio')
    fecha_fin = data.get('fechaFin')
    search_input = data.get('searchInput', '').upper()

    conn = mysql.connect()
    cursor = conn.cursor()

    query = '''
        SELECT
            aux_sal.id,
            salidas_medicinas.fecha,
            med.descripcion,
            med.marca,
            med.dosis,
            aux_sal.cantidad,
            med.unidad,
            salidas_medicinas.id_salida,
            salidas_medicinas.destino
        FROM
            salidas_medicinas
        JOIN
            aux_salidas_medicinas aux_sal ON salidas_medicinas.id_salida = aux_sal.id_salida
        JOIN
            medicinas med ON aux_sal.id_medicina = med.id_medicina
        WHERE
            (salidas_medicinas.fecha BETWEEN %s AND %s)
            AND (UPPER(med.descripcion) LIKE %s
            OR UPPER(med.marca) LIKE %s
            OR UPPER(med.dosis) LIKE %s
            OR UPPER(aux_sal.cantidad) LIKE %s
            OR UPPER(med.unidad) LIKE %s
            OR UPPER(salidas_medicinas.destino) LIKE %s)
        ORDER BY
            salidas_medicinas.id_salida ASC, salidas_medicinas.fecha ASC
    '''

    like_term = f"%{search_input}%"
    cursor.execute(query, (fecha_inicio, fecha_fin, like_term, like_term, like_term, like_term, like_term, like_term))
    salidas = cursor.fetchall()

    grouped_salidas = []
    current_id_salida = None
    group = []

    for salida in salidas:
        if salida[7] != current_id_salida:
            if group:
                grouped_salidas.append(group)
            group = [salida]
            current_id_salida = salida[7]
        else:
            group.append(salida)

    if group:
        grouped_salidas.append(group)

    cursor.close()
    conn.close()

    # Crear un archivo Excel en memoria
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Salidas"

    # Estilo para los encabezados
    header_font = Font(bold=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Agregar el título "Salidas" en la primera fila
    sheet.merge_cells('A1:H1')  # Combina las celdas de la A1 a la H1
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "Salidas"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = alignment_center

    # Crear encabezados (en la segunda fila)
    headers = ['ID', 'Fecha', 'Medicina', 'Marca', 'Dosis', 'Cantidad', 'Unidad', 'Destino']
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = border_style
        cell.fill = header_fill

    # Rellenar datos a partir de la tercera fila
    row_num = 3
    total_cantidad = 0  # Variable para almacenar la suma de cantidades
    for group in grouped_salidas:
        rowspan = len(group)
        for i, salida in enumerate(group):
            sheet.cell(row=row_num, column=3).value = salida[2]  # Descripción de la medicina
            sheet.cell(row=row_num, column=4).value = salida[3]  # Marca
            sheet.cell(row=row_num, column=5).value = salida[4]  # Dosis
            sheet.cell(row=row_num, column=6).value = salida[5]  # Cantidad
            sheet.cell(row=row_num, column=7).value = salida[6]  # Unidad
            total_cantidad += salida[5]  # Sumar la cantidad

            for col in range(3, 8):
                sheet.cell(row=row_num, column=col).border = border_style
                sheet.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True)

            if i == 0:
                sheet.cell(row=row_num, column=1).value = salida[7]  # ID de la salida
                sheet.cell(row=row_num, column=2).value = salida[1]  # Fecha
                sheet.cell(row=row_num, column=8).value = salida[8]  # Destino
                for col in [1, 2, 8]:
                    sheet.merge_cells(start_row=row_num, start_column=col, end_row=row_num+rowspan-1, end_column=col)
                    for merged_row in range(row_num, row_num + rowspan):
                        merged_cell = sheet.cell(row=merged_row, column=col)
                        merged_cell.alignment = alignment_center
                        merged_cell.border = border_style
                        merged_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

            row_num += 1

    # Agregar la fila con el total de cantidades al final
    sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
    total_label_cell = sheet.cell(row=row_num, column=1)
    total_label_cell.value = "TOTAL:"
    total_label_cell.font = Font(bold=True)
    total_label_cell.alignment = alignment_center
    total_label_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    total_value_cell = sheet.cell(row=row_num, column=6)
    total_value_cell.value = total_cantidad
    total_value_cell.font = Font(bold=True)
    total_value_cell.alignment = alignment_center
    total_value_cell.border = border_style

    # Ajustar el ancho de las columnas automáticamente
    for col in sheet.columns:
        max_length = 0
        column = col[0].coordinate[0]  # Obtén la letra de la columna a partir del atributo 'coordinate'
        for cell in col:
            if isinstance(cell, MergedCell):  # Saltar celdas fusionadas
                continue
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    workbook.save(output)
    output.seek(0)

    # Asignar un nombre específico al archivo Excel
    filename = "Salidas_Medicinas.xlsx"

    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')




@app.route('/inventario_excel', methods=['POST'])
def inventario_excel():
    data = request.get_json()
    search_input = data.get('searchInput', '').upper()

    conn = mysql.connect()
    cur = conn.cursor()

    query = '''
        SELECT
            stock.id_articulo,
            articulos.articulo,
            stock.cantidad,
            articulos.unidad,
            articulos.area
        FROM
            stock_inventario stock
        JOIN
            articulos ON stock.id_articulo = articulos.id_articulo
        WHERE
            stock.cantidad > 0
            AND (UPPER(stock.id_articulo) LIKE %s
            OR UPPER(articulos.articulo) LIKE %s
            OR UPPER(stock.cantidad) LIKE %s
            OR UPPER(articulos.unidad) LIKE %s
            OR UPPER(articulos.area) LIKE %s)
        ORDER BY
            articulos.articulo ASC '''

    like_term = f"%{search_input}%"
    cur.execute(query, (like_term, like_term, like_term, like_term, like_term))
    rows = cur.fetchall()

    cur.close()
    conn.close()

    # Crear un archivo Excel en memoria
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inventario Actual"

    # Estilo para los encabezados
    header_font = Font(bold=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Agregar el título "Inventario Disponible" en la primera fila
    sheet.merge_cells('A1:E1')  # Combina las celdas de la A1 a la E1
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "Inventario Disponible"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = alignment_center

    # Crear encabezados (en la segunda fila)
    headers = ['ID Artículo', 'Artículo', 'Cantidad', 'Unidad', 'Área']
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = border_style
        cell.fill = header_fill

    # Rellenar datos a partir de la tercera fila
    row_num = 3
    total_cantidad = 0  # Variable para almacenar la suma de cantidades
    for row in rows:
        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border_style
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        total_cantidad += row[2]  # Suponiendo que la cantidad está en la tercera columna
        row_num += 1

    # Agregar la fila con el total de cantidades al final
    sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
    total_label_cell = sheet.cell(row=row_num, column=1)
    total_label_cell.value = "TOTAL:"
    total_label_cell.font = Font(bold=True)
    total_label_cell.alignment = alignment_center
    total_label_cell.border = border_style

    total_value_cell = sheet.cell(row=row_num, column=3)
    total_value_cell.value = total_cantidad
    total_value_cell.font = Font(bold=True)
    total_value_cell.alignment = alignment_center
    total_value_cell.border = border_style

    # Ajustar el ancho de las columnas automáticamente
    for col in sheet.columns:
        max_length = 0
        column = col[0].coordinate[0]  # Obtén la letra de la columna a partir del atributo 'coordinate'
        for cell in col:
            if isinstance(cell, MergedCell):  # Saltar celdas fusionadas
                continue
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width


    workbook.save(output)
    output.seek(0)

    # Asignar un nombre específico al archivo Excel
    filename = "Inventario_Actual.xlsx"

    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')










### Gestion de Inventarios ####
# Gestion de Articulos

@app.route('/ferreteria/agregar_articulo')
def agregar_articulo_ferreteria():

    conn = mysql.connect()
    cursor = conn.cursor()

    #Vista para el usuario director (o administrador)

    cursor.execute("SELECT * FROM municipios ORDER BY nombre")

    municipios = cursor.fetchall()

    conn.commit()

    return render_template('ferreteria/agregar_articulo.html', municipios = municipios)

@app.route('/ferreteria/agregar_articulo_BD', methods=['POST'])
def ferreteria_agregar_articulo_BD():
    """Agrega un articulo a la base de datos"""
    try:
        fecha = request.form['txtFecha']
        articulos = [articulo.upper() for articulo in request.form.getlist('articulo[]')]
        cantidades = request.form.getlist('cantidad[]')
        unidades = [unidad.upper() for unidad in request.form.getlist('unidad[]')]
        areas = [area.upper() for area in request.form.getlist('area[]')]

        conn = mysql.connect()
        cursor = conn.cursor()
        id_generator = IDGenerator(conn)

        # Obtener el nuevo ID de entrada
        new_id_entrada = id_generator.generar_id_entrada_inventario()

        # Insertar en la tabla de entradas_inventario
        cursor.execute("INSERT INTO entradas_inventario (id_entrada, fecha) VALUES (%s, %s)", (new_id_entrada, fecha))

        articulos_existentes = []
        detalles_historial = []

        for articulo, cantidad, unidad, area in zip(articulos, cantidades, unidades, areas):
            # Verificar si el artículo ya existe en la tabla articulos
            cursor.execute("SELECT id_articulo FROM articulos WHERE articulo = %s AND unidad = %s", (articulo, unidad))
            result = cursor.fetchone()

            if result:
                id_articulo = result[0]
                articulos_existentes.append(articulo)
            else:
                # Generar un nuevo ID de artículo
                id_articulo = id_generator.generar_id_articulo()

                # Insertar el artículo en la tabla articulos
                cursor.execute("INSERT INTO articulos (id_articulo, articulo, unidad, area) VALUES (%s, %s, %s, %s)", (id_articulo, articulo, unidad, area))

            # Insertar en la tabla aux_entradas_inventario
            cursor.execute("INSERT INTO aux_entradas_inventario (id_entrada, id_articulo, cantidad) VALUES (%s, %s, %s)", (new_id_entrada, id_articulo, cantidad))

            # Actualizar la tabla de stock
            cursor.execute("SELECT cantidad FROM stock_inventario WHERE id_articulo = %s", (id_articulo,))
            stock_result = cursor.fetchone()

            if stock_result:
                nueva_cantidad = stock_result[0] + int(cantidad)
                cursor.execute("UPDATE stock_inventario SET cantidad = %s WHERE id_articulo = %s", (nueva_cantidad, id_articulo))
            else:
                cursor.execute("INSERT INTO stock_inventario (id_articulo, cantidad) VALUES (%s, %s)", (id_articulo, cantidad))

            # Agregar detalles al historial
            detalle = f"ARTÍCULO: {articulo}, CANTIDAD: {cantidad}, UNIDAD: {unidad}, ÁREA: {area}"
            detalles_historial.append(detalle)

        conn.commit()

        # Generar el mensaje de éxito y error
        if articulos_existentes:
            mensaje = 'Algunos artículos ya existían y fueron sumados al inventario: ' + ', '.join(articulos_existentes)
            flash(mensaje, 'warning')
        else:
            flash('¡Artículos agregados exitosamente!', 'success')

        # Registrar en el historial
        if detalles_historial:
            _detalles = "\n".join(detalles_historial)
            registrar_historial(conn, current_user.id, "REGISTRO", "FERRETERÍA", _detalles)

    except Exception as e:
        conn.rollback()
        flash(f'Error al agregar artículos: {str(e)}', 'error')
    finally:
        cursor.close()
        conn.close()

    return redirect('/ferreteria/inventario')

@app.route('/ferreteria/editar_articulo/<string:id_articulo>')
def editar_articulo_ferreteria(id_articulo):
    """Accede al Front de Editar Articulo"""

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM articulos WHERE id_articulo = %s", (id_articulo))

    articulo = cursor.fetchone()

    return render_template('ferreteria/editar_articulo.html', articulo = articulo)

@app.route('/ferreteria/editar_articulo_BD', methods=['POST'])
def ferreteria_editar_articulo_BD():
    """Guarda la modificación en la base de datos"""
    try:

        id_articulo = request.form['txtID']
        _articulo = request.form["articulo[]"].upper()
        _unidad = request.form["unidad[]"].upper()
        _area = request.form["area[]"].upper()

        conn = mysql.connect()
        cursor = conn.cursor()

        # Actualizar la fecha en la tabla de entradas_inventario
        cursor.execute("UPDATE articulos SET articulo = %s, unidad =%s, area =%s WHERE id_articulo = %s", (_articulo, _unidad, _area, id_articulo))

        conn.commit()

        flash('¡Articulo editado exitosamente!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Error al actualizar el articulo: {str(e)}', 'error')

    finally:
        cursor.close()
        conn.close()

    return redirect('/ferreteria/inventario')

@app.route('/eliminar_multiples_articulos', methods=['POST'])
@login_required
def eliminar_multiples_articulos():
    ids = request.json.get('ids', [])
    if isinstance(ids, str):  # Si ids es un solo ID en lugar de una lista
        ids = [ids]
    elif not ids:  # Si no se proporciona ningún ID
        return jsonify(success=False, message='No IDs provided'), 400

    try:
        conn = mysql.connect()
        cursor = conn.cursor()

        # Obtener información de los artículos antes de eliminarlos (para el registro)
        cursor.execute("SELECT id_articulo, articulo, unidad, area FROM articulos WHERE id_articulo IN %s", (tuple(ids),))
        articulos = cursor.fetchall()

        # Eliminar los artículos
        format_strings = ','.join(['%s'] * len(ids))
        query = f"DELETE FROM articulos WHERE id_articulo IN ({format_strings})"
        cursor.execute(query, ids)

        conn.commit()

        # Registrar cada eliminación en el historial de acciones
        _id_usuario = current_user.id
        _area = "OFICINA"
        _accion = "ELIMINACION"

        for articulo in articulos:
            codigo_articulo = articulo[0]
            nombre_articulo = articulo[1]
            _detalles = f"SE ELIMINÓ EL ARTICULO: {nombre_articulo}, CODIGO: {codigo_articulo}"
            registrar_historial(conn, _id_usuario, _accion, _area, _detalles)

        return jsonify(success=True)

    except Exception as e:
        conn.rollback()
        return jsonify(success=False, message=str(e)), 500

    finally:
        cursor.close()
        conn.close()

### Gestión de Oficina ###

##Articulos

@app.route('/oficina/agregar_articulo')
def agregar_articulo_oficina():

    return render_template('oficina/agregar_articulo.html')

@app.route('/oficina/agregar_articulo_BD', methods=['POST'])
@login_required
def oficina_agregar_articulo_BD():
    """Controlador para agregar nuevos artículos al inventario"""
    try:
        # Procesar datos del formulario
        form_data = procesar_formulario_articulos(request.form)
        ubicacion = "oficina"
        conn = mysql.connect()
        with conn:
            manager = InventarioManager(conn)
            
            # Registrar la entrada de artículos
            id_entrada = manager.agregar_entrada(
                fecha=form_data['fecha'],
                ubicacion=ubicacion,
                items=form_data['articulos_inventario']
            )
            
            # Registrar en el historial
            registrar_historial_entrada(
                conn,
                current_user.id,
                form_data['articulos_inventario']
            )
            
            flash('¡Artículos agregados exitosamente!', 'success')
            return redirect('/oficina/inventario')

    except Exception as e:
        print(e)
        return redirect('/oficina/inventario')

#Funcion auxiliar de agregar articulos
def procesar_formulario_articulos(form_data):
    """
    Procesa datos de formulario para agregar artículos (sin extras ni destino)
    :param form_data: Datos del formulario
    :return: Diccionario con datos estructurados
    """
    processed = {
        'fecha': form_data.get('txtFecha', ''),
        'articulos_inventario': []
    }

    # Obtener listas de campos del formulario
    articulos = [a.upper() for a in form_data.getlist('articulo[]')]
    cantidades = form_data.getlist('cantidad[]')
    unidades = [u.upper() for u in form_data.getlist('unidad[]')]
    areas = [area.upper() for area in form_data.getlist('area[]')]

    # Procesar cada artículo
    for idx, (articulo, cantidad, unidad, area) in enumerate(zip(
        articulos, cantidades, unidades, areas
    )):
        processed['articulos_inventario'].append({
            'articulo': articulo,
            'cantidad': cantidad,
            'unidad': unidad,
            'area': area,
            'orden': idx
        })

    return processed


@app.route('/oficina/editar_articulo/<string:id_articulo>')
def editar_articulo_oficina(id_articulo):
    """Accede al Front de Editar Articulo"""

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM articulos WHERE id_articulo = %s", (id_articulo))

    articulo = cursor.fetchone()

    return render_template('oficina/editar_articulo.html', articulo = articulo)

@app.route('/oficina/editar_articulo_BD', methods=['POST'])
def oficina_editar_articulo_BD():
    """Guarda la modificación en la base de datos"""
    try:

        id_articulo = request.form['txtID']
        _articulo = request.form["articulo[]"].upper()
        _unidad = request.form["unidad[]"].upper()
        _area = request.form["area[]"].upper()

        conn = mysql.connect()
        cursor = conn.cursor()

        # Actualizar la fecha en la tabla de entradas_inventario
        cursor.execute("UPDATE articulos SET articulo = %s, unidad =%s, area =%s WHERE id_articulo = %s", (_articulo, _unidad, _area, id_articulo))

        conn.commit()

        flash('¡Articulo editado exitosamente!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Error al actualizar el articulo: {str(e)}', 'error')

    finally:
        cursor.close()
        conn.close()

    return redirect('/oficina/inventario')



# Consultar inventario
@app.route('/oficina/inventario')
@login_required
def oficina_inventario():
    # Conectar a la base de datos
    conn = mysql.connect()

    try:
        # Crear el manejador de inventario
        inventario_manager = InventarioManager(conn)

        # Obtener inventario y total de cantidades
        inventario = inventario_manager.obtener_inventario(ubicacion="oficina")
        total_cantidad = inventario_manager.obtener_total_cantidad('inventario','oficina')

        # Renderizar el template con los datos
        return render_template('/oficina/inventario.html', inventario=inventario, total=total_cantidad)

    except Exception as e:
        # Manejo de errores en caso de que algo falle
        flash(f'Error al obtener inventario: {str(e)}', 'error')
        return render_template('general/notfound.html'), 404  # Redirigir a la página de inicio en caso de error

    finally:
        # Asegurarse de cerrar la conexión
        inventario_manager.cerrar_conexion()

# Gestion de Entradas (Oficina) #

### Gestion de Medicinas ENTRADAS ###

@app.route('/oficina/entradas')
@login_required
def oficina_entradas():

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT
            aux_entradas_inventario.id,
            DATE_FORMAT(entradas_inventario.fecha, '%d/%m/%Y') as fecha,
            articulos.articulo,
            aux_entradas_inventario.cantidad,
            articulos.unidad,
            articulos.area,
            entradas_inventario.id_entrada
        FROM
            entradas_inventario
        JOIN
            aux_entradas_inventario ON entradas_inventario.id_entrada = aux_entradas_inventario.id_entrada
        JOIN
            articulos ON aux_entradas_inventario.id_articulo = articulos.id_articulo
        WHERE
            entradas_inventario.ubicacion = "oficina"
        ORDER BY
            entradas_inventario.id_entrada ASC, entradas_inventario.fecha ASC
    ''')

    entradas = cursor.fetchall()
    print(entradas)

    # Convertir resultados a diccionarios
    entradas = [
        {
            'id': entrada[0],
            'fecha': entrada[1],
            'articulo': entrada[2],
            'cantidad': entrada[3],
            'unidad': entrada[4],
            'area': entrada[5],
            'id_entrada': entrada[6]
        }
        for entrada in entradas
    ]

    # Agrupar las entradas por ID Entrada
    grouped_entradas = []
    current_id_entrada = None
    group = []

    for entrada in entradas:
        if entrada['id_entrada'] != current_id_entrada:
            if group:
                grouped_entradas.append(group)
            group = [entrada]
            current_id_entrada = entrada['id_entrada']
        else:
            group.append(entrada)

    if group:
        grouped_entradas.append(group)

    print("Las entradas agrupadas son: ", grouped_entradas)

    # Calcular la cantidad total de artículos ingresados
    cursor.execute('''
        SELECT
            SUM(cantidad) as cantidad_total
        FROM
            aux_entradas_inventario
        WHERE
            id_entrada LIKE 'EOF-%'
    ''')
    cantidad_total = cursor.fetchone()[0]
    print("Cantidad Total es: ", cantidad_total)

    cursor.close()
    conn.close()

    return render_template('/oficina/entradas.html', grouped_entradas=grouped_entradas, cantidad_total=cantidad_total)

@app.route('/oficina/agregar_entrada', methods=['GET', 'POST'])
def agregar_entrada_oficina():
    """ Accede al Front de Agregar Entrada (Oficina) """

    # Crear una instancia de InventarioManager
    conn = mysql.connect()
    inventario_manager = InventarioManager(conn)

    # Definir la ubicación (puede ser "galpon" o "oficina")
    ubicacion = "oficina"

    # Obtener los artículos y su cantidad según la ubicación
    articulos = inventario_manager.obtener_articulos_con_stock(ubicacion)

    # Cerrar la conexión
    inventario_manager.cerrar_conexion()

    # Renderizar la plantilla con los artículos obtenidos
    return render_template('oficina/agregar_entrada.html', articulos=articulos)

# Ruta actualizada
@app.route('/oficina/agregar_entrada_BD', methods=['POST'])
@login_required
def oficina_agregar_entrada_BD():
    """Controlador para agregar entradas de inventario"""
    try:
        form_data = procesar_formulario_articulos(request.form)
        ubicacion = "oficina"
        conn = mysql.connect()
        with conn:
            manager = InventarioManager(conn)
            
            # Registrar entrada
            id_entrada = manager.agregar_entrada(
                fecha=form_data['fecha'],
                ubicacion=ubicacion,
                items=form_data['articulos_inventario']
            )
            
            # Registrar historial
            registrar_historial_entrada(
                conn,
                current_user.id,
                form_data['articulos_inventario']
            )

            conn.commit()
            conn.close()
            
            flash('¡Entradas agregadas y stock actualizado exitosamente!', 'success')
            return redirect('/oficina/entradas')

    except Exception as e:
        print(e)
        return redirect('/oficina/entradas')

# Funciones auxiliares
def registrar_historial_entrada(conn, id_usuario, articulos):
    """Registra múltiples entradas en el historial"""
    with conn.cursor() as cursor:
        for articulo in articulos:
            query = """
                INSERT INTO historial 
                (id_usuario, accion, area, detalles)
                VALUES (%s, %s, %s, %s)
            """
            detalles = (
                f"AGREGADA {articulo['cantidad']} {articulo['unidad']} "
                f"DE '{articulo['articulo']}', ÁREA: {articulo['area']}"
            )
            cursor.execute(query, (
                id_usuario,
                "REGISTRO",
                "oficina",
                detalles
            ))


@app.route('/oficina/editar_entrada/<string:id_entrada>')
def editar_entrada_oficina(id_entrada):

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM articulos WHERE ubicacion = 'oficina' ORDER BY articulo")

    articulos = cursor.fetchall()

    cursor.execute("""
        SELECT
            aux_ent.id,
            aux_ent.id_entrada,
            art.articulo AS nombre_articulo,
            art.unidad,
            art.area,
            aux_ent.cantidad,
            ent.fecha
        FROM
            aux_entradas_inventario aux_ent
        JOIN
            articulos art ON aux_ent.id_articulo = art.id_articulo
        JOIN
            entradas_inventario ent ON aux_ent.id_entrada = ent.id_entrada
        WHERE
            aux_ent.id_entrada = %s
        """, (id_entrada,))

    entradas = cursor.fetchall()

    return render_template('oficina/editar_entrada.html', articulos = articulos, entradas = entradas)


@app.route('/oficina/editar_entrada_BD', methods=['POST'])
def oficina_editar_entrada_BD():

    # try:
        # Obtener datos del formulario
        ubicacion = 'oficina'
        id_entrada = request.form['txtID']
        fecha = request.form['txtFecha']
        articulos = [articulo.upper() for articulo in request.form.getlist('articulo[]')]
        cantidades = request.form.getlist('cantidad[]')
        unidades = [unidad.upper() for unidad in request.form.getlist('unidad[]')]
        areas = [area.upper() for area in request.form.getlist('area[]')]

        # Conectar a la base de datos
        conn = mysql.connect()
        inventario_manager = InventarioManager(conn)

        # Editar la entrada y actualizar el stock
        inventario_manager.editar_entrada(id_entrada, fecha, articulos, cantidades, unidades, areas, ubicacion)
        inventario_manager.actualizar_stock(ubicacion="oficina")

        # Registrar la acción en el historial
        _id_usuario = current_user.id
        _accion = "EDICIÓN"
        _area = "FERRETERÍA"
        for articulo, cantidad, unidad, area in zip(articulos, cantidades, unidades, areas):
            detalles = f"SE ACTUALIZÓ {cantidad} {unidad} DE {articulo} EN LA ENTRADA {id_entrada} EL {fecha}"
            registrar_historial(conn, _id_usuario, _accion, _area, detalles)

        # Confirmar cambios
        inventario_manager.commit()
        flash('¡Entrada actualizada exitosamente!', 'success')

    # except Exception as e:
        # Revertir cambios en caso de error
        # inventario_manager.rollback()
        # flash(f'Error al actualizar la entrada: {str(e)}', 'error')

    # finally:
        # Cerrar conexión
        inventario_manager.cerrar_conexion()

        return redirect('/oficina/entradas')

### Gestión de Salidas (Oficina) ###
@app.route('/oficina/salidas')
@login_required
def oficina_salidas():
    """Consultar el FrontEnd de las Salidas en Oficina"""

    conn = mysql.connect()
    manager = InventarioManager(conn)  # Instanciar la clase InventarioManager

    # Obtener las salidas agrupadas
    salidas = manager.obtener_entradas_o_salidas(tipo = "salidas", ubicacion="oficina")
    # print("Salidas es: ", salidas)

    # Agrupar las salidas por ID de salida
    grouped_salidas = []
    current_id_salida = None
    group = []

    for salida in salidas:
        if salida[6] != current_id_salida:  # Index 6 corresponde a id_salida
            if group:
                grouped_salidas.append(group)
            group = [list(salida)]  # Convertir cada salida en una lista
            current_id_salida = salida[6]
        else:
            group.append(list(salida))  # Convertir cada salida en una lista

    if group:
        grouped_salidas.append(group)

    # Obtener los artículos extra asociados a cada salida
    for salida_group in grouped_salidas:
        for salida in salida_group:
            id_salida = salida[6]
            articulos_extra = manager.obtener_articulos_extra_por_salida(id_salida)
            salida.append(articulos_extra)  # Ahora 'salida' es una lista, por lo que puedes usar append

    # Calcular la cantidad total de artículos en las salidas
    cantidad_total = 0 #sum(salida[3] for salida in salidas)  # Index 3 corresponde a cantidad

    # Recorremos cada grupo en grouped_salidas
    for group in grouped_salidas:
        # Recorremos cada salida en el grupo
        for salida in group:
            # Recorremos cada elemento en salida y lo sustituimos por una cadena vacía ' ' si es None
            for i in range(len(salida)):
                if salida[i] is None:
                    salida[i] = ' '



    # print("\n \n Grouped salidas es: ", grouped_salidas, "\n \n")

    return render_template(
        '/oficina/salidas.html',
        grouped_salidas=grouped_salidas,
        cantidad_total=cantidad_total
    )

@app.route('/oficina/agregar_salida')
def agregar_salida_oficina():
    """Accede al Front de Agregar las Salidas (Galpon)"""

    conn = mysql.connect()
    inventario_manager = InventarioManager(conn)
    # Definir la ubicación (puede ser "galpon" o "oficina")
    ubicacion = "oficina"
    articulos = inventario_manager.obtener_articulos_con_stock(ubicacion)
    # Cerrar la conexión
    inventario_manager.cerrar_conexion()

    return render_template('oficina/agregar_salida.html', articulos = articulos)


def procesar_formulario_salida(form_data, es_edicion=False):
    """
    Procesa datos de formulario para salidas, compatible con agregar y editar
    :param form_data: Datos del formulario
    :param es_edicion: True si es una operación de edición
    :return: Diccionario con datos estructurados
    """
    processed = {
        'fecha': form_data.get('txtFecha', ''),
        'destino': form_data.get('txtDestino', '').upper(),
        'articulos_inventario': [],
        'articulos_extra': []
    }
    
    if es_edicion:
        processed['id_salida'] = form_data.get('txtID', '')

    # Obtener listas de campos del formulario
    articulos = [a.upper() for a in form_data.getlist('articulo[]')]
    cantidades = form_data.getlist('cantidad[]')
    unidades = [u.upper() for u in form_data.getlist('unidad[]')]
    areas = [area.upper() for area in form_data.getlist('area[]')]
    extras = form_data.getlist('class_extra[]')

    # Procesar cada artículo
    for idx, (articulo, cantidad, unidad, area, extra) in enumerate(zip(
        articulos, cantidades, unidades, areas, extras
    )):
        articulo_data = {
            'articulo': articulo,
            'cantidad': cantidad,
            'unidad': unidad,
            'area': area,
            # 'orden': idx  # Mantener orden original si es necesario
        }

        if extra == 'on':
            processed['articulos_extra'].append(articulo_data)
        else:
            processed['articulos_inventario'].append(articulo_data)

    return processed


@app.route('/oficina/agregar_salida_BD', methods=['POST'])
@login_required
def oficina_agregar_salida_BD():
    """Controlador para registrar nuevas salidas de inventario en oficina"""
    try:
        ubicacion = "oficina"
        
        # 1. Procesamiento de datos del formulario
        form_data = procesar_formulario_salida(request.form)
        
        # 2. Conexión y gestión de transacciones
        conn = mysql.connect()
        inventario_manager = InventarioManager(conn)
        
        # 3. Registrar la salida principal
        id_salida = registrar_salida_principal(
            inventario_manager, 
            form_data, 
            ubicacion
        )

       # print("Form data es: ", form_data)
        
        # 4. Registrar artículos extra si existen
        if form_data['articulos_extra']:
            registrar_articulos_extra(
                inventario_manager,
                id_salida,
                form_data['articulos_extra']
            )

        # 5. Actualización final del stock
        inventario_manager.actualizar_stock(ubicacion)
        conn.commit()
        
        flash('¡Salidas agregadas y stock actualizado exitosamente!', 'success')
        return redirect('/oficina/salidas')

    except Exception as e:
        print("Error: ", e)
        conn.rollback()
        return redirect('/oficina/salidas')


def registrar_salida_principal(manager, form_data, ubicacion):
    """Registra la salida principal y sus artículos asociados"""
    return manager.agregar_salida(
        fecha=form_data['fecha'],
        destino=form_data['destino'],
        ubicacion=ubicacion,
        items=form_data['articulos_inventario']
    )

def registrar_articulos_extra(manager, id_salida, articulos_extra):
    """Registra múltiples artículos extra asociados a una salida"""

    for articulo_extra in articulos_extra:
        
        manager.registrar_articulo_extra(
            id_salida = id_salida,
            articulo = articulo_extra['articulo'],
            cantidad = articulo_extra['cantidad'],
            unidad = articulo_extra['unidad'],
            area = articulo_extra['area'],
        )

@app.route('/oficina/editar_salida/<string:id_salida>')
def editar_salida_oficina(id_salida):

    ubicacion = "oficina"
    conn = mysql.connect()
    cursor = conn.cursor()
    inventario_manager = InventarioManager(conn)

    articulos = inventario_manager.obtener_articulos_con_stock(ubicacion)

    # obtener los datos de los artículos comunes asociados a la salida
    cursor.execute("""
        SELECT
            aux_salidas_inventario.id,
            aux_salidas_inventario.id_salida,
            articulos.articulo AS nombre_articulo,
            articulos.unidad,
            articulos.area,
            aux_salidas_inventario.cantidad,
            salidas_inventario.fecha,
            salidas_inventario.destino
        FROM
            aux_salidas_inventario
        JOIN
            articulos ON aux_salidas_inventario.id_articulo = articulos.id_articulo
        JOIN
            salidas_inventario ON aux_salidas_inventario.id_salida = salidas_inventario.id_salida
        WHERE
            aux_salidas_inventario.id_salida = %s
        """, (id_salida,))

    salidas = cursor.fetchall()

    # obtener los datos de los artículos extra asociados a la salida
    cursor.execute("""
        SELECT
            articulos_extra_salida.id,
            articulos_extra_salida.id_salida,
            articulos_extra_salida.articulo AS nombre_articulo,
            articulos_extra_salida.unidad,
            articulos_extra_salida.area,
            articulos_extra_salida.cantidad,
            salidas_inventario.fecha,
            salidas_inventario.destino
        FROM
            articulos_extra_salida
        JOIN
            salidas_inventario ON articulos_extra_salida.id_salida = salidas_inventario.id_salida
        WHERE
            articulos_extra_salida.id_salida = %s
        """, (id_salida,))

    salidas_extra = cursor.fetchall()
    print("\n \n Salidas Extra es: \n", salidas_extra, "\n\n")

    inventario_manager.cerrar_conexion()

    return render_template('oficina/editar_salida.html', articulos=articulos, salidas=salidas, salidas_extra=salidas_extra)


@app.route('/oficina/editar_salida_BD', methods=['POST'])
@login_required
def oficina_editar_salida_BD():
    # try:
        form_data = procesar_formulario_salida(request.form, es_edicion=True)
        print("\n\n\n\n\n")
        print("Los datos son: ", form_data)
        print("\n\n\n\n")

        ubicacion = "oficina"
        conn = mysql.connect()
        with conn:
            manager = InventarioManager(conn)
            
            # 1. Actualizar cabecera con control de versión
            manager.actualizar_salida(
                form_data['id_salida'],
                form_data['fecha'],
                form_data['destino']
            )

            # 2. Eliminar registros antiguos
            manager.eliminar_articulos_salida(form_data['id_salida'])

            # 3. Insertar nuevamente los articulos

            #3.1 Articulos comunes
            for item in form_data['articulos_inventario']:
                # Obtener el id_articulo usando el nombre del artículo (requiere una consulta a la base de datos)
                id_articulo = manager._registrar_articulo(
                    item['articulo'],
                    item['unidad'],
                    item['area'],
                    ubicacion
                )
                
                # Registrar en detalle de salida
                manager._registrar_detalle_salida(
                    form_data['id_salida'], 
                    id_articulo, 
                    item['cantidad']
                )

            #3.2 Articulos extras
            if form_data['articulos_extra']:
                registrar_articulos_extra(
                    manager,
                    form_data['id_salida'],
                    form_data['articulos_extra']
                )
            
            # 4. Actualizar stock general
            manager.actualizar_stock(ubicacion)
            
            flash('¡Salida actualizada exitosamente!', 'success')
            return redirect('/oficina/salidas')

    # except Exception as e:
    #     handle_database_error(e)
        return redirect('/oficina/salidas')





# @app.route('/oficina/editar_salida_BD', methods=['POST'])
# def oficina_editar_salida_BD():
#     """Guarda la modificacion en la base de datos"""
#     ubicacion = "oficina"
#     try:
#         id_salida = request.form['txtID']
#         fecha = request.form['txtFecha']
#         destino = request.form['txtDestino'].upper()
#         articulos = [articulo.upper() for articulo in request.form.getlist('articulo[]')]
#         cantidades = request.form.getlist('cantidad[]')
#         unidades = [unidad.upper() for unidad in request.form.getlist('unidad[]')]
#         areas = [area.upper() for area in request.form.getlist('area[]')]

#         conn = mysql.connect()
#         inventario_manager = InventarioManager(conn)
#         cursor = conn.cursor()

#         # Actualizar la fecha en la tabla de salidas_inventario
#         cursor.execute("UPDATE salidas_inventario SET fecha = %s, destino = %s WHERE id_salida = %s", (fecha, destino, id_salida))

#         # Obtener los artículos existentes en la salida
#         cursor.execute("SELECT id_articulo, cantidad FROM aux_salidas_inventario WHERE id_salida = %s", (id_salida,))
#         existing_entries = cursor.fetchall()
#         existing_articles = {row[0]: row[1] for row in existing_entries}

#         new_articles = set()
#         for articulo, cantidad, unidad, area in zip(articulos, cantidades, unidades, areas):

#             # Obtener el ID del artículo directamente
#             cursor.execute("SELECT id_articulo FROM articulos WHERE articulo = %s AND unidad = %s", (articulo, unidad))
#             result = cursor.fetchone()

#             if result:
#                 id_articulo = result[0]
#                 new_articles.add(id_articulo)
#             else:
#                 flash(f"El artículo '{articulo}' con unidad '{unidad}' no existe en la base de datos.", 'warning')
#                 continue

#             cantidad = int(cantidad)

#             # Verificar si el artículo ya existe en la salida
#             if id_articulo in existing_articles:
#                 old_cantidad = existing_articles[id_articulo]
#                 diferencia = cantidad - old_cantidad

#                 # Actualizar la cantidad en aux_salidas_inventario
#                 cursor.execute("UPDATE aux_salidas_inventario SET cantidad = %s WHERE id_salida = %s AND id_articulo = %s", (cantidad, id_salida, id_articulo))

#                 # Actualizar la tabla stock_inventario
#                 #cursor.execute("UPDATE stock_inventario SET cantidad = cantidad - %s WHERE id_articulo = %s", (diferencia, id_articulo))
#             else:
#                 # Si no existe, insertarlo en aux_salidas_inventario
#                 cursor.execute("INSERT INTO aux_salidas_inventario (id_salida, id_articulo, cantidad) VALUES (%s, %s, %s)", (id_salida, id_articulo, cantidad))

#                 # Actualizar el stock aumentando la cantidad
#                 #cursor.execute("UPDATE stock_inventario SET cantidad = cantidad - %s WHERE id_articulo = %s", (cantidad, id_articulo))

#         # Eliminar artículos que ya no están en la lista
#         articles_to_remove = set(existing_articles.keys()) - new_articles
#         for art_id in articles_to_remove:
#             old_cantidad = existing_articles[art_id]

#             # Aumentar el stock por la cantidad eliminada
#             #cursor.execute("UPDATE stock_inventario SET cantidad = cantidad + %s WHERE id_articulo = %s", (old_cantidad, art_id))

#             # Eliminar el artículo de aux_salidas_inventario
#             cursor.execute("DELETE FROM aux_salidas_inventario WHERE id_salida = %s AND id_articulo = %s", (id_salida, art_id))



#         # # Registrar la acción en el historial
#         # for articulo, cantidad, unidad, area in zip(articulos, cantidades, unidades, areas):
#         #     detalles = f"SE ACTUALIZÓ {cantidad} {unidad} DE {articulo} EN LA SALIDA {id_salida} EL {fecha}"
#         #     registrar_historial(conn, current_user.id, "EDICIÓN", "FERRETERÍA", detalles)

#         flash('¡Salida actualizada exitosamente!', 'success')


#         inventario_manager.actualizar_stock(ubicacion)
#         conn.commit()
#         inventario_manager.cerrar_conexion()


#     except Exception as e:
#         conn.rollback()
#         flash(f'Error al actualizar la salida: {str(e)}', 'error')
#         conn.close()
#     finally:
#         cursor.close()
#         # conn.close()

#     return redirect('/oficina/salidas')

### Reportes de la Oficina ###

@app.route('/oficina/inventario_pdf', methods=['POST'])
def oficina_inventario_pdf():
    conn = mysql.connect()
    cursor = conn.cursor()

    # Extraer solo oficinas con cantidad mayor a 0
    cursor.execute('''
        SELECT
            articulos.id_articulo,
            articulos.articulo,
            IFNULL(stock_inventario.cantidad, 0) AS cantidad,
            articulos.unidad,
            articulos.area
        FROM
            articulos
        LEFT JOIN
            stock_inventario ON articulos.id_articulo = stock_inventario.id_articulo
        WHERE
            articulos.ubicacion = "oficina" AND stock_inventario.cantidad > 0
        ORDER BY
            articulos.articulo ASC
    ''')


    rows = cursor.fetchall()

    # Obtener la sumatoria total de todas las cantidades en stock mayores a 0
    cursor.execute('''
        SELECT
            SUM(cantidad) AS total_cantidad
        FROM
            stock_inventario
        WHERE
            ubicacion = "oficina" AND cantidad > 0
    ''')

    # Convertir el resultado de la consulta a un número entero
    total_result = cursor.fetchone()
    total = int(total_result[0]) if total_result[0] is not None else 0
    # print("total es: ", total)

    conn.close()

    # Renderizar el template con los datos agrupados y el total
    rendered = render_template('oficina/reportes/reporte_inventario.html', datos=rows, total=total)

    # Generar el PDF con WeasyPrint
    html = HTML(string=rendered)
    pdf = html.write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=reporte_inventario.pdf'

    return response

@app.route('/oficina/inventario_excel', methods=['POST'])
def oficina_inventario_excel():
    data = request.get_json()
    search_input = data.get('searchInput', '').upper()

    conn = mysql.connect()
    cur = conn.cursor()

    query = '''
        SELECT
            stock.id_articulo,
            articulos.articulo,
            stock.cantidad,
            articulos.unidad,
            articulos.area
        FROM
            stock_inventario stock
        JOIN
            articulos ON stock.id_articulo = articulos.id_articulo
        WHERE
            stock.cantidad > 0
            AND stock.ubicacion = "oficina"
            AND (UPPER(stock.id_articulo) LIKE %s
            OR UPPER(articulos.articulo) LIKE %s
            OR UPPER(stock.cantidad) LIKE %s
            OR UPPER(articulos.unidad) LIKE %s
            OR UPPER(articulos.area) LIKE %s)
        ORDER BY
            articulos.articulo ASC '''

    like_term = f"%{search_input}%"
    cur.execute(query, (like_term, like_term, like_term, like_term, like_term))
    rows = cur.fetchall()

    cur.close()
    conn.close()

    # Crear un archivo Excel en memoria
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inventario Actual"

    # Estilo para los encabezados
    header_font = Font(bold=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Agregar el título "Inventario Disponible" en la primera fila
    sheet.merge_cells('A1:E1')  # Combina las celdas de la A1 a la E1
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "Inventario Disponible"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = alignment_center

    # Crear encabezados (en la segunda fila)
    headers = ['ID Artículo', 'Artículo', 'Cantidad', 'Unidad', 'Área']
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = border_style
        cell.fill = header_fill

    # Rellenar datos a partir de la tercera fila
    row_num = 3
    total_cantidad = 0  # Variable para almacenar la suma de cantidades
    for row in rows:
        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border_style
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        total_cantidad += row[2]  # Suponiendo que la cantidad está en la tercera columna
        row_num += 1

    # Agregar la fila con el total de cantidades al final
    sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
    total_label_cell = sheet.cell(row=row_num, column=1)
    total_label_cell.value = "TOTAL:"
    total_label_cell.font = Font(bold=True)
    total_label_cell.alignment = alignment_center
    total_label_cell.border = border_style

    total_value_cell = sheet.cell(row=row_num, column=3)
    total_value_cell.value = total_cantidad
    total_value_cell.font = Font(bold=True)
    total_value_cell.alignment = alignment_center
    total_value_cell.border = border_style

    # Ajustar el ancho de columnas específicas
    sheet.column_dimensions['A'].width = 15  # Ancho de la columna ID Artículo
    sheet.column_dimensions['B'].width = 40  # Ancho de la columna Artículo

    # Ajustar el ancho de las demás columnas automáticamente
    for col in sheet.columns:
        max_length = 0
        column = col[0].coordinate[0]  # Obtén la letra de la columna a partir del atributo 'coordinate'
        for cell in col:
            if isinstance(cell, MergedCell):  # Saltar celdas fusionadas
                continue
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if column not in ['A','B']: # Evita la columna A y B ya que la establecemos manualmente
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width


    workbook.save(output)
    output.seek(0)

    # Asignar el nombre especificado al archivo Excel
    filename = "Inventario_Actual.xlsx"

    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')






























### Gestión del Galpón ###

# Consultar inventario
@app.route('/ferreteria/inventario')
@login_required
def ferreteria_inventario():
    # Conectar a la base de datos
    conn = mysql.connect()

    try:
        # Crear el manejador de inventario
        inventario_manager = InventarioManager(conn)

        # Obtener inventario y total de cantidades
        inventario = inventario_manager.obtener_inventario()
        total_cantidad = inventario_manager.obtener_total_cantidad('inventario','galpon')

        # Renderizar el template con los datos
        return render_template('/ferreteria/inventario.html', inventario=inventario, total=total_cantidad)

    except Exception as e:
        # Manejo de errores en caso de que algo falle
        flash(f'Error al obtener inventario: {str(e)}', 'error')
        return render_template('general/notfound.html'), 404  # Redirigir a la página de inicio en caso de error

    finally:
        # Asegurarse de cerrar la conexión
        inventario_manager.cerrar_conexion()

# Gestión de Entradas (Galpon) #

@app.route('/ferreteria/entradas')
@login_required
def ferreteria_entradas():
    """Consultar el FrontEnd de las Entradas (Galpón)"""

    # Conectar a la base de datos
    conn = mysql.connect()
    inventario_manager = InventarioManager(conn)

    # Obtener las entradas del galpón
    tipo = "entradas"  # O puede ser "salidas" si es necesario
    ubicacion = "galpon"  # O puede ser "palacio" según corresponda

    # Llamar al método de InventarioManager
    entradas = inventario_manager.obtener_entradas_o_salidas(tipo, ubicacion)

    # Calcular la cantidad total de artículos ingresados en el galpón
    total_entradas_galpon = inventario_manager.obtener_total_cantidad(tipo=tipo, ubicacion=ubicacion)

    conn.close()

    # Pasar los datos al template
    return render_template('/ferreteria/entradas.html', entradas=entradas, cantidad_total=total_entradas_galpon)


@app.route('/ferreteria/agregar_entrada', methods=['GET', 'POST'])
def agregar_entrada_ferreteria():
    """ Accede al Front de Agregar Entrada (Galpon) """

    # Crear una instancia de InventarioManager
    conn = mysql.connect()
    inventario_manager = InventarioManager(conn)

    # Definir la ubicación (puede ser "galpon" o "oficina")
    ubicacion = "galpon"

    # Obtener los artículos y su cantidad según la ubicación
    articulos = inventario_manager.obtener_articulos_con_stock(ubicacion)

    # Cerrar la conexión
    inventario_manager.cerrar_conexion()

    # Renderizar la plantilla con los artículos obtenidos
    return render_template('ferreteria/agregar_entrada.html', articulos=articulos)

@app.route('/ferreteria/agregar_entrada_BD', methods=['POST'])
@login_required
def ferreteria_agregar_entrada_BD():
    """ Agrega una entrada de artículos al Galpón """
    try:

        ubicacion = "galpon"

        # Recibir datos del formulario
        fecha = request.form['txtFecha']
        articulos = [articulo.upper() for articulo in request.form.getlist('articulo[]')]
        cantidades = request.form.getlist('cantidad[]')
        unidades = [unidad.upper() for unidad in request.form.getlist('unidad[]')]
        areas = [area.upper() for area in request.form.getlist('area[]')]

        # Conectar a la base de datos
        conn = mysql.connect()
        inventario_manager = InventarioManager(conn)

        # Agregar entrada con la ubicación
        inventario_manager.agregar_entrada(fecha, articulos, cantidades, unidades, areas, ubicacion)

        # Registrar las acciones en el historial
        _id_usuario = current_user.id
        _accion = "REGISTRO"
        _area = "FERRETERIA"

        for articulo, cantidad, unidad, area in zip(articulos, cantidades, unidades, areas):
            _detalles = f'AGREGADA {cantidad} {unidad} DE "{articulo}", ÁREA: {area}'
            registrar_historial(conn, _id_usuario, _accion, _area, _detalles)

        # Confirmar cambios
        inventario_manager.commit()
        flash('¡Entradas agregadas y stock actualizado exitosamente!', 'success')

    except Exception as e:
        # Revertir en caso de error
        inventario_manager.rollback()
        flash(f'Error al agregar entradas: {str(e)}', 'error')

    finally:
        # Cerrar conexión
        inventario_manager.cerrar_conexion()

    return redirect('/ferreteria/entradas')

@app.route('/ferreteria/editar_entrada/<string:id_entrada>')
def editar_entrada_ferreteria(id_entrada):

    conn = mysql.connect()
    cursor = conn.cursor()

    #Vista para el usuario director (o administrador)

    cursor.execute("SELECT * FROM articulos ORDER BY articulo")

    articulos = cursor.fetchall()

    cursor.execute("""
        SELECT
            aux_ent.id,
            aux_ent.id_entrada,
            art.articulo AS nombre_articulo,
            art.unidad,
            art.area,
            aux_ent.cantidad,
            ent.fecha
        FROM
            aux_entradas_inventario aux_ent
        JOIN
            articulos art ON aux_ent.id_articulo = art.id_articulo
        JOIN
            entradas_inventario ent ON aux_ent.id_entrada = ent.id_entrada
        WHERE
            aux_ent.id_entrada = %s
        """, (id_entrada,))

    entradas = cursor.fetchall()

    return render_template('ferreteria/editar_entrada.html', articulos = articulos, entradas = entradas)


@app.route('/ferreteria/editar_entrada_BD', methods=['POST'])
def ferreteria_editar_entrada_BD():
    try:
        # Obtener datos del formulario
        id_entrada = request.form['txtID']
        fecha = request.form['txtFecha']
        articulos = [articulo.upper() for articulo in request.form.getlist('articulo[]')]
        cantidades = request.form.getlist('cantidad[]')
        unidades = [unidad.upper() for unidad in request.form.getlist('unidad[]')]
        areas = [area.upper() for area in request.form.getlist('area[]')]

        # Conectar a la base de datos
        conn = mysql.connect()
        inventario_manager = InventarioManager(conn)

        # Editar la entrada y actualizar el stock
        inventario_manager.editar_entrada(id_entrada, fecha, articulos, cantidades, unidades, areas)

        # Registrar la acción en el historial
        _id_usuario = current_user.id
        _accion = "EDICIÓN"
        _area = "FERRETERÍA"
        for articulo, cantidad, unidad, area in zip(articulos, cantidades, unidades, areas):
            detalles = f"SE ACTUALIZÓ {cantidad} {unidad} DE {articulo} EN LA ENTRADA {id_entrada} EL {fecha}"
            registrar_historial(conn, _id_usuario, _accion, _area, detalles)

        # Confirmar cambios
        inventario_manager.commit()
        flash('¡Entrada actualizada exitosamente!', 'success')

    except Exception as e:
        # Revertir cambios en caso de error
        inventario_manager.rollback()
        flash(f'Error al actualizar la entrada: {str(e)}', 'error')

    finally:
        # Cerrar conexión
        inventario_manager.cerrar_conexion()

    return redirect('/ferreteria/entradas')

@app.route('/eliminar_multiples_entradas', methods=['POST'])
@login_required
def eliminar_multiples_entradas():
    ids = request.json.get('ids')

    if not ids:
        return jsonify({'success': False, 'message': 'No se proporcionaron IDs'})

    conn = mysql.connect()
    cursor = conn.cursor()

    try:
        entradas_a_verificar = set()
        articulos_cache = {}

        for id_entrada in ids:
            # Obtener todos los registros asociados al id_entrada en aux_entradas_inventario
            cursor.execute('''
                            SELECT id_articulo, cantidad
                            FROM aux_entradas_inventario
                            WHERE id_entrada=%s
                           ''', (id_entrada,))
            resultados = cursor.fetchall()

            if not resultados:
                continue

            for id_articulo, cantidad in resultados:
                # Obtener nombre del artículo desde la caché o la base de datos si no está en la caché
                if id_articulo in articulos_cache:
                    nombre_articulo = articulos_cache[id_articulo]
                else:
                    cursor.execute('''
                                    SELECT articulo
                                    FROM articulos
                                    WHERE id_articulo=%s
                                   ''', (id_articulo,))
                    nombre_articulo = cursor.fetchone()[0]
                    articulos_cache[id_articulo] = nombre_articulo

                # Actualizar la tabla stock_inventario disminuyendo la cantidad eliminada
                cursor.execute('''
                                UPDATE stock_inventario
                                SET cantidad = GREATEST(cantidad - %s, 0)
                                WHERE id_articulo=%s
                               ''', (cantidad, id_articulo))

                # Registrar la eliminación de cada artículo en el historial
                detalles = f'SE ELIMINÓ {cantidad} DEL ARTÍCULO: "{nombre_articulo}" EN LA SALIDA: {id_entrada}.'
                registrar_historial(conn, current_user.id, "ELIMINACIÓN", "ARTICULOS", detalles)

            # Eliminar todos los registros de aux_entradas_inventario asociados a esta salida
            cursor.execute('''
                            DELETE FROM aux_entradas_inventario
                            WHERE id_entrada=%s
                           ''', (id_entrada,))

            # Registrar la eliminación de la salida completa en el historial
            detalles = f"SE ELIMINÓ COMPLETAMENTE LA SALIDA: ID {id_entrada}."
            registrar_historial(conn, current_user.id, "ELIMINACIÓN", "FERRETERÍA", detalles)

            # Eliminar la salida completa de entradas_inventario
            cursor.execute('''
                            DELETE FROM entradas_inventario
                            WHERE id_entrada=%s
                           ''', (id_entrada,))

        # manager = InventarioManager(conn)
        # manager.actualizar_stock()
        conn.commit()
        return jsonify({'success': True})

    except Exception as e:
        conn.rollback()
        print(f"Error: {e}")
        return jsonify({'success': False, 'message': str(e)})

    finally:
        cursor.close()
        conn.close()

### Gestión de Salidas (Galpón) ###

@app.route('/ferreteria/salidas')
@login_required
def ferreteria_salidas():
    """Consultar el FrontEnd de las Salidas"""

    conn = mysql.connect()
    manager = InventarioManager(conn)  # Instanciar la clase InventarioManager

    # Obtener las salidas agrupadas
    salidas = manager.obtener_entradas_o_salidas(tipo="salidas", ubicacion="galpon")

    # Agrupar las salidas por ID de salida
    grouped_salidas = []
    current_id_salida = None
    group = []

    for salida in salidas:
        if salida[6] != current_id_salida:  # Index 6 corresponde a id_salida
            if group:
                grouped_salidas.append(group)
            group = [salida]
            current_id_salida = salida[6]
        else:
            group.append(salida)

    if group:
        grouped_salidas.append(group)

    # Calcular la cantidad total de artículos en las salidas
    cantidad_total = sum(salida[3] for salida in salidas)  # Index 3 corresponde a cantidad

    return render_template(
        '/ferreteria/salidas.html',
        grouped_salidas=grouped_salidas,
        cantidad_total=cantidad_total
    )


@app.route('/ferreteria/agregar_salida')
def agregar_salida_ferreteria():
    """Accede al Front de Agregar las Salida (Galpon)"""

    conn = mysql.connect()

    inventario_manager = InventarioManager(conn)

    # Definir la ubicación (puede ser "galpon" o "oficina")
    ubicacion = "galpon"

    articulos = inventario_manager.obtener_articulos_con_stock(ubicacion)

    # Cerrar la conexión
    inventario_manager.cerrar_conexion()

    return render_template('ferreteria/agregar_salida.html', articulos = articulos)

@app.route('/ferreteria/agregar_salida_BD', methods=['POST'])
@login_required
def ferreteria_agregar_salida_BD():
    """Agrega una salida a la base de Datos"""
    try:
        fecha = request.form['txtFecha']
        destino = request.form['txtDestino'].upper()
        articulos = [articulo.upper() for articulo in request.form.getlist('articulo[]')]
        cantidades = request.form.getlist('cantidad[]')
        unidades = [unidad.upper() for unidad in request.form.getlist('unidad[]')]
        areas = [area.upper() for area in request.form.getlist('area[]')]

        conn = mysql.connect()
        cursor = conn.cursor()

        # Obtener el último ID de salida
        cursor.execute("SELECT id_salida FROM salidas_inventario ORDER BY id_salida DESC LIMIT 1")
        last_id = cursor.fetchone()

        if last_id:
            last_id_num = int(last_id[0].split('-')[1])
            new_id_num = last_id_num + 1
        else:
            new_id_num = 1

        new_id_salida = f"SA-{new_id_num:04d}"

        # Insertar en la tabla de salidas_inventario con el nuevo ID
        cursor.execute("INSERT INTO salidas_inventario (id_salida, fecha, destino) VALUES (%s, %s, %s)", (new_id_salida, fecha, destino))

        _id_usuario = current_user.id
        _accion = "SALIDAS"
        detalles_historial = []

        for articulo, cantidad, unidad, area in zip(articulos, cantidades, unidades, areas):

            # Verificar si el artículo ya existe en la base de datos
            cursor.execute("SELECT id_articulo FROM articulos WHERE articulo = %s AND unidad = %s", (articulo, unidad))
            result = cursor.fetchone()

            if not result:
                # Si no existe el artículo, crearlo
                prefijo = area[:3]
                cursor.execute("SELECT id_articulo FROM articulos WHERE id_articulo LIKE %s ORDER BY id_articulo DESC LIMIT 1", (f"{prefijo}-%",))
                last_id_articulo = cursor.fetchone()

                if last_id_articulo:
                    last_number = int(last_id_articulo[0].split('-')[1])
                    new_number = last_number + 1
                else:
                    new_number = 1

                # Generar el nuevo código para el artículo
                id_articulo = f"{prefijo}-{new_number:03}"
                cursor.execute("INSERT INTO articulos (id_articulo, articulo, unidad, area) VALUES (%s, %s, %s, %s)", (id_articulo, articulo, unidad, area))



                # Obtener el último ID de entrada
                cursor.execute("SELECT id_entrada FROM entradas_inventario ORDER BY id_entrada DESC LIMIT 1")
                last_id = cursor.fetchone()

                if last_id:
                    last_id_num = int(last_id[0].split('-')[1])
                    new_id_num = last_id_num + 1
                else:
                    new_id_num = 1

                new_id_entrada = f"EN-{new_id_num:04d}"

                # Insertar en la tabla de entradas_inventario
                cursor.execute("INSERT INTO entradas_inventario (id_entrada, fecha) VALUES (%s, %s)", (new_id_entrada, fecha))

                # Insertar en la tabla aux_entradas_inventario usando el correcto id_entrada
                cursor.execute("INSERT INTO aux_entradas_inventario (id_entrada, id_articulo, cantidad) VALUES (%s, %s, %s)", (new_id_entrada, id_articulo, cantidad))

                # Insertar el articulo en la tabla Stock Ferreteria
                cursor.execute("INSERT INTO stock_inventario (id_articulo, cantidad) VALUES (%s, %s)", (id_articulo, cantidad))

            else: #El articulo sí existe.
                id_articulo = result[0]

            # Insertar en la tabla aux_salidas_inventario
            cursor.execute("INSERT INTO aux_salidas_inventario (id_salida, id_articulo, cantidad) VALUES (%s, %s, %s)", (new_id_salida, id_articulo, cantidad))

            # Actualizar la tabla de stock_inventario
            cursor.execute("SELECT cantidad FROM stock_inventario WHERE id_articulo = %s", (id_articulo,))
            stock_result = cursor.fetchone()

            nueva_cantidad = stock_result[0] - int(cantidad)

            if nueva_cantidad >= 0:
                cursor.execute("UPDATE stock_inventario SET cantidad = %s WHERE id_articulo = %s", (nueva_cantidad, id_articulo))
            else:
                raise ValueError("La cantidad de stock no puede ser negativa.")

            # Agregar detalles al historial
            detalle = f'ENTREGADO {cantidad} {unidad} DE "{articulo}", ÁREA: {area}'
            detalles_historial.append(detalle)

        conn.commit()

        if detalles_historial:
            _detalles = "\n".join(detalles_historial)
            registrar_historial(conn, _id_usuario, _accion, "FERRETERIA", _detalles)

        conn.rollback()

        flash('¡Salidas agregadas y stock actualizado exitosamente!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Error al agregar salidas: {str(e)}', 'error')
    finally:
        cursor.close()
        conn.close()

    return redirect('/ferreteria/salidas')

@app.route('/ferreteria/editar_salida/<string:id_salida>')
def editar_salida_ferreteria(id_salida):
    """Accede al Front de Editar salida"""

    conn = mysql.connect()
    cursor = conn.cursor()

    #Vista para el usuario director (o administrador)

    cursor.execute("""
        SELECT a.id_articulo,
                a.articulo,
                a.unidad,
                a.area,
               IFNULL(SUM(s.cantidad), 0) AS total_cantidad
        FROM articulos a
        LEFT JOIN stock_inventario s ON a.id_articulo = s.id_articulo
        GROUP BY a.id_articulo, a.articulo, a.unidad, a.area
        ORDER BY a.articulo
    """)
    articulos = cursor.fetchall()

    cursor.execute("""
        SELECT
            aux_ent.id,
            aux_ent.id_salida,
            art.articulo AS nombre_articulo,
            art.unidad,
            art.area,
            aux_ent.cantidad,
            ent.fecha,
            ent.destino
        FROM
            aux_salidas_inventario aux_ent
        JOIN
            articulos art ON aux_ent.id_articulo = art.id_articulo
        JOIN
            salidas_inventario ent ON aux_ent.id_salida = ent.id_salida
        WHERE
            aux_ent.id_salida = %s
        """, (id_salida,))

    salidas = cursor.fetchall()

    return render_template('ferreteria/editar_salida.html', articulos = articulos, salidas = salidas)

@app.route('/ferreteria/editar_salida_BD', methods=['POST'])
def editar_salida_BD():
    """Guarda la modificacion en la base de datos"""
    try:
        id_salida = request.form['txtID']
        fecha = request.form['txtFecha']
        destino = request.form['txtDestino'].upper()
        articulos = [articulo.upper() for articulo in request.form.getlist('articulo[]')]
        cantidades = request.form.getlist('cantidad[]')
        unidades = [unidad.upper() for unidad in request.form.getlist('unidad[]')]
        areas = [area.upper() for area in request.form.getlist('area[]')]

        conn = mysql.connect()
        cursor = conn.cursor()

        # Actualizar la fecha en la tabla de salidas_inventario
        cursor.execute("UPDATE salidas_inventario SET fecha = %s, destino = %s WHERE id_salida = %s", (fecha, destino, id_salida))

        # Obtener los artículos existentes en la salida
        cursor.execute("SELECT id_articulo, cantidad FROM aux_salidas_inventario WHERE id_salida = %s", (id_salida,))
        existing_entries = cursor.fetchall()
        existing_articles = {row[0]: row[1] for row in existing_entries}

        new_articles = set()
        for articulo, cantidad, unidad, area in zip(articulos, cantidades, unidades, areas):

            # Obtener el ID del artículo directamente
            cursor.execute("SELECT id_articulo FROM articulos WHERE articulo = %s AND unidad = %s", (articulo, unidad))
            result = cursor.fetchone()

            if result:
                id_articulo = result[0]
                new_articles.add(id_articulo)
            else:
                flash(f"El artículo '{articulo}' con unidad '{unidad}' no existe en la base de datos.", 'warning')
                continue

            cantidad = int(cantidad)

            # Verificar si el artículo ya existe en la salida
            if id_articulo in existing_articles:
                old_cantidad = existing_articles[id_articulo]
                diferencia = cantidad - old_cantidad

                # Actualizar la cantidad en aux_salidas_inventario
                cursor.execute("UPDATE aux_salidas_inventario SET cantidad = %s WHERE id_salida = %s AND id_articulo = %s", (cantidad, id_salida, id_articulo))

                # Actualizar la tabla stock_inventario
                #cursor.execute("UPDATE stock_inventario SET cantidad = cantidad - %s WHERE id_articulo = %s", (diferencia, id_articulo))
            else:
                # Si no existe, insertarlo en aux_salidas_inventario
                cursor.execute("INSERT INTO aux_salidas_inventario (id_salida, id_articulo, cantidad) VALUES (%s, %s, %s)", (id_salida, id_articulo, cantidad))

                # Actualizar el stock aumentando la cantidad
                #cursor.execute("UPDATE stock_inventario SET cantidad = cantidad - %s WHERE id_articulo = %s", (cantidad, id_articulo))

        # Eliminar artículos que ya no están en la lista
        articles_to_remove = set(existing_articles.keys()) - new_articles
        for art_id in articles_to_remove:
            old_cantidad = existing_articles[art_id]

            # Aumentar el stock por la cantidad eliminada
            #cursor.execute("UPDATE stock_inventario SET cantidad = cantidad + %s WHERE id_articulo = %s", (old_cantidad, art_id))

            # Eliminar el artículo de aux_salidas_inventario
            cursor.execute("DELETE FROM aux_salidas_inventario WHERE id_salida = %s AND id_articulo = %s", (id_salida, art_id))

        conn.commit()

        # Registrar la acción en el historial
        for articulo, cantidad, unidad, area in zip(articulos, cantidades, unidades, areas):
            detalles = f"SE ACTUALIZÓ {cantidad} {unidad} DE {articulo} EN LA SALIDA {id_salida} EL {fecha}"
            registrar_historial(conn, current_user.id, "EDICIÓN", "FERRETERÍA", detalles)

        flash('¡Salida actualizada exitosamente!', 'success')

        Actualizar_stock_inventario()

    except Exception as e:
        conn.rollback()
        flash(f'Error al actualizar la salida: {str(e)}', 'error')
    finally:
        cursor.close()
        conn.close()

    return redirect('/ferreteria/salidas')




def Actualizar_stock_inventario():
    try:
        conn = mysql.connect()
        inventario_manager = InventarioManager(conn)

        # Obtener los datos
        todos_los_articulos = inventario_manager.obtener_ID_articulos()
        entradas = inventario_manager.obtener_totales("aux_entradas_inventario")
        salidas = inventario_manager.obtener_totales("aux_salidas_inventario")

        # Actualizar el stock de cada artículo
        for id_articulo in todos_los_articulos:
            total_entrada = entradas.get(id_articulo, 0)
            total_salida = salidas.get(id_articulo, 0)
            nuevo_stock = total_entrada - total_salida

            # Si no hay movimientos, establecer stock en 0
            if total_entrada == 0 and total_salida == 0:
                nuevo_stock = 0

            # Actualizar el stock en la base de datos
            inventario_manager.actualizar_stock(id_articulo, nuevo_stock)

        # Confirmar cambios
        inventario_manager.commit()
        flash('¡Stock de inventario actualizado exitosamente!', 'success')
    except Exception as e:
        # Revertir cambios en caso de error
        inventario_manager.rollback()
        flash(f'Error al actualizar el stock: {str(e)}', 'error')
    finally:
        # Cerrar la conexión
        inventario_manager.cerrar_conexion()









# @app.route('/ferreteria/salidas')
# @login_required
# def ferreteria_salidas():
#     # Establecer conexión con la base de datos
#     conn = mysql.connect()
#     cursor = conn.cursor()

#     # Ejecutar la consulta para obtener todas las salidas y sus detalles
#     cursor.execute('''
#         SELECT
#             aux_sal.id,
#             DATE_FORMAT(salidas_inventario.fecha, '%d/%m/%Y') as fecha,
#             art.articulo,
#             aux_sal.cantidad,
#             art.unidad,
#             art.area,
#             salidas_inventario.id_salida,
#             salidas_inventario.destino
#         FROM
#             salidas_inventario
#         JOIN
#             aux_salidas_inventario aux_sal ON salidas_inventario.id_salida = aux_sal.id_salida
#         JOIN
#             articulos art ON aux_sal.id_articulo = art.id_articulo
#         ORDER BY
#             salidas_inventario.fecha ASC, salidas_inventario.id_salida ASC
#     ''')

#     # Obtener los resultados de la consulta
#     salidas = cursor.fetchall()

#     # Agrupar las salidas por id_salida
#     grouped_salidas = {}
#     for salida in salidas:
#         id, fecha, articulo, cantidad, unidad, area, id_salida, destino = salida

#         # Convertir cantidad a entero si es un valor entero
#         if cantidad.is_integer():
#             cantidad = int(cantidad)

#         # Si id_salida no está en el diccionario, inicializar una nueva entrada
#         if id_salida not in grouped_salidas:
#             grouped_salidas[id_salida] = {
#                 'fecha': fecha,
#                 'destino': destino,
#                 'articulos': []
#             }

#         # Agregar los detalles del artículo al grupo correspondiente
#         grouped_salidas[id_salida]['articulos'].append({
#             'id': id,
#             'articulo': articulo,
#             'cantidad': cantidad,
#             'unidad': unidad,
#             'area': area
#         })

#     # Calcular la cantidad total de artículos ingresados
#     cursor.execute('''
#         SELECT
#             SUM(ae.cantidad) as cantidad_total
#         FROM
#             aux_salidas_inventario ae
#     ''')
#     cantidad_total = cursor.fetchone()[0]

#     # Cerrar el cursor y la conexión a la base de datos
#     cursor.close()
#     conn.close()

#     # print("Las salidas agrupadas son: ", grouped_salidas)

#     # Renderizar la plantilla, pasando las salidas agrupadas y la cantidad total
#     return render_template('/ferreteria/salidas.html', salidas=grouped_salidas, cantidad_total=cantidad_total)




























@app.route('/eliminar_multiples_salidas', methods=['POST'])
@login_required
def eliminar_multiples_salidas():
    ids = request.json.get('ids')

    if not ids:
        return jsonify({'success': False, 'message': 'No se proporcionaron IDs'})

    conn = mysql.connect()
    cursor = conn.cursor()

    try:
        salidas_a_verificar = set()
        articulos_cache = {}

        for id_salida in ids:
            # Eliminar todos los registros asociados al id_salida en aux_salidas_inventario
            cursor.execute('''
                            SELECT id_articulo, cantidad
                            FROM aux_salidas_inventario
                            WHERE id_salida=%s
                           ''', (id_salida,))
            resultados = cursor.fetchall()

            for id_articulo, cantidad in resultados:
                if id_articulo in articulos_cache:
                    nombre_articulo = articulos_cache[id_articulo]
                else:
                    cursor.execute('''
                                    SELECT articulo
                                    FROM articulos
                                    WHERE id_articulo=%s
                                   ''', (id_articulo,))
                    nombre_articulo = cursor.fetchone()[0]
                    articulos_cache[id_articulo] = nombre_articulo

                cursor.execute('''
                                UPDATE stock_inventario
                                SET cantidad = GREATEST(cantidad + %s, 0)
                                WHERE id_articulo=%s
                               ''', (cantidad, id_articulo))

                detalles = f'SE ELIMINÓ {cantidad} DEL ARTÍCULO: "{nombre_articulo}" EN LA SALIDA: {id_salida}.'
                registrar_historial(conn, current_user.id, "ELIMINACIÓN", "FERRETERÍA", detalles)

            cursor.execute('''
                            DELETE FROM aux_salidas_inventario
                            WHERE id_salida=%s
                           ''', (id_salida,))

            # Eliminar todos los registros asociados al id_salida en articulos_extra_salida
            cursor.execute('''
                            SELECT articulo, cantidad
                            FROM articulos_extra_salida
                            WHERE id_salida=%s
                           ''', (id_salida,))
            resultados_extras = cursor.fetchall()

            # for articulo_extra, cantidad_extra in resultados_extras:
            #     cursor.execute('''
            #                     UPDATE stock_inventario
            #                     SET cantidad = GREATEST(cantidad + %s, 0)
            #                     WHERE articulo=%s
            #                    ''', (cantidad_extra, articulo_extra))

            #     detalles_extra = f'SE ELIMINÓ {cantidad_extra} DEL ARTÍCULO EXTRA: "{articulo_extra}" EN LA SALIDA: {id_salida}.'
            #     registrar_historial(conn, current_user.id, "ELIMINACIÓN", "FERRETERÍA", detalles_extra)

            cursor.execute('''
                            DELETE FROM articulos_extra_salida
                            WHERE id_salida=%s
                           ''', (id_salida,))

            detalles = f"SE ELIMINÓ COMPLETAMENTE LA SALIDA: ID {id_salida}."
            registrar_historial(conn, current_user.id, "ELIMINACIÓN", "FERRETERÍA", detalles)

            cursor.execute('''
                            DELETE FROM salidas_inventario
                            WHERE id_salida=%s
                           ''', (id_salida,))

        conn.commit()
        return jsonify({'success': True})

    except Exception as e:
        conn.rollback()
        print(f"Error: {e}")
        return jsonify({'success': False, 'message': str(e)})

    finally:
        cursor.close()
        conn.close()



# @app.route('/ferreteria/eliminar_salida/<string:id_salida>')
# @login_required
# def eliminar_salida(id_salida):

#     conn = mysql.connect()
#     cursor = conn.cursor()

#     try:
#         cursor.execute("DELETE FROM salidas_inventario WHERE id_salida = %s", (id_salida))


#     except Exception as e:
#         conn.rollback()
#         print(f"Error: {e}")

#     finally:
#         conn.commit()
#         conn.close()
#         cursor.close()

#     flash("¡La salida: " + id_salida + " ha sido eliminada exitosamente!")
#     return redirect('/ferreteria/salidas')



mode = "dev"

if __name__ == '__main__':

    if mode == "dev":
        app.run(host='0.0.0.0', port=5000, debug=True)
    else:
        serve(app,host='0.0.0.0',port=5000,threads=6)
        serve(app,host='0.0.0.0',port=5000,threads=6)

